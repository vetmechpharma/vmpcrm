from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File, Request
from fastapi.responses import Response
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import uuid
import json
import os

from deps import db, logger, get_current_user, hash_password, verify_password, create_token, security
from models.schemas import ItemCreate, ItemUpdate, ItemResponse, CategoryResponse, CustomField
from utils.image import process_image_to_webp
from utils.code_gen import generate_item_code
import base64
from io import BytesIO
from PIL import Image
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

router = APIRouter(prefix="/api")

# ============== ITEM ROUTES ==============

@router.get("/item-categories", response_model=List[CategoryResponse])
async def get_item_categories(current_user: dict = Depends(get_current_user)):
    """Get all unique item categories with counts"""
    pipeline = [
        {'$match': {'category': {'$ne': None, '$ne': '', '$exists': True}}},
        {'$group': {'_id': '$category', 'count': {'$sum': 1}}},
        {'$match': {'_id': {'$ne': None}}},
        {'$sort': {'_id': 1}}
    ]
    categories = await db.items.aggregate(pipeline).to_list(100)
    return [CategoryResponse(name=cat['_id'], count=cat['count']) for cat in categories if cat['_id']]

@router.post("/items", response_model=ItemResponse)
async def create_item(item_data: ItemCreate, current_user: dict = Depends(get_current_user)):
    item_id = str(uuid.uuid4())
    
    # Use custom item_code if provided, otherwise auto-generate
    if item_data.item_code and item_data.item_code.strip():
        item_code = item_data.item_code.strip()
        # Check if code already exists
        existing = await db.items.find_one({'item_code': item_code}, {'_id': 0})
        if existing:
            raise HTTPException(status_code=400, detail="Item code already exists")
    else:
        item_code = await generate_item_code()
    
    now = datetime.now(timezone.utc)
    
    # Process image if provided
    processed_image = None
    if item_data.image_base64:
        try:
            # Decode base64 and process
            image_bytes = base64.b64decode(item_data.image_base64)
            processed_image = process_image_to_webp(image_bytes)
        except Exception as e:
            logger.error(f"Image processing error: {str(e)}")
    
    # Set default role-based pricing from rate if not provided
    rate = item_data.rate or 0
    
    item_doc = {
        'id': item_id,
        'item_code': item_code,
        'item_name': item_data.item_name,
        'main_categories': item_data.main_categories or [],
        'subcategories': item_data.subcategories or [],
        'composition': item_data.composition,
        'mrp': item_data.mrp,
        'gst': item_data.gst,
        # Role-based pricing - default to rate if not provided
        'rate_doctors': item_data.rate_doctors if item_data.rate_doctors is not None else rate,
        'offer_doctors': item_data.offer_doctors or item_data.offer,
        'special_offer_doctors': item_data.special_offer_doctors or item_data.special_offer,
        'rate_medicals': item_data.rate_medicals if item_data.rate_medicals is not None else rate,
        'offer_medicals': item_data.offer_medicals or item_data.offer,
        'special_offer_medicals': item_data.special_offer_medicals or item_data.special_offer,
        'rate_agencies': item_data.rate_agencies if item_data.rate_agencies is not None else rate,
        'offer_agencies': item_data.offer_agencies or item_data.offer,
        'special_offer_agencies': item_data.special_offer_agencies or item_data.special_offer,
        # Special Offer 2 - for dashboard scroll
        'special_offer_2_doctors': item_data.special_offer_2_doctors,
        'special_offer_2_doctors_desc': item_data.special_offer_2_doctors_desc,
        'special_offer_2_medicals': item_data.special_offer_2_medicals,
        'special_offer_2_medicals_desc': item_data.special_offer_2_medicals_desc,
        'special_offer_2_agencies': item_data.special_offer_2_agencies,
        'special_offer_2_agencies_desc': item_data.special_offer_2_agencies_desc,
        # Legacy fields
        'rate': rate,
        'offer': item_data.offer,
        'special_offer': item_data.special_offer,
        'custom_fields': [cf.model_dump() for cf in (item_data.custom_fields or [])],
        'image_webp': processed_image,
        'has_image': processed_image is not None,
        'out_of_stock': item_data.out_of_stock or False,
        'is_hidden': item_data.is_hidden or False,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': current_user['id']
    }
    
    await db.items.insert_one(item_doc)
    
    return ItemResponse(
        id=item_id,
        item_code=item_code,
        item_name=item_data.item_name,
        main_categories=item_data.main_categories or [],
        subcategories=item_data.subcategories or [],
        composition=item_data.composition,
        mrp=item_data.mrp,
        gst=item_data.gst,
        rate_doctors=item_doc['rate_doctors'],
        offer_doctors=item_doc['offer_doctors'],
        special_offer_doctors=item_doc['special_offer_doctors'],
        rate_medicals=item_doc['rate_medicals'],
        offer_medicals=item_doc['offer_medicals'],
        special_offer_medicals=item_doc['special_offer_medicals'],
        rate_agencies=item_doc['rate_agencies'],
        offer_agencies=item_doc['offer_agencies'],
        special_offer_agencies=item_doc['special_offer_agencies'],
        # Special Offer 2 - for dashboard scroll
        special_offer_2_doctors=item_data.special_offer_2_doctors,
        special_offer_2_doctors_desc=item_data.special_offer_2_doctors_desc,
        special_offer_2_medicals=item_data.special_offer_2_medicals,
        special_offer_2_medicals_desc=item_data.special_offer_2_medicals_desc,
        special_offer_2_agencies=item_data.special_offer_2_agencies,
        special_offer_2_agencies_desc=item_data.special_offer_2_agencies_desc,
        rate=rate,
        offer=item_data.offer,
        special_offer=item_data.special_offer,
        custom_fields=item_data.custom_fields or [],
        image_url=f"/api/items/{item_id}/image" if processed_image else None,
        out_of_stock=item_data.out_of_stock or False,
        is_hidden=item_data.is_hidden or False,
        created_at=now
    )

@router.get("/items/{item_id}/image")
async def get_item_image(item_id: str, fmt: Optional[str] = None):
    """Get item image as WebP or JPEG (use ?fmt=jpg for JPEG)"""
    item = await db.items.find_one({'id': item_id}, {'image_webp': 1})
    if not item or not item.get('image_webp'):
        raise HTTPException(status_code=404, detail="Image not found")
    
    image_data = base64.b64decode(item['image_webp'])
    
    if fmt == 'jpg' or fmt == 'jpeg':
        img = Image.open(BytesIO(image_data))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        jpg_buffer = BytesIO()
        img.save(jpg_buffer, format='JPEG', quality=85)
        jpg_buffer.seek(0)
        return Response(content=jpg_buffer.read(), media_type="image/jpeg")
    
    return Response(content=image_data, media_type="image/webp")


@router.get("/items/{item_id}/image.jpg")
async def get_item_image_jpg(item_id: str):
    """Get item image as JPEG with proper .jpg extension for WhatsApp compatibility"""
    return await get_item_image(item_id, fmt='jpg')


@router.get("/items", response_model=List[ItemResponse])
async def get_items(
    search: Optional[str] = None,
    main_category: Optional[str] = None,
    subcategory: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if search:
        query['$or'] = [
            {'item_name': {'$regex': search, '$options': 'i'}},
            {'item_code': {'$regex': search, '$options': 'i'}},
            {'composition': {'$regex': search, '$options': 'i'}}
        ]
    
    if main_category:
        query['main_categories'] = main_category
    
    if subcategory:
        query['subcategories'] = subcategory
    
    items = await db.items.find(query, {'_id': 0, 'image_webp': 0}).sort('created_at', -1).to_list(1000)
    
    result = []
    for item in items:
        created_at = item.get('created_at', '')
        if created_at and isinstance(created_at, str):
            try:
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            except (ValueError, TypeError):
                created_at = datetime.now(timezone.utc)
        elif not created_at:
            created_at = datetime.now(timezone.utc)
        
        custom_fields = [CustomField(**cf) for cf in item.get('custom_fields', [])]
        
        # Check if image exists
        has_image = await db.items.find_one({'id': item['id'], 'image_webp': {'$ne': None}}, {'_id': 1})
        
        result.append(ItemResponse(
            id=item['id'],
            item_code=item['item_code'],
            item_name=item['item_name'],
            main_categories=item.get('main_categories', []) or item.get('main_category', []) if isinstance(item.get('main_category'), list) else ([item.get('main_category')] if item.get('main_category') else []),
            subcategories=item.get('subcategories', []),
            composition=item.get('composition'),
            mrp=item.get('mrp', 0),
            gst=item.get('gst', 0),
            # Role-based pricing
            rate_doctors=item.get('rate_doctors') or item.get('rate', 0),
            offer_doctors=item.get('offer_doctors') or item.get('offer'),
            special_offer_doctors=item.get('special_offer_doctors') or item.get('special_offer'),
            rate_medicals=item.get('rate_medicals') or item.get('rate', 0),
            offer_medicals=item.get('offer_medicals') or item.get('offer'),
            special_offer_medicals=item.get('special_offer_medicals') or item.get('special_offer'),
            rate_agencies=item.get('rate_agencies') or item.get('rate', 0),
            offer_agencies=item.get('offer_agencies') or item.get('offer'),
            special_offer_agencies=item.get('special_offer_agencies') or item.get('special_offer'),
            # Special Offer 2 - for dashboard scroll
            special_offer_2_doctors=item.get('special_offer_2_doctors'),
            special_offer_2_doctors_desc=item.get('special_offer_2_doctors_desc'),
            special_offer_2_medicals=item.get('special_offer_2_medicals'),
            special_offer_2_medicals_desc=item.get('special_offer_2_medicals_desc'),
            special_offer_2_agencies=item.get('special_offer_2_agencies'),
            special_offer_2_agencies_desc=item.get('special_offer_2_agencies_desc'),
            # Legacy
            rate=item.get('rate', 0),
            offer=item.get('offer'),
            special_offer=item.get('special_offer'),
            custom_fields=custom_fields,
            image_url=f"/api/items/{item['id']}/image" if has_image else None,
            out_of_stock=item.get('out_of_stock', False),
            is_hidden=item.get('is_hidden', False),
            created_at=created_at
        ))
    
    return result

@router.get("/items/offers/active")
async def get_active_offers(role: Optional[str] = None):
    """Get items with active Special Offer 2 - for dashboard scrolling"""
    items = await db.items.find({'out_of_stock': {'$ne': True}, 'is_hidden': {'$ne': True}}, {
        '_id': 0, 'image_webp': 0, 'custom_fields': 0
    }).to_list(500)
    
    offers = []
    for item in items:
        so2_key = f'special_offer_2_{role}s' if role else 'special_offer_2_doctors'
        desc_key = f'special_offer_2_{role}s_desc' if role else 'special_offer_2_doctors_desc'
        rate_key = f'rate_{role}s' if role else 'rate'
        
        offer_text = item.get(so2_key) or ''
        desc_text = item.get(desc_key) or ''
        
        if offer_text:
            has_image = await db.items.find_one({'id': item['id'], 'image_webp': {'$ne': None}}, {'_id': 1})
            offers.append({
                'id': item['id'],
                'item_name': item.get('item_name', ''),
                'item_code': item.get('item_code', ''),
                'mrp': item.get('mrp', 0),
                'rate': item.get(rate_key) or item.get('rate', 0),
                'offer': offer_text,
                'description': desc_text,
                'image_url': f"/api/items/{item['id']}/image" if has_image else None,
            })
    return offers

@router.get("/items/{item_id}", response_model=ItemResponse)
async def get_item(item_id: str, current_user: dict = Depends(get_current_user)):
    item = await db.items.find_one({'id': item_id}, {'_id': 0, 'image_webp': 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    created_at = item.get('created_at', '')
    if created_at and isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        except (ValueError, TypeError):
            created_at = datetime.now(timezone.utc)
    elif not created_at:
        created_at = datetime.now(timezone.utc)
    
    custom_fields = [CustomField(**cf) for cf in item.get('custom_fields', [])]
    
    # Check if image exists
    has_image = await db.items.find_one({'id': item_id, 'image_webp': {'$ne': None}}, {'_id': 1})
    
    return ItemResponse(
        id=item['id'],
        item_code=item['item_code'],
        item_name=item['item_name'],
        main_categories=item.get('main_categories', []) or ([item.get('main_category')] if item.get('main_category') else []),
        subcategories=item.get('subcategories', []),
        composition=item.get('composition'),
        offer=item.get('offer'),
        special_offer=item.get('special_offer'),
        mrp=item.get('mrp', 0),
        rate=item['rate'],
        gst=item.get('gst', 0),
        # Role-based pricing - Doctors
        rate_doctors=item.get('rate_doctors'),
        offer_doctors=item.get('offer_doctors'),
        special_offer_doctors=item.get('special_offer_doctors'),
        # Role-based pricing - Medicals
        rate_medicals=item.get('rate_medicals'),
        offer_medicals=item.get('offer_medicals'),
        special_offer_medicals=item.get('special_offer_medicals'),
        # Role-based pricing - Agencies
        rate_agencies=item.get('rate_agencies'),
        offer_agencies=item.get('offer_agencies'),
        special_offer_agencies=item.get('special_offer_agencies'),
        # Special Offer 2 - for dashboard scroll
        special_offer_2_doctors=item.get('special_offer_2_doctors'),
        special_offer_2_doctors_desc=item.get('special_offer_2_doctors_desc'),
        special_offer_2_medicals=item.get('special_offer_2_medicals'),
        special_offer_2_medicals_desc=item.get('special_offer_2_medicals_desc'),
        special_offer_2_agencies=item.get('special_offer_2_agencies'),
        special_offer_2_agencies_desc=item.get('special_offer_2_agencies_desc'),
        custom_fields=custom_fields,
        image_url=f"/api/items/{item['id']}/image" if has_image else None,
        out_of_stock=item.get('out_of_stock', False),
        is_hidden=item.get('is_hidden', False),
        created_at=created_at
    )

@router.put("/items/{item_id}", response_model=ItemResponse)
async def update_item(item_id: str, item_data: ItemUpdate, current_user: dict = Depends(get_current_user)):
    item = await db.items.find_one({'id': item_id}, {'_id': 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    update_data = {}
    for k, v in item_data.model_dump().items():
        if v is not None:
            if k == 'custom_fields':
                update_data[k] = [cf.model_dump() if hasattr(cf, 'model_dump') else cf for cf in v]
            elif k == 'image_base64':
                # Process and update image
                try:
                    image_bytes = base64.b64decode(v)
                    update_data['image_webp'] = process_image_to_webp(image_bytes)
                    update_data['has_image'] = True
                except Exception as e:
                    logger.error(f"Image processing error: {str(e)}")
            elif k == 'item_code':
                # Check if new code already exists (excluding current item)
                if v != item.get('item_code'):
                    existing = await db.items.find_one({'item_code': v, 'id': {'$ne': item_id}}, {'_id': 0})
                    if existing:
                        raise HTTPException(status_code=400, detail="Item code already exists")
                update_data[k] = v
            else:
                update_data[k] = v
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.items.update_one({'id': item_id}, {'$set': update_data})
    
    updated_item = await db.items.find_one({'id': item_id}, {'_id': 0, 'image_webp': 0})
    
    created_at = updated_item.get('created_at', '')
    if created_at and isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        except (ValueError, TypeError):
            created_at = datetime.now(timezone.utc)
    elif not created_at:
        created_at = datetime.now(timezone.utc)
    
    custom_fields = [CustomField(**cf) for cf in updated_item.get('custom_fields', [])]
    
    # Check if image exists
    has_image = await db.items.find_one({'id': item_id, 'image_webp': {'$ne': None}}, {'_id': 1})
    
    return ItemResponse(
        id=updated_item['id'],
        item_code=updated_item['item_code'],
        item_name=updated_item['item_name'],
        main_categories=updated_item.get('main_categories', []) or ([updated_item.get('main_category')] if updated_item.get('main_category') else []),
        subcategories=updated_item.get('subcategories', []),
        composition=updated_item.get('composition'),
        offer=updated_item.get('offer'),
        special_offer=updated_item.get('special_offer'),
        mrp=updated_item.get('mrp', 0),
        rate=updated_item['rate'],
        gst=updated_item.get('gst', 0),
        # Role-based pricing - Doctors
        rate_doctors=updated_item.get('rate_doctors'),
        offer_doctors=updated_item.get('offer_doctors'),
        special_offer_doctors=updated_item.get('special_offer_doctors'),
        # Role-based pricing - Medicals
        rate_medicals=updated_item.get('rate_medicals'),
        offer_medicals=updated_item.get('offer_medicals'),
        special_offer_medicals=updated_item.get('special_offer_medicals'),
        # Role-based pricing - Agencies
        rate_agencies=updated_item.get('rate_agencies'),
        offer_agencies=updated_item.get('offer_agencies'),
        special_offer_agencies=updated_item.get('special_offer_agencies'),
        # Special Offer 2 - for dashboard scroll
        special_offer_2_doctors=updated_item.get('special_offer_2_doctors'),
        special_offer_2_doctors_desc=updated_item.get('special_offer_2_doctors_desc'),
        special_offer_2_medicals=updated_item.get('special_offer_2_medicals'),
        special_offer_2_medicals_desc=updated_item.get('special_offer_2_medicals_desc'),
        special_offer_2_agencies=updated_item.get('special_offer_2_agencies'),
        special_offer_2_agencies_desc=updated_item.get('special_offer_2_agencies_desc'),
        custom_fields=custom_fields,
        image_url=f"/api/items/{item_id}/image" if has_image else None,
        out_of_stock=updated_item.get('out_of_stock', False),
        is_hidden=updated_item.get('is_hidden', False),
        created_at=created_at
    )

@router.delete("/items/{item_id}/image")
async def delete_item_image(item_id: str, current_user: dict = Depends(get_current_user)):
    """Delete item image"""
    result = await db.items.update_one({'id': item_id}, {'$set': {'image_webp': None}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Image deleted successfully"}

@router.delete("/items/{item_id}")
async def delete_item(item_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.items.delete_one({'id': item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted successfully"}

@router.patch("/items/{item_id}/stock")
async def toggle_item_stock(item_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Toggle out_of_stock status for an item"""
    out_of_stock = data.get('out_of_stock', False)
    result = await db.items.update_one({'id': item_id}, {'$set': {'out_of_stock': out_of_stock}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Stock status updated", "out_of_stock": out_of_stock}

@router.patch("/items/{item_id}/visibility")
async def toggle_item_visibility(item_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Toggle is_hidden status for an item (global hide)"""
    is_hidden = data.get('is_hidden', False)
    result = await db.items.update_one({'id': item_id}, {'$set': {'is_hidden': is_hidden}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Visibility updated", "is_hidden": is_hidden}

# ============== SUBCATEGORY ORDER & ITEM EXPORT ==============

DEFAULT_SUBCATEGORY_ORDER = [
    'Injection', 'Dry Injections', 'Hormones', 'Schedule X Drugs',
    'Liquids', 'Bolus', 'Powder', 'Feed Supplements',
    'Shampoo / Soap', 'Spray / Ointments', 'Tablets', 'Syrups', 'Vaccines'
]

@router.get("/subcategory-order")
async def get_subcategory_order(current_user: dict = Depends(get_current_user)):
    """Get subcategory display order"""
    doc = await db.subcategory_order.find_one({}, {'_id': 0})
    if not doc:
        return {"order": DEFAULT_SUBCATEGORY_ORDER}
    return {"order": doc.get('order', DEFAULT_SUBCATEGORY_ORDER)}

@router.put("/subcategory-order")
async def update_subcategory_order(data: dict, current_user: dict = Depends(get_current_user)):
    """Update subcategory display order"""
    order = data.get('order', [])
    await db.subcategory_order.update_one({}, {'$set': {'order': order}}, upsert=True)
    return {"message": "Subcategory order updated", "order": order}

@router.get("/items/export/pdf")
async def export_items_pdf(
    main_category: Optional[str] = None,
    role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export items as PDF grouped by subcategory, role-based pricing"""
    query = {}
    if main_category:
        query['main_categories'] = main_category
    
    items = await db.items.find(query, {'_id': 0, 'image_webp': 0}).sort('item_name', 1).to_list(5000)
    
    # Get subcategory order
    order_doc = await db.subcategory_order.find_one({}, {'_id': 0})
    sub_order = order_doc.get('order', DEFAULT_SUBCATEGORY_ORDER) if order_doc else DEFAULT_SUBCATEGORY_ORDER
    
    # Get company settings
    company = await db.company_settings.find_one({}, {'_id': 0})
    company_name = company.get('company_name', 'VMP CRM') if company else 'VMP CRM'
    
    # Group items by subcategory
    grouped = {}
    for item in items:
        subs = item.get('subcategories', [])
        if not subs:
            subs = ['Uncategorized']
        for sub in subs:
            if sub not in grouped:
                grouped[sub] = []
            grouped[sub].append(item)
    
    def sort_key(sub_name):
        try:
            return sub_order.index(sub_name)
        except ValueError:
            return len(sub_order)
    
    sorted_subs = sorted(grouped.keys(), key=sort_key)
    
    # Role label mapping
    role_labels = {'doctor': 'Doctors', 'medical': 'Medicals', 'agency': 'Agencies'}
    role_label = role_labels.get(role, 'All Roles')
    
    # Create PDF
    pdf = FPDF(orientation='L', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    
    title = company_name
    subtitle_parts = []
    if main_category:
        subtitle_parts.append(main_category)
    if role:
        subtitle_parts.append(f"Pricing: {role_label}")
    subtitle = ' | '.join(subtitle_parts) if subtitle_parts else 'All Items'
    
    # Column config
    col_widths = [10, 22, 55, 70, 18, 22, 30, 30, 30]
    headers = ['S.No', 'Item Code', 'Item Name', 'Composition', 'MRP', 'Rate', 'Offer', 'Special Offer', 'Special Offer 2']
    
    # Single continuous document
    pdf = FPDF(orientation='L', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # Title header
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, title, ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, subtitle, ln=True, align='C')
    pdf.ln(3)
    
    serial = 0
    
    for sub_name in sorted_subs:
        sub_items = grouped[sub_name]
        
        # Subcategory header row
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(41, 128, 185)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(sum(col_widths), 7, f"  {sub_name}", 1, 1, 'L', True)
        
        # Table header
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_fill_color(52, 73, 94)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 7, h, 1, 0, 'C', True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        
        # Table rows
        pdf.set_font('Helvetica', '', 7)
        for item in sub_items:
            serial += 1
            row_h = 6
            pdf.set_fill_color(245, 245, 245) if serial % 2 == 0 else pdf.set_fill_color(255, 255, 255)
            fill = serial % 2 == 0
            
            if role == 'doctor':
                rate = item.get('rate_doctors') or item.get('rate', 0) or 0
                offer = item.get('offer_doctors') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_doctors') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_doctors') or ''
            elif role == 'medical':
                rate = item.get('rate_medicals') or item.get('rate', 0) or 0
                offer = item.get('offer_medicals') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_medicals') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_medicals') or ''
            elif role == 'agency':
                rate = item.get('rate_agencies') or item.get('rate', 0) or 0
                offer = item.get('offer_agencies') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_agencies') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_agencies') or ''
            else:
                rate = item.get('rate_doctors') or item.get('rate', 0) or 0
                offer = item.get('offer_doctors') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_doctors') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_doctors') or ''
            
            pdf.cell(col_widths[0], row_h, str(serial), 1, 0, 'C', fill)
            pdf.cell(col_widths[1], row_h, str(item.get('item_code', ''))[:12], 1, 0, 'C', fill)
            pdf.cell(col_widths[2], row_h, str(item.get('item_name', ''))[:32], 1, 0, 'L', fill)
            pdf.cell(col_widths[3], row_h, str(item.get('composition', '') or '')[:45], 1, 0, 'L', fill)
            pdf.cell(col_widths[4], row_h, f"{item.get('mrp', 0):.0f}", 1, 0, 'R', fill)
            pdf.cell(col_widths[5], row_h, f"{rate:.0f}" if rate else '', 1, 0, 'R', fill)
            pdf.cell(col_widths[6], row_h, str(offer)[:18], 1, 0, 'L', fill)
            pdf.cell(col_widths[7], row_h, str(sp_offer)[:18], 1, 0, 'L', fill)
            pdf.cell(col_widths[8], row_h, str(sp_offer_2)[:18], 1, 0, 'L', fill)
            pdf.ln()
    
    # Total count at end
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(0, 8, f"Total items: {serial}", ln=True, align='R')
    
    pdf_output = pdf.output()
    filename = f"items_{main_category or 'all'}_{role or 'all'}.pdf".replace(' ', '_').lower()
    
    return Response(
        content=bytes(pdf_output),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/items/export/excel")
async def export_items_excel(
    main_category: Optional[str] = None,
    role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export items as Excel grouped by subcategory, role-based pricing"""
    query = {}
    if main_category:
        query['main_categories'] = main_category
    
    items = await db.items.find(query, {'_id': 0, 'image_webp': 0}).sort('item_name', 1).to_list(5000)
    
    # Get subcategory order
    order_doc = await db.subcategory_order.find_one({}, {'_id': 0})
    sub_order = order_doc.get('order', DEFAULT_SUBCATEGORY_ORDER) if order_doc else DEFAULT_SUBCATEGORY_ORDER
    
    # Group items by subcategory
    grouped = {}
    for item in items:
        subs = item.get('subcategories', [])
        if not subs:
            subs = ['Uncategorized']
        for sub in subs:
            if sub not in grouped:
                grouped[sub] = []
            grouped[sub].append(item)
    
    def sort_key(sub_name):
        try:
            return sub_order.index(sub_name)
        except ValueError:
            return len(sub_order)
    
    sorted_subs = sorted(grouped.keys(), key=sort_key)
    
    role_labels = {'doctor': 'Doctors', 'medical': 'Medicals', 'agency': 'Agencies'}
    role_label = role_labels.get(role, 'All')
    
    # Create Excel
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    sub_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    sub_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    sheet_name = f"{main_category or 'All'} - {role_label}"[:31]
    ws = wb.create_sheet(sheet_name)
    row_num = 1
    
    col_headers = ['S.No', 'Item Code', 'Item Name', 'Composition', 'MRP', 'Rate', 'Offer', 'Special Offer', 'Special Offer 2']
    col_count = len(col_headers)
    serial = 0
    
    for sub_name in sorted_subs:
        sub_items = grouped[sub_name]
        
        # Subcategory header row
        cell = ws.cell(row=row_num, column=1, value=sub_name)
        cell.font = sub_font
        cell.fill = sub_fill
        for c in range(1, col_count + 1):
            ws.cell(row=row_num, column=c).fill = sub_fill
            ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1
        
        # Data rows (continuous serial)
        for item in sub_items:
            serial += 1
            if role == 'doctor':
                rate = item.get('rate_doctors') or item.get('rate', 0) or 0
                offer = item.get('offer_doctors') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_doctors') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_doctors') or ''
            elif role == 'medical':
                rate = item.get('rate_medicals') or item.get('rate', 0) or 0
                offer = item.get('offer_medicals') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_medicals') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_medicals') or ''
            elif role == 'agency':
                rate = item.get('rate_agencies') or item.get('rate', 0) or 0
                offer = item.get('offer_agencies') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_agencies') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_agencies') or ''
            else:
                rate = item.get('rate_doctors') or item.get('rate', 0) or 0
                offer = item.get('offer_doctors') or item.get('offer', '') or ''
                sp_offer = item.get('special_offer_doctors') or item.get('special_offer', '') or ''
                sp_offer_2 = item.get('special_offer_2_doctors') or ''
            
            values = [serial, item.get('item_code', ''), item.get('item_name', ''),
                      item.get('composition', '') or '', item.get('mrp', 0),
                      rate, offer, sp_offer, sp_offer_2]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
            row_num += 1
    
    # Column widths
    for col, w in enumerate([6, 14, 28, 45, 10, 10, 22, 22, 22], 1):
        ws.column_dimensions[chr(64 + col)].width = w
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"items_{main_category or 'all'}_{role or 'all'}.xlsx".replace(' ', '_').lower()
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============== BULK IMPORT ENDPOINTS ==============

@router.get("/items/import/template")
async def get_import_template(current_user: dict = Depends(get_current_user)):
    """Generate and return Excel template for bulk item import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Import"
    
    # Define headers
    headers = [
        "Item Code*", "Item Name*", "Main Categories", "Subcategories", 
        "Composition", "MRP*", "GST %",
        "Rate Doctors", "Offer Doctors", "Special Offer Doctors", "Special Offer 2 Doctors",
        "Rate Medicals", "Offer Medicals", "Special Offer Medicals", "Special Offer 2 Medicals",
        "Rate Agencies", "Offer Agencies", "Special Offer Agencies", "Special Offer 2 Agencies"
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add sample data rows
    sample_data = [
        ["ITM-0001", "Paracetamol 500mg", "Large Animals, Poultry", "Injection, Powder", "Paracetamol IP 500mg", "50.00", "12", "35.00", "Buy 10 Get 1 Free", "", "", "32.00", "", "", "", "30.00", "", "", ""],
        ["ITM-0002", "Vitamin D3 1000IU", "Pets", "Liquids", "Cholecalciferol 1000IU", "250.00", "5", "180.00", "", "", "", "175.00", "", "", "", "170.00", "", "", ""],
        ["", "Sample Item 3", "Large Animals", "Bolus", "Sample composition", "100.00", "18", "75.00", "", "", "", "70.00", "", "", "", "65.00", "", "", ""],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    instructions = wb.create_sheet("Instructions")
    instructions_data = [
        ["BULK ITEM IMPORT INSTRUCTIONS"],
        [""],
        ["Required Fields (marked with *):"],
        ["- Item Code: Unique code (leave empty for auto-generation)"],
        ["- Item Name: Product name (required)"],
        ["- MRP: Maximum Retail Price (required, number)"],
        [""],
        ["Optional Fields:"],
        ["- Main Categories: Comma-separated (e.g., 'Large Animals, Poultry, Pets')"],
        ["- Subcategories: Comma-separated (e.g., 'Injection, Powder, Liquids')"],
        ["- Composition: Product composition text"],
        ["- GST %: GST percentage (number, default 0)"],
        ["- Rate Doctors / Offer Doctors / Special Offer Doctors / Special Offer 2 Doctors: Pricing for Doctor customers"],
        ["- Rate Medicals / Offer Medicals / Special Offer Medicals / Special Offer 2 Medicals: Pricing for Medical Store customers"],
        ["- Rate Agencies / Offer Agencies / Special Offer Agencies / Special Offer 2 Agencies: Pricing for Agency customers"],
        [""],
        ["Notes:"],
        ["- Images can be added manually after import"],
        ["- Default logo will be used for items without images"],
        ["- Delete sample rows before importing"],
        ["- Maximum 500 items per import"],
    ]
    
    for row_num, row_data in enumerate(instructions_data, 1):
        cell = instructions.cell(row=row_num, column=1, value=row_data[0])
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
    
    # Adjust column widths
    column_widths = [15, 25, 30, 30, 35, 12, 10, 14, 20, 20, 20, 14, 20, 20, 20, 14, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items_import_template.xlsx"}
    )

@router.post("/items/import")
async def bulk_import_items(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Bulk import items from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), sheet_name=0)
        
        # Normalize column names (remove * and extra spaces)
        df.columns = [col.replace('*', '').strip() for col in df.columns]
        
        # Expected columns mapping
        column_mapping = {
            'Item Code': 'item_code',
            'Item Name': 'item_name',
            'Main Categories': 'main_categories',
            'Subcategories': 'subcategories',
            'Composition': 'composition',
            'MRP': 'mrp',
            'GST %': 'gst',
            'Rate Doctors': 'rate_doctors',
            'Offer Doctors': 'offer_doctors',
            'Special Offer Doctors': 'special_offer_doctors',
            'Special Offer 2 Doctors': 'special_offer_2_doctors',
            'Rate Medicals': 'rate_medicals',
            'Offer Medicals': 'offer_medicals',
            'Special Offer Medicals': 'special_offer_medicals',
            'Special Offer 2 Medicals': 'special_offer_2_medicals',
            'Rate Agencies': 'rate_agencies',
            'Offer Agencies': 'offer_agencies',
            'Special Offer Agencies': 'special_offer_agencies',
            'Special Offer 2 Agencies': 'special_offer_2_agencies',
            # Legacy support
            'Rate': 'rate',
            'Offer': 'offer',
            'Special Offer': 'special_offer'
        }
        
        # Rename columns
        df = df.rename(columns=column_mapping)
        
        # Validate required columns
        required_cols = ['item_name', 'mrp']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_cols)}")
        
        # Remove rows where item_name is empty
        df = df.dropna(subset=['item_name'])
        df = df[df['item_name'].astype(str).str.strip() != '']
        
        if len(df) == 0:
            raise HTTPException(status_code=400, detail="No valid items found in the file")
        
        if len(df) > 500:
            raise HTTPException(status_code=400, detail="Maximum 500 items per import. Please split your file.")
        
        # Get company logo for default image
        company_settings = await db.company_settings.find_one({}, {'_id': 0})
        default_image = company_settings.get('logo_base64') if company_settings else None
        
        now = datetime.now(timezone.utc)
        items_created = 0
        items_updated = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                item_name = str(row.get('item_name', '')).strip()
                if not item_name:
                    continue
                
                # Parse item_code
                item_code = str(row.get('item_code', '')).strip() if pd.notna(row.get('item_code')) else ''
                
                # Parse main_categories (comma-separated)
                main_cats_str = str(row.get('main_categories', '')) if pd.notna(row.get('main_categories')) else ''
                main_categories = [cat.strip() for cat in main_cats_str.split(',') if cat.strip()]
                
                # Parse subcategories (comma-separated)
                sub_cats_str = str(row.get('subcategories', '')) if pd.notna(row.get('subcategories')) else ''
                subcategories = [cat.strip() for cat in sub_cats_str.split(',') if cat.strip()]
                
                # Parse numeric fields
                mrp = float(row.get('mrp', 0)) if pd.notna(row.get('mrp')) else 0
                gst = float(row.get('gst', 0)) if pd.notna(row.get('gst')) else 0
                
                # Parse role-based pricing
                rate_doctors = float(row.get('rate_doctors', 0)) if pd.notna(row.get('rate_doctors')) else None
                offer_doctors = str(row.get('offer_doctors', '')) if pd.notna(row.get('offer_doctors')) else ''
                special_offer_doctors = str(row.get('special_offer_doctors', '')) if pd.notna(row.get('special_offer_doctors')) else ''
                rate_medicals = float(row.get('rate_medicals', 0)) if pd.notna(row.get('rate_medicals')) else None
                offer_medicals = str(row.get('offer_medicals', '')) if pd.notna(row.get('offer_medicals')) else ''
                special_offer_medicals = str(row.get('special_offer_medicals', '')) if pd.notna(row.get('special_offer_medicals')) else ''
                rate_agencies = float(row.get('rate_agencies', 0)) if pd.notna(row.get('rate_agencies')) else None
                offer_agencies = str(row.get('offer_agencies', '')) if pd.notna(row.get('offer_agencies')) else ''
                special_offer_agencies = str(row.get('special_offer_agencies', '')) if pd.notna(row.get('special_offer_agencies')) else ''
                
                # Parse Special Offer 2 fields
                special_offer_2_doctors = str(row.get('special_offer_2_doctors', '')) if pd.notna(row.get('special_offer_2_doctors')) else ''
                special_offer_2_medicals = str(row.get('special_offer_2_medicals', '')) if pd.notna(row.get('special_offer_2_medicals')) else ''
                special_offer_2_agencies = str(row.get('special_offer_2_agencies', '')) if pd.notna(row.get('special_offer_2_agencies')) else ''
                
                # Legacy fallback: if old Rate/Offer/Special Offer columns exist, use as defaults
                legacy_rate = float(row.get('rate', 0)) if pd.notna(row.get('rate')) else 0
                legacy_offer = str(row.get('offer', '')) if pd.notna(row.get('offer')) else ''
                legacy_special_offer = str(row.get('special_offer', '')) if pd.notna(row.get('special_offer')) else ''
                
                # Apply legacy fallback to role-based if not set
                if rate_doctors is None:
                    rate_doctors = legacy_rate or 0
                if rate_medicals is None:
                    rate_medicals = legacy_rate or 0
                if rate_agencies is None:
                    rate_agencies = legacy_rate or 0
                
                # Parse text fields
                composition = str(row.get('composition', '')) if pd.notna(row.get('composition')) else ''
                
                # Check if item with same code exists
                existing_item = None
                if item_code:
                    existing_item = await db.items.find_one({'item_code': item_code}, {'_id': 0})
                
                if existing_item:
                    # Update existing item
                    update_data = {
                        'item_name': item_name,
                        'main_categories': main_categories,
                        'subcategories': subcategories,
                        'composition': composition,
                        'mrp': mrp,
                        'rate': 0,
                        'gst': gst,
                        'offer': legacy_offer,
                        'special_offer': legacy_special_offer,
                        'rate_doctors': rate_doctors,
                        'offer_doctors': offer_doctors or legacy_offer,
                        'special_offer_doctors': special_offer_doctors or legacy_special_offer,
                        'special_offer_2_doctors': special_offer_2_doctors,
                        'rate_medicals': rate_medicals,
                        'offer_medicals': offer_medicals or legacy_offer,
                        'special_offer_medicals': special_offer_medicals or legacy_special_offer,
                        'special_offer_2_medicals': special_offer_2_medicals,
                        'rate_agencies': rate_agencies,
                        'offer_agencies': offer_agencies or legacy_offer,
                        'special_offer_agencies': special_offer_agencies or legacy_special_offer,
                        'special_offer_2_agencies': special_offer_2_agencies,
                        'updated_at': now.isoformat()
                    }
                    await db.items.update_one({'item_code': item_code}, {'$set': update_data})
                    items_updated += 1
                else:
                    # Generate item_code if not provided
                    if not item_code:
                        item_code = await generate_item_code()
                    
                    # Create new item
                    item_doc = {
                        'id': str(uuid.uuid4()),
                        'item_code': item_code,
                        'item_name': item_name,
                        'main_categories': main_categories,
                        'subcategories': subcategories,
                        'composition': composition,
                        'mrp': mrp,
                        'rate': 0,
                        'gst': gst,
                        'offer': legacy_offer,
                        'special_offer': legacy_special_offer,
                        'rate_doctors': rate_doctors,
                        'offer_doctors': offer_doctors or legacy_offer,
                        'special_offer_doctors': special_offer_doctors or legacy_special_offer,
                        'special_offer_2_doctors': special_offer_2_doctors,
                        'rate_medicals': rate_medicals,
                        'offer_medicals': offer_medicals or legacy_offer,
                        'special_offer_medicals': special_offer_medicals or legacy_special_offer,
                        'special_offer_2_medicals': special_offer_2_medicals,
                        'rate_agencies': rate_agencies,
                        'offer_agencies': offer_agencies or legacy_offer,
                        'special_offer_agencies': special_offer_agencies or legacy_special_offer,
                        'special_offer_2_agencies': special_offer_2_agencies,
                        'has_image': default_image is not None,
                        'image_webp': default_image,
                        'custom_fields': [],
                        'created_at': now.isoformat(),
                        'updated_at': now.isoformat()
                    }
                    await db.items.insert_one(item_doc)
                    items_created += 1
                    
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        result_message = f"Import completed: {items_created} items created, {items_updated} items updated"
        if errors:
            result_message += f". {len(errors)} errors occurred."
        
        return {
            "message": result_message,
            "items_created": items_created,
            "items_updated": items_updated,
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")


# ============== ITEM IMAGES ZIP DOWNLOAD ==============

@router.get("/items/images/download")
async def download_item_images_zip(current_user: dict = Depends(get_current_user)):
    """Download all item images as a ZIP file"""
    import zipfile
    import io
    
    items = await db.items.find(
        {'image_webp': {'$ne': None}},
        {'item_code': 1, 'item_name': 1, 'image_webp': 1, '_id': 0}
    ).to_list(5000)
    
    if not items:
        raise HTTPException(status_code=404, detail="No item images found")
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for item in items:
            if item.get('image_webp'):
                try:
                    image_data = base64.b64decode(item['image_webp'])
                    safe_code = str(item.get('item_code', 'unknown')).replace('/', '_').replace('\\', '_')
                    filename = f"{safe_code}.webp"
                    zf.writestr(filename, image_data)
                except Exception as e:
                    logger.error(f"Error adding image for {item.get('item_code')}: {e}")
    
    zip_buffer.seek(0)
    return Response(
        content=zip_buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=item_images.zip"}
    )


