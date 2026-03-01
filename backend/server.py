from fastapi import FastAPI, APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import json
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import base64
from io import BytesIO
from PIL import Image
import random
import string
import httpx
import asyncio
from contextlib import asynccontextmanager
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from india_locations import get_all_states, get_districts_by_state

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'vmp-crm-secret-key-2024')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI(title="VMP CRM API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Global variable for background task
daily_reminder_task = None

async def send_daily_reminder_summary():
    """Background task to send daily reminder summary to admin via WhatsApp"""
    while True:
        try:
            # Calculate time until next 8 AM
            now = datetime.now(timezone.utc)
            # For IST (UTC+5:30), 8 AM IST = 2:30 AM UTC
            target_hour = 2  # 2 AM UTC = ~7:30 AM IST
            target_minute = 30
            
            next_run = now.replace(hour=target_hour, minute=target_minute, second=0, microsecond=0)
            if now >= next_run:
                # Already past today's run time, schedule for tomorrow
                next_run += timedelta(days=1)
            
            wait_seconds = (next_run - now).total_seconds()
            logger.info(f"Daily reminder task scheduled. Next run in {wait_seconds/3600:.1f} hours")
            
            await asyncio.sleep(wait_seconds)
            
            # Execute reminder summary send
            logger.info("Executing daily reminder summary...")
            
            # Get WhatsApp config
            wa_config = await db.whatsapp_config.find_one({}, {'_id': 0})
            if not wa_config:
                logger.warning("WhatsApp not configured, skipping daily reminder")
                continue
            
            # Get company settings for admin number
            company = await db.company_settings.find_one({}, {'_id': 0})
            admin_phone = company.get('phone') if company else None
            
            if not admin_phone:
                logger.warning("Admin phone not configured, skipping daily reminder")
                continue
            
            # Get today's date for reminders
            today = datetime.now(timezone.utc).date()
            today_str = today.strftime('%Y-%m-%d')
            
            # Get today's reminders count
            reminders = []
            
            # Manual reminders for today
            manual_count = await db.reminders.count_documents({
                'reminder_date': today_str,
                'is_completed': {'$ne': True}
            })
            
            # Get birthday/anniversary counts
            month_day = today.strftime('%m-%d')
            
            # Doctors with birthdays/follow-ups
            doctors = await db.doctors.find({
                'lead_status': {'$nin': ['Not Interested', 'Closed']}
            }, {'_id': 0, 'name': 1, 'dob': 1, 'follow_up_date': 1, 'last_contact_date': 1}).to_list(1000)
            
            birthday_count = 0
            followup_count = 0
            for doc in doctors:
                if doc.get('dob') and doc['dob'][5:] == month_day:
                    birthday_count += 1
                if doc.get('follow_up_date') and doc['follow_up_date'] <= today_str:
                    followup_count += 1
            
            # Medicals and Agencies birthdays/anniversaries
            medicals = await db.medicals.find({}, {'_id': 0, 'name': 1, 'birthday': 1, 'anniversary': 1, 'follow_up_date': 1}).to_list(1000)
            agencies = await db.agencies.find({}, {'_id': 0, 'name': 1, 'birthday': 1, 'anniversary': 1, 'follow_up_date': 1}).to_list(1000)
            
            anniversary_count = 0
            for entity in medicals + agencies:
                if entity.get('birthday') and entity['birthday'][5:] == month_day:
                    birthday_count += 1
                if entity.get('anniversary') and entity['anniversary'][5:] == month_day:
                    anniversary_count += 1
                if entity.get('follow_up_date') and entity['follow_up_date'] <= today_str:
                    followup_count += 1
            
            total_count = manual_count + birthday_count + anniversary_count + followup_count
            
            if total_count == 0:
                logger.info("No reminders today, skipping notification")
                continue
            
            # Collect detailed lists per type
            followup_list = []
            birthday_list = []
            anniversary_list = []

            for doc in doctors:
                if doc.get('dob') and doc['dob'][5:] == month_day:
                    birthday_list.append({'name': doc['name'], 'phone': doc.get('phone', ''), 'type': 'Doctor'})
                if doc.get('follow_up_date') and doc['follow_up_date'] <= today_str:
                    followup_list.append({'name': doc['name'], 'phone': doc.get('phone', ''), 'type': 'Doctor', 'status': doc.get('lead_status', '')})

            for entity in medicals + agencies:
                etype = 'Medical' if entity in medicals else 'Agency'
                if entity.get('birthday') and entity['birthday'][5:] == month_day:
                    birthday_list.append({'name': entity['name'], 'phone': entity.get('phone', ''), 'type': etype})
                if entity.get('anniversary') and entity['anniversary'][5:] == month_day:
                    anniversary_list.append({'name': entity['name'], 'phone': entity.get('phone', ''), 'type': etype})
                if entity.get('follow_up_date') and entity['follow_up_date'] <= today_str:
                    followup_list.append({'name': entity['name'], 'phone': entity.get('phone', ''), 'type': etype, 'status': entity.get('lead_status', '')})

            # Get manual/custom reminders
            manual_reminders = await db.reminders.find({
                'reminder_date': today_str,
                'is_completed': {'$ne': True}
            }, {'_id': 0, 'title': 1, 'entity_name': 1}).to_list(50)

            # Build message
            message_lines = [
                f"🌅 *Good Morning!*",
                f"📅 *Today's Reminders ({today_str})*",
                f"Total: {total_count} reminder(s)",
                ""
            ]

            if followup_list:
                message_lines.append(f"📞 *Follow-ups ({len(followup_list)}):*")
                for item in followup_list[:10]:
                    status_tag = f" [{item['status']}]" if item.get('status') else ''
                    message_lines.append(f"  • {item['name']} - {item['phone']}{status_tag}")
                if len(followup_list) > 10:
                    message_lines.append(f"  +{len(followup_list) - 10} more...")
                message_lines.append("")

            if birthday_list:
                message_lines.append(f"🎂 *Birthdays ({len(birthday_list)}):*")
                for item in birthday_list:
                    message_lines.append(f"  • {item['name']} - {item['phone']} ({item['type']})")
                message_lines.append("")

            if anniversary_list:
                message_lines.append(f"🎉 *Anniversaries ({len(anniversary_list)}):*")
                for item in anniversary_list:
                    message_lines.append(f"  • {item['name']} - {item['phone']} ({item['type']})")
                message_lines.append("")

            if manual_reminders:
                message_lines.append(f"📝 *Custom ({len(manual_reminders)}):*")
                for r in manual_reminders[:5]:
                    name_part = f" - {r['entity_name']}" if r.get('entity_name') else ''
                    message_lines.append(f"  • {r['title']}{name_part}")
                if len(manual_reminders) > 5:
                    message_lines.append(f"  +{len(manual_reminders) - 5} more...")
                message_lines.append("")

            message_lines.append("Login to CRM to view details.")
            
            message = "\n".join(message_lines)
            
            # Send WhatsApp
            try:
                async with httpx.AsyncClient() as http_client:
                    params = {
                        'action': 'send',
                        'senderId': wa_config['sender_id'],
                        'authToken': wa_config['auth_token'],
                        'messageText': message,
                        'receiverId': admin_phone
                    }
                    response = await http_client.get(wa_config['api_url'], params=params, timeout=30)
                    
                    if response.status_code == 200:
                        logger.info(f"Daily reminder sent to admin: {admin_phone}")
                    else:
                        logger.error(f"Failed to send daily reminder: {response.text}")
            except Exception as e:
                logger.error(f"Error sending daily reminder: {str(e)}")
            
        except asyncio.CancelledError:
            logger.info("Daily reminder task cancelled")
            break
        except Exception as e:
            logger.error(f"Error in daily reminder task: {str(e)}")
            # Wait 1 hour before retrying on error
            await asyncio.sleep(3600)


async def send_birthday_anniversary_greetings():
    """Background task to auto-send birthday/anniversary greetings at 10 AM IST via WhatsApp & Email"""
    while True:
        try:
            now = datetime.now(timezone.utc)
            # 10 AM IST = 4:30 AM UTC
            target_hour = 4
            target_minute = 30

            next_run = now.replace(hour=target_hour, minute=target_minute, second=0, microsecond=0)
            if now >= next_run:
                next_run += timedelta(days=1)

            wait_seconds = (next_run - now).total_seconds()
            logger.info(f"Birthday/Anniversary greeting task scheduled. Next run in {wait_seconds/3600:.1f} hours")
            await asyncio.sleep(wait_seconds)

            logger.info("Executing birthday/anniversary greetings...")

            today = datetime.now(timezone.utc).date()
            today_md = today.strftime('%m-%d')

            # Get company name
            company = await db.company_settings.find_one({}, {'_id': 0})
            company_name = company.get('name', 'Our Company') if company else 'Our Company'

            # Get configs
            wa_config = await db.whatsapp_config.find_one({}, {'_id': 0})
            smtp_config = await db.smtp_config.find_one({}, {'_id': 0})

            # Collect all birthday/anniversary people
            contacts = []

            # Doctors (use dob field)
            doctors = await db.doctors.find({'dob': {'$exists': True, '$ne': None, '$ne': ''}}, {'_id': 0}).to_list(1000)
            for doc in doctors:
                if doc.get('dob') and doc['dob'][5:] == today_md:
                    contacts.append({'name': doc['name'], 'phone': doc.get('phone', ''), 'email': doc.get('email', ''), 'type': 'birthday', 'entity_type': 'doctor', 'entity_id': doc['id']})

            # Medicals
            medicals = await db.medicals.find({}, {'_id': 0, 'name': 1, 'id': 1, 'phone': 1, 'email': 1, 'birthday': 1, 'anniversary': 1, 'proprietor_name': 1}).to_list(1000)
            for m in medicals:
                if m.get('birthday') and m['birthday'][5:] == today_md:
                    contacts.append({'name': m.get('proprietor_name') or m['name'], 'phone': m.get('phone', ''), 'email': m.get('email', ''), 'type': 'birthday', 'entity_type': 'medical', 'entity_id': m['id']})
                if m.get('anniversary') and m['anniversary'][5:] == today_md:
                    contacts.append({'name': m['name'], 'phone': m.get('phone', ''), 'email': m.get('email', ''), 'type': 'anniversary', 'entity_type': 'medical', 'entity_id': m['id']})

            # Agencies
            agencies = await db.agencies.find({}, {'_id': 0, 'name': 1, 'id': 1, 'phone': 1, 'email': 1, 'birthday': 1, 'anniversary': 1, 'proprietor_name': 1}).to_list(1000)
            for a in agencies:
                if a.get('birthday') and a['birthday'][5:] == today_md:
                    contacts.append({'name': a.get('proprietor_name') or a['name'], 'phone': a.get('phone', ''), 'email': a.get('email', ''), 'type': 'birthday', 'entity_type': 'agency', 'entity_id': a['id']})
                if a.get('anniversary') and a['anniversary'][5:] == today_md:
                    contacts.append({'name': a['name'], 'phone': a.get('phone', ''), 'email': a.get('email', ''), 'type': 'anniversary', 'entity_type': 'agency', 'entity_id': a['id']})

            if not contacts:
                logger.info("No birthdays/anniversaries today")
                continue

            logger.info(f"Found {len(contacts)} birthday/anniversary contacts today")

            for contact in contacts:
                try:
                    # Get a random active template
                    templates = await db.greeting_templates.find(
                        {'type': contact['type'], 'is_active': True}, {'_id': 0}
                    ).to_list(50)

                    if not templates:
                        logger.warning(f"No active {contact['type']} templates found")
                        continue

                    import random
                    template = random.choice(templates)
                    message = template['message'].replace('{customer_name}', contact['name']).replace('{company_name}', company_name)
                    image_url = template.get('image_url', '')

                    # Send WhatsApp
                    if wa_config and contact.get('phone'):
                        try:
                            async with httpx.AsyncClient() as http_client:
                                if image_url:
                                    params = {
                                        'action': 'send',
                                        'senderId': wa_config['sender_id'],
                                        'authToken': wa_config['auth_token'],
                                        'messageText': message,
                                        'mediaUrl': image_url,
                                        'receiverId': contact['phone']
                                    }
                                else:
                                    params = {
                                        'action': 'send',
                                        'senderId': wa_config['sender_id'],
                                        'authToken': wa_config['auth_token'],
                                        'messageText': message,
                                        'receiverId': contact['phone']
                                    }
                                resp = await http_client.get(wa_config['api_url'], params=params, timeout=30)
                                logger.info(f"WhatsApp greeting sent to {contact['name']} ({contact['phone']}): {resp.status_code}")
                        except Exception as we:
                            logger.error(f"WhatsApp greeting failed for {contact['name']}: {we}")

                    # Send Email
                    if smtp_config and contact.get('email'):
                        try:
                            subject = f"Happy {'Birthday' if contact['type'] == 'birthday' else 'Anniversary'}! - {company_name}"
                            html_body = f"""
                            <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:20px;">
                                {'<img src="' + image_url + '" style="width:100%;max-width:600px;border-radius:12px;margin-bottom:20px;" />' if image_url else ''}
                                <div style="white-space:pre-line;font-size:16px;line-height:1.6;color:#333;">{message}</div>
                                <hr style="margin:20px 0;border:none;border-top:1px solid #eee;" />
                                <p style="font-size:12px;color:#999;">Sent with love from {company_name}</p>
                            </div>"""
                            msg = MIMEMultipart()
                            msg['From'] = f"{smtp_config['from_name']} <{smtp_config['from_email']}>"
                            msg['To'] = f"{contact['name']} <{contact['email']}>"
                            msg['Subject'] = subject
                            msg.attach(MIMEText(html_body, 'html'))
                            with smtplib.SMTP(smtp_config['smtp_server'], smtp_config['smtp_port'], timeout=30) as server:
                                server.starttls()
                                server.login(smtp_config['smtp_username'], smtp_config['smtp_password'])
                                server.sendmail(smtp_config['from_email'], [contact['email']], msg.as_string())
                            logger.info(f"Email greeting sent to {contact['name']} ({contact['email']})")
                        except Exception as ee:
                            logger.error(f"Email greeting failed for {contact['name']}: {ee}")

                    # Log the greeting
                    await db.greeting_logs.insert_one({
                        'id': str(uuid.uuid4()),
                        'contact_name': contact['name'],
                        'contact_phone': contact.get('phone', ''),
                        'contact_email': contact.get('email', ''),
                        'entity_type': contact['entity_type'],
                        'entity_id': contact['entity_id'],
                        'greeting_type': contact['type'],
                        'template_id': template.get('id'),
                        'message': message,
                        'sent_at': datetime.now(timezone.utc).isoformat()
                    })

                except Exception as ce:
                    logger.error(f"Error sending greeting to {contact['name']}: {ce}")

        except asyncio.CancelledError:
            logger.info("Birthday/Anniversary greeting task cancelled")
            break
        except Exception as e:
            logger.error(f"Error in greeting task: {str(e)}")
            await asyncio.sleep(3600)


async def seed_default_greeting_templates():
    """Seed default greeting templates if none exist"""
    count = await db.greeting_templates.count_documents({})
    if count > 0:
        return

    templates = [
        # Birthday templates
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Dear {customer_name},\n\nWishing you a very Happy Birthday! May this special day bring you endless joy and happiness.\n\nWarm regards,\n{company_name} Family",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Happy Birthday, {customer_name}!\n\nMay your birthday be filled with sunshine and smiles, laughter and love. Here's to another wonderful year!\n\nBest wishes,\n{company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Many happy returns of the day, {customer_name}!\n\nOn your special day, we wish you nothing but the best. May all your dreams come true.\n\nWith love,\nTeam {company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Dear {customer_name},\n\nHappiest birthday to you! We value our association with you and wish you good health and prosperity.\n\nCheers,\n{company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Happy Birthday {customer_name}!\n\nAnother year of success, health and happiness. Wishing you the best on your big day!\n\nRegards,\n{company_name} Team",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Dear {customer_name},\n\nBirthdays are nature's way of telling us to eat more cake! Wishing you a day filled with sweet moments.\n\nHappy Birthday!\n{company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Warmest wishes on your birthday, {customer_name}!\n\nMay this year bring new goals, new achievements and a whole lot of new inspirations.\n\nFrom all of us at {company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "birthday", "is_active": True,
         "message": "Dear {customer_name},\n\nWishing you a birthday that's as special as you are! Thank you for being a valued part of our family.\n\nHappy Birthday!\n{company_name}",
         "image_url": ""},
        # Anniversary templates
        {"id": str(uuid.uuid4()), "type": "anniversary", "is_active": True,
         "message": "Dear {customer_name},\n\nHappy Anniversary! Congratulations on another successful year. Wishing you continued growth and prosperity.\n\nBest regards,\n{company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "anniversary", "is_active": True,
         "message": "Congratulations on your anniversary, {customer_name}!\n\nAnother milestone achieved! We are proud to be associated with your journey of success.\n\nWarm wishes,\n{company_name} Team",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "anniversary", "is_active": True,
         "message": "Happy Anniversary {customer_name}!\n\nCelebrating your remarkable journey today. May the years ahead bring even more success and achievements.\n\nBest wishes,\n{company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "anniversary", "is_active": True,
         "message": "Dear {customer_name},\n\nWishing you a wonderful anniversary! Thank you for your continued trust and partnership with us.\n\nWith appreciation,\n{company_name} Family",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "anniversary", "is_active": True,
         "message": "Happy Anniversary {customer_name}!\n\nEvery year together is a year of growth. Here's to many more successful years ahead!\n\nCheers,\nTeam {company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "anniversary", "is_active": True,
         "message": "Dear {customer_name},\n\nOn this special anniversary, we celebrate the wonderful bond we share. May it grow stronger with each passing year.\n\nHeartfelt wishes,\n{company_name}",
         "image_url": ""},
        {"id": str(uuid.uuid4()), "type": "anniversary", "is_active": True,
         "message": "Congratulations {customer_name}!\n\nAnother glorious year! Your dedication inspires us. Wishing you an amazing anniversary and a bright future.\n\nRegards,\n{company_name}",
         "image_url": ""},
    ]

    for t in templates:
        t['created_at'] = datetime.now(timezone.utc).isoformat()
    await db.greeting_templates.insert_many(templates)
    logger.info(f"Seeded {len(templates)} default greeting templates")

# ============== MODELS ==============

# Greeting Template Models
class GreetingTemplateCreate(BaseModel):
    type: str  # birthday, anniversary
    message: str  # Template with {customer_name}, {company_name} placeholders
    image_url: Optional[str] = None
    is_active: bool = True

class GreetingTemplateUpdate(BaseModel):
    type: Optional[str] = None
    message: Optional[str] = None
    image_url: Optional[str] = None
    is_active: Optional[bool] = None

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "staff"  # admin or staff

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    permissions: Optional[dict] = None
    created_at: datetime

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class DoctorCreate(BaseModel):
    name: str
    reg_no: str
    address: Optional[str] = None  # Legacy field for backward compatibility
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    district: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None  # Preferred transport
    email: EmailStr
    phone: str
    lead_status: str = "Pipeline"
    dob: Optional[str] = None

class DoctorUpdate(BaseModel):
    name: Optional[str] = None
    reg_no: Optional[str] = None
    address: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    district: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    lead_status: Optional[str] = None
    dob: Optional[str] = None
    priority: Optional[str] = None  # low, moderate, high
    last_contact_date: Optional[str] = None
    follow_up_date: Optional[str] = None

class DoctorResponse(BaseModel):
    id: str
    customer_code: str
    name: str
    reg_no: str
    address: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    district: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None  # Populated from transport lookup
    email: str
    phone: str
    lead_status: str
    dob: Optional[str] = None
    priority: Optional[str] = None
    last_contact_date: Optional[str] = None
    follow_up_date: Optional[str] = None
    is_portal_customer: Optional[bool] = False  # Flag for portal customers
    portal_customer_id: Optional[str] = None  # Link to portal_customers
    created_at: datetime
    updated_at: datetime

class DoctorNoteCreate(BaseModel):
    note: str

class DoctorNoteResponse(BaseModel):
    id: str
    doctor_id: str
    note: str
    created_by: str
    created_at: datetime

# ============== MEDICAL MODELS ==============

class MedicalCreate(BaseModel):
    name: str
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address: Optional[str] = None  # Legacy field
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None  # Preferred transport
    email: Optional[EmailStr] = None
    phone: str
    alternate_phone: Optional[str] = None
    lead_status: str = "Pipeline"
    birthday: Optional[str] = None  # Format: YYYY-MM-DD
    anniversary: Optional[str] = None  # Format: YYYY-MM-DD

class MedicalUpdate(BaseModel):
    name: Optional[str] = None
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    alternate_phone: Optional[str] = None
    lead_status: Optional[str] = None
    priority: Optional[str] = None
    last_contact_date: Optional[str] = None
    follow_up_date: Optional[str] = None
    birthday: Optional[str] = None
    anniversary: Optional[str] = None

class MedicalResponse(BaseModel):
    id: str
    customer_code: str
    name: str
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None  # Populated from transport lookup
    email: Optional[str] = None
    phone: str
    alternate_phone: Optional[str] = None
    lead_status: str
    priority: Optional[str] = None
    last_contact_date: Optional[str] = None
    follow_up_date: Optional[str] = None
    birthday: Optional[str] = None
    anniversary: Optional[str] = None
    is_portal_customer: Optional[bool] = False  # Flag for portal customers
    portal_customer_id: Optional[str] = None  # Link to portal_customers
    created_at: datetime
    updated_at: datetime

class MedicalNoteCreate(BaseModel):
    note: str

class MedicalNoteResponse(BaseModel):
    id: str
    medical_id: str
    note: str
    created_by: str
    created_at: datetime

# ============== AGENCY MODELS ==============

class AgencyCreate(BaseModel):
    name: str
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address: Optional[str] = None  # Legacy field
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None  # Preferred transport
    email: Optional[EmailStr] = None
    phone: str
    alternate_phone: Optional[str] = None
    lead_status: str = "Pipeline"
    birthday: Optional[str] = None
    anniversary: Optional[str] = None

class AgencyUpdate(BaseModel):
    name: Optional[str] = None
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    alternate_phone: Optional[str] = None
    lead_status: Optional[str] = None
    priority: Optional[str] = None
    last_contact_date: Optional[str] = None
    follow_up_date: Optional[str] = None
    birthday: Optional[str] = None
    anniversary: Optional[str] = None

class AgencyResponse(BaseModel):
    id: str
    customer_code: str
    name: str
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None  # Populated from transport lookup
    email: Optional[str] = None
    phone: str
    alternate_phone: Optional[str] = None
    lead_status: str
    priority: Optional[str] = None
    last_contact_date: Optional[str] = None
    follow_up_date: Optional[str] = None
    birthday: Optional[str] = None
    anniversary: Optional[str] = None
    is_portal_customer: Optional[bool] = False  # Flag for portal customers
    portal_customer_id: Optional[str] = None  # Link to portal_customers
    created_at: datetime
    updated_at: datetime

class AgencyNoteCreate(BaseModel):
    note: str

class AgencyNoteResponse(BaseModel):
    id: str
    agency_id: str
    note: str
    created_by: str
    created_at: datetime

# ============== TASK MODELS ==============

class TaskCreate(BaseModel):
    doctor_id: Optional[str] = None
    medical_id: Optional[str] = None
    agency_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    due_date: Optional[str] = None
    priority: Optional[str] = "moderate"  # low, moderate, high

class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    due_date: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None  # pending, completed

class TaskResponse(BaseModel):
    id: str
    doctor_id: str
    doctor_name: Optional[str] = None
    title: str
    description: Optional[str] = None
    due_date: Optional[str] = None
    priority: str
    status: str
    created_at: datetime

class SMTPConfigCreate(BaseModel):
    smtp_server: str
    smtp_port: int
    smtp_username: str
    smtp_password: str
    from_email: EmailStr
    from_name: str = "VMP CRM"

class SMTPConfigResponse(BaseModel):
    id: str
    smtp_server: str
    smtp_port: int
    smtp_username: str
    from_email: str
    from_name: str
    created_at: datetime
    updated_at: datetime

class SendEmailRequest(BaseModel):
    doctor_id: str
    subject: str
    body: str
    is_html: bool = False

class EmailLogResponse(BaseModel):
    id: str
    doctor_id: str
    doctor_name: str
    doctor_email: str
    subject: str
    status: str
    sent_at: Optional[datetime] = None
    error_message: Optional[str] = None

class DashboardStats(BaseModel):
    total_doctors: int
    by_status: dict
    recent_emails: int
    recent_doctors: List[DoctorResponse]

# ============== ITEM MODELS ==============

class CustomField(BaseModel):
    field_name: str
    field_value: str

class ItemCreate(BaseModel):
    item_name: str
    item_code: Optional[str] = None  # Custom item code, auto-generated if not provided
    main_categories: Optional[List[str]] = []  # Main categories: Large Animals, Poultry, Pets (multiple)
    subcategories: Optional[List[str]] = []  # Subcategories: Injection, Liquids, Bolus, Powder
    composition: Optional[str] = None
    # Common pricing
    mrp: float
    gst: float = 0
    # Role-based pricing - Doctors
    rate_doctors: Optional[float] = None
    offer_doctors: Optional[str] = None
    special_offer_doctors: Optional[str] = None
    # Role-based pricing - Medicals
    rate_medicals: Optional[float] = None
    offer_medicals: Optional[str] = None
    special_offer_medicals: Optional[str] = None
    # Role-based pricing - Agencies
    rate_agencies: Optional[float] = None
    offer_agencies: Optional[str] = None
    special_offer_agencies: Optional[str] = None
    # Legacy fields (for backward compatibility)
    rate: Optional[float] = None
    offer: Optional[str] = None
    special_offer: Optional[str] = None
    custom_fields: Optional[List[CustomField]] = []
    image_base64: Optional[str] = None  # Base64 encoded image

class ItemUpdate(BaseModel):
    item_name: Optional[str] = None
    item_code: Optional[str] = None
    main_categories: Optional[List[str]] = None
    subcategories: Optional[List[str]] = None
    composition: Optional[str] = None
    mrp: Optional[float] = None
    gst: Optional[float] = None
    # Role-based pricing - Doctors
    rate_doctors: Optional[float] = None
    offer_doctors: Optional[str] = None
    special_offer_doctors: Optional[str] = None
    # Role-based pricing - Medicals
    rate_medicals: Optional[float] = None
    offer_medicals: Optional[str] = None
    special_offer_medicals: Optional[str] = None
    # Role-based pricing - Agencies
    rate_agencies: Optional[float] = None
    offer_agencies: Optional[str] = None
    special_offer_agencies: Optional[str] = None
    # Legacy fields
    rate: Optional[float] = None
    offer: Optional[str] = None
    special_offer: Optional[str] = None
    custom_fields: Optional[List[CustomField]] = None
    image_base64: Optional[str] = None

class ItemResponse(BaseModel):
    id: str
    item_code: str
    item_name: str
    main_categories: List[str] = []
    subcategories: List[str] = []
    composition: Optional[str] = None
    mrp: float
    gst: float
    # Role-based pricing - Doctors
    rate_doctors: Optional[float] = None
    offer_doctors: Optional[str] = None
    special_offer_doctors: Optional[str] = None
    # Role-based pricing - Medicals
    rate_medicals: Optional[float] = None
    offer_medicals: Optional[str] = None
    special_offer_medicals: Optional[str] = None
    # Role-based pricing - Agencies
    rate_agencies: Optional[float] = None
    offer_agencies: Optional[str] = None
    special_offer_agencies: Optional[str] = None
    # Legacy fields (for backward compatibility)
    rate: float
    offer: Optional[str] = None
    special_offer: Optional[str] = None
    custom_fields: List[CustomField] = []
    image_url: Optional[str] = None
    created_at: datetime

class CategoryResponse(BaseModel):
    name: str
    count: int

# ============== CUSTOMER PORTAL MODELS ==============

class CustomerRegister(BaseModel):
    """Customer registration request"""
    name: str
    phone: str
    email: Optional[EmailStr] = None
    password: str
    role: str  # doctor, medical, agency
    # Role-specific fields (Doctor)
    reg_no: Optional[str] = None
    dob: Optional[str] = None
    # Role-specific fields (Medical/Agency)
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    alternate_phone: Optional[str] = None
    birthday: Optional[str] = None
    anniversary: Optional[str] = None
    # Address fields
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None

class CustomerLogin(BaseModel):
    phone: str
    password: str

class CustomerOTPRequest(BaseModel):
    phone: str
    purpose: str = "register"  # register, reset_password

class CustomerOTPVerify(BaseModel):
    phone: str
    otp: str
    purpose: str = "register"

class CustomerResetPassword(BaseModel):
    phone: str
    otp: str
    new_password: str

class CustomerProfileUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    reg_no: Optional[str] = None
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None

class CustomerResponse(BaseModel):
    id: str
    customer_code: str
    name: str
    phone: str
    email: Optional[str] = None
    role: str  # doctor, medical, agency
    status: str  # pending_approval, approved, rejected, suspended
    reg_no: Optional[str] = None
    proprietor_name: Optional[str] = None
    gst_number: Optional[str] = None
    drug_license: Optional[str] = None
    address_line_1: Optional[str] = None
    address_line_2: Optional[str] = None
    state: Optional[str] = None
    district: Optional[str] = None
    pincode: Optional[str] = None
    delivery_station: Optional[str] = None
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None
    created_at: datetime
    approved_at: Optional[datetime] = None
    approved_by: Optional[str] = None

class CustomerApproval(BaseModel):
    status: str  # approved, rejected, suspended
    rejection_reason: Optional[str] = None

# ============== FALLBACK OTP MODELS ==============

class FallbackOTPCreate(BaseModel):
    otp: str

class FallbackOTPResponse(BaseModel):
    id: str
    otp: str
    is_active: bool
    used_count: int
    created_at: datetime
    created_by: str

# ============== MARKETING MODELS ==============

class MarketingTemplateCreate(BaseModel):
    name: str
    category: str  # greeting, product_promo, announcement, circular
    message: str
    is_active: bool = True

class MarketingTemplateResponse(BaseModel):
    id: str
    name: str
    category: str
    message: str
    is_active: bool
    created_at: datetime
    created_by: str

class MarketingCampaignCreate(BaseModel):
    name: str
    campaign_type: str  # product_promo, greeting, announcement, circular
    target_entity: str  # doctors, medicals, agencies, all
    target_status: str  # all, pipeline, customer, contacted, not_interested, closed
    recipient_ids: List[str]  # Selected recipient IDs
    message: str
    item_ids: Optional[List[str]] = None  # For product promotions
    image_base64: Optional[str] = None  # Optional image attachment
    scheduled_at: Optional[str] = None  # ISO datetime for scheduling
    batch_size: int = 10  # Messages per batch
    batch_delay_seconds: int = 60  # Delay between batches (to avoid ban)

class MarketingCampaignResponse(BaseModel):
    id: str
    name: str
    campaign_type: str
    target_entity: str
    target_status: str
    total_recipients: int
    sent_count: int
    failed_count: int
    pending_count: int
    status: str  # draft, scheduled, sending, completed, cancelled
    message_preview: str
    has_image: bool
    scheduled_at: Optional[datetime] = None
    started_at: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    created_at: datetime
    created_by: str

class CampaignLogResponse(BaseModel):
    id: str
    campaign_id: str
    recipient_id: str
    recipient_name: str
    recipient_phone: str
    recipient_type: str  # doctor, medical, agency
    reference_number: str  # Random ref for anti-ban
    status: str  # pending, sent, failed
    error_message: Optional[str] = None
    sent_at: Optional[datetime] = None

# ============== SUPPORT TICKET MODELS ==============

class TicketCreate(BaseModel):
    subject: str
    description: str
    order_id: Optional[str] = None  # Link to order if related
    priority: str = "medium"  # low, medium, high

class TicketReply(BaseModel):
    message: str

class TicketResponse(BaseModel):
    id: str
    ticket_number: str
    customer_id: str
    customer_name: str
    customer_phone: str
    customer_role: str
    subject: str
    description: str
    order_id: Optional[str] = None
    order_number: Optional[str] = None
    priority: str
    status: str  # open, in_progress, resolved, closed
    replies: List[dict] = []
    created_at: datetime
    updated_at: datetime
    resolved_at: Optional[datetime] = None

# ============== COMPANY SETTINGS MODELS ==============

class CompanySettingsCreate(BaseModel):
    company_name: str
    address: str
    email: EmailStr
    phone: Optional[str] = None
    gst_number: str
    drug_license: str
    logo_base64: Optional[str] = None
    terms_conditions: Optional[str] = None
    # Login page customization
    login_tagline: Optional[str] = None  # Custom tagline for login page
    login_background_color: Optional[str] = None  # Hex color for background
    login_background_image: Optional[str] = None  # Base64 image for background

class CompanySettingsResponse(BaseModel):
    id: str
    company_name: str
    address: str
    email: str
    phone: Optional[str] = None
    gst_number: str
    drug_license: str
    logo_url: Optional[str] = None
    terms_conditions: Optional[str] = None
    login_tagline: Optional[str] = None
    login_background_color: Optional[str] = None
    login_background_image_url: Optional[str] = None
    updated_at: datetime

# ============== WHATSAPP LOG MODELS ==============

class WhatsAppLogResponse(BaseModel):
    id: str
    recipient_phone: str
    recipient_name: Optional[str] = None
    message_type: str  # otp, order_confirmation, status_update, reminder, stock_arrived, etc.
    message_preview: str
    status: str  # success, failed
    error_message: Optional[str] = None
    created_at: datetime

# ============== USER MANAGEMENT MODELS ==============

class UserPermissions(BaseModel):
    doctors: bool = True
    medicals: bool = True
    agencies: bool = True
    items: bool = True
    orders: bool = True
    delete_orders: bool = False
    expenses: bool = True
    reminders: bool = True
    pending_items: bool = True
    marketing: bool = True
    support: bool = True
    portal_customers: bool = True
    email_logs: bool = False
    whatsapp_logs: bool = False
    users: bool = False
    smtp_settings: bool = False
    company_settings: bool = False
    whatsapp_settings: bool = False
    backup: bool = False

class UserCreateByAdmin(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "staff"
    permissions: Optional[UserPermissions] = None

class UserUpdateByAdmin(BaseModel):
    email: Optional[EmailStr] = None
    name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None
    permissions: Optional[UserPermissions] = None

class UserWithPermissions(BaseModel):
    id: str
    email: str
    name: str
    role: str
    permissions: Optional[dict] = None
    created_at: datetime

# ============== ORDER MODELS ==============

class OrderItem(BaseModel):
    item_id: str
    item_code: str
    item_name: str
    quantity: str  # Can be "10" or "10+2" format
    mrp: float
    rate: float

class OrderCreate(BaseModel):
    mobile: str
    items: List[OrderItem]
    terms_accepted: bool = True

class OTPRequest(BaseModel):
    mobile: str

class OTPVerify(BaseModel):
    mobile: str
    otp: str
    items: List[OrderItem]
    doctor_info: Optional[dict] = None
    ip_address: Optional[str] = None
    location: Optional[str] = None
    device_info: Optional[str] = None

class ManualOrderCreate(BaseModel):
    customer_name: str
    customer_phone: str
    customer_email: Optional[str] = None
    customer_address: Optional[str] = None
    customer_type: str = "doctor"  # doctor, medical, agency
    customer_id: Optional[str] = None  # If linking to existing entity
    items: List[OrderItem]
    pending_items: Optional[List[dict]] = None  # Items to mark as pending (out of stock)

class OrderResponse(BaseModel):
    id: str
    order_number: str
    doctor_id: Optional[str] = None
    doctor_name: Optional[str] = None
    doctor_phone: str
    doctor_email: Optional[str] = None
    doctor_address: Optional[str] = None
    items: List[OrderItem]
    status: str
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None
    tracking_number: Optional[str] = None
    tracking_url: Optional[str] = None
    delivery_station: Optional[str] = None
    payment_mode: Optional[str] = None
    payment_amount: Optional[float] = None
    expense_paid_by: Optional[str] = None
    expense_account: Optional[str] = None
    # Shipping package details
    boxes_count: Optional[int] = None
    cans_count: Optional[int] = None
    bags_count: Optional[int] = None
    # Invoice details
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    invoice_value: Optional[float] = None
    # Cancellation
    cancellation_reason: Optional[str] = None
    ip_address: Optional[str] = None
    location: Optional[str] = None
    device_info: Optional[str] = None
    created_at: datetime

# ============== WHATSAPP CONFIG MODELS ==============

class WhatsAppConfigCreate(BaseModel):
    api_url: str
    auth_token: str
    sender_id: str

class WhatsAppConfigResponse(BaseModel):
    id: str
    api_url: str
    sender_id: str
    updated_at: datetime

# ============== TRANSPORT MODELS ==============

class TransportCreate(BaseModel):
    name: str
    tracking_url_template: Optional[str] = None  # Transport website URL
    is_local: bool = False  # Local supply - no tracking needed
    contact_number: Optional[str] = None  # Transport incharge contact number
    alternate_number: Optional[str] = None  # Alternate contact number

class TransportResponse(BaseModel):
    id: str
    name: str
    tracking_url_template: Optional[str] = None
    is_local: bool
    contact_number: Optional[str] = None
    alternate_number: Optional[str] = None
    created_at: datetime

class OrderStatusUpdate(BaseModel):
    status: str
    # Transport details (only for shipped status)
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None
    tracking_number: Optional[str] = None
    tracking_url: Optional[str] = None
    delivery_station: Optional[str] = None
    payment_mode: Optional[str] = None  # "to_pay" or "paid"
    payment_amount: Optional[float] = None  # Amount for to_pay or paid
    # Expense details (only for 'paid' payment mode)
    expense_paid_by: Optional[str] = None  # Who paid for the transport
    expense_account: Optional[str] = None  # Which account was used
    # Package counts (only for shipped status)
    boxes_count: Optional[int] = None
    cans_count: Optional[int] = None
    bags_count: Optional[int] = None
    # Invoice details (only for shipped status)
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None
    invoice_value: Optional[float] = None
    # Cancellation reason (only for cancelled status)
    cancellation_reason: Optional[str] = None

# ============== PENDING ITEMS MODELS ==============

class PendingItemCreate(BaseModel):
    doctor_phone: str
    doctor_name: Optional[str] = None
    item_id: str
    item_code: str
    item_name: str
    quantity: str
    original_order_id: str
    original_order_number: str

class PendingItemResponse(BaseModel):
    id: str
    doctor_phone: str
    doctor_name: Optional[str] = None
    item_id: str
    item_code: str
    item_name: str
    quantity: str
    original_order_id: str
    original_order_number: str
    original_order_date: datetime
    created_at: datetime

class OrderItemsUpdate(BaseModel):
    items: List[OrderItem]
    pending_items: Optional[List[dict]] = None  # Items to mark as pending

class OrderCustomerUpdate(BaseModel):
    doctor_name: Optional[str] = None
    doctor_email: Optional[str] = None
    doctor_address: Optional[str] = None
    doctor_phone: Optional[str] = None
    link_to_doctor: Optional[bool] = False  # If true, link/create doctor record

# ============== EXPENSE MODELS ==============

class ExpenseCategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None

class ExpenseCategoryResponse(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    is_default: bool = False
    created_at: datetime

class ExpenseCreate(BaseModel):
    category_id: str
    date: str
    amount: float
    payment_type: str  # cash, card, upi, net_banking
    payment_account: str  # company_account, admin_user, employee_user
    paid_by: Optional[str] = None  # Name of person who paid
    reason: str
    # For transport expenses
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None
    transport_location: Optional[str] = None
    order_id: Optional[str] = None
    order_number: Optional[str] = None

class ExpenseUpdate(BaseModel):
    category_id: Optional[str] = None
    date: Optional[str] = None
    amount: Optional[float] = None
    payment_type: Optional[str] = None
    payment_account: Optional[str] = None
    paid_by: Optional[str] = None
    reason: Optional[str] = None

class ExpenseResponse(BaseModel):
    id: str
    category_id: str
    category_name: Optional[str] = None
    date: str
    amount: float
    payment_type: str
    payment_account: str
    paid_by: Optional[str] = None
    reason: str
    transport_id: Optional[str] = None
    transport_name: Optional[str] = None
    transport_location: Optional[str] = None
    order_id: Optional[str] = None
    order_number: Optional[str] = None
    is_auto_generated: bool = False
    created_at: datetime
    updated_at: datetime

# ============== REMINDER MODELS ==============

class ReminderCreate(BaseModel):
    title: str
    description: Optional[str] = None
    reminder_type: str  # follow_up, birthday, anniversary, custom
    reminder_date: str  # YYYY-MM-DD
    reminder_time: Optional[str] = None  # HH:MM
    entity_type: Optional[str] = None  # doctor, medical, agency
    entity_id: Optional[str] = None
    entity_name: Optional[str] = None
    priority: str = "moderate"  # low, moderate, high

class ReminderUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    reminder_type: Optional[str] = None
    reminder_date: Optional[str] = None
    reminder_time: Optional[str] = None
    priority: Optional[str] = None
    is_completed: Optional[bool] = None

class ReminderResponse(BaseModel):
    id: str
    title: str
    description: Optional[str] = None
    reminder_type: str
    reminder_date: str
    reminder_time: Optional[str] = None
    entity_type: Optional[str] = None
    entity_id: Optional[str] = None
    entity_name: Optional[str] = None
    priority: str
    is_completed: bool = False
    is_auto_generated: bool = False
    created_at: datetime

# ============== FOLLOWUP MODELS ==============

class FollowUpCreate(BaseModel):
    entity_type: str  # doctor, medical, agency
    entity_id: str
    notes: str
    new_status: Optional[str] = None  # Update lead status
    next_follow_up_date: Optional[str] = None  # YYYY-MM-DD
    next_follow_up_time: Optional[str] = None  # HH:MM

class FollowUpResponse(BaseModel):
    id: str
    entity_type: str
    entity_id: str
    entity_name: str
    notes: str
    new_status: Optional[str] = None
    next_follow_up_date: Optional[str] = None
    next_follow_up_time: Optional[str] = None
    status: str  # open, closed
    created_by: str
    created_at: str

# ============== AUTH HELPERS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({'id': payload['user_id']}, {'_id': 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============== CUSTOMER CODE GENERATOR ==============

async def generate_customer_code() -> str:
    last_doctor = await db.doctors.find_one(
        {},
        {'customer_code': 1},
        sort=[('customer_code', -1)]
    )
    if last_doctor and 'customer_code' in last_doctor:
        last_num = int(last_doctor['customer_code'].replace('VMP-', ''))
        new_num = last_num + 1
    else:
        new_num = 1
    return f"VMP-{str(new_num).zfill(4)}"

async def generate_medical_code() -> str:
    last_medical = await db.medicals.find_one(
        {'customer_code': {'$regex': '^MED-'}},
        {'customer_code': 1},
        sort=[('customer_code', -1)]
    )
    if last_medical and 'customer_code' in last_medical:
        try:
            last_num = int(last_medical['customer_code'].replace('MED-', ''))
            new_num = last_num + 1
        except ValueError:
            new_num = 1
    else:
        new_num = 1
    return f"MED-{str(new_num).zfill(4)}"

async def generate_agency_code() -> str:
    last_agency = await db.agencies.find_one(
        {'customer_code': {'$regex': '^AGY-'}},
        {'customer_code': 1},
        sort=[('customer_code', -1)]
    )
    if last_agency and 'customer_code' in last_agency:
        try:
            last_num = int(last_agency['customer_code'].replace('AGY-', ''))
            new_num = last_num + 1
        except ValueError:
            new_num = 1
    else:
        new_num = 1
    return f"AGY-{str(new_num).zfill(4)}"

async def generate_item_code() -> str:
    last_item = await db.items.find_one(
        {},
        {'item_code': 1},
        sort=[('item_code', -1)]
    )
    if last_item and 'item_code' in last_item:
        # Try to extract number from existing code
        code = last_item['item_code']
        if code.startswith('ITM-'):
            try:
                last_num = int(code.replace('ITM-', ''))
                new_num = last_num + 1
                return f"ITM-{str(new_num).zfill(4)}"
            except ValueError:
                pass
    return "ITM-0001"

def process_image_to_webp(image_data: bytes, max_size_kb: int = 25, target_size: tuple = (100, 100)) -> str:
    """
    Process image: resize to 100x100, convert to WebP, compress to under 25KB
    Returns base64 encoded WebP image
    """
    try:
        # Open image
        img = Image.open(BytesIO(image_data))
        
        # Convert to RGB if necessary (for PNG with transparency)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # Resize to target size (100x100) maintaining aspect ratio and crop
        img.thumbnail(target_size, Image.Resampling.LANCZOS)
        
        # Create a new image with exact target size and paste resized image centered
        new_img = Image.new('RGB', target_size, (255, 255, 255))
        offset = ((target_size[0] - img.size[0]) // 2, (target_size[1] - img.size[1]) // 2)
        new_img.paste(img, offset)
        
        # Compress to WebP with quality adjustment to meet size requirement
        quality = 85
        while quality > 10:
            buffer = BytesIO()
            new_img.save(buffer, format='WEBP', quality=quality, optimize=True)
            size_kb = buffer.tell() / 1024
            
            if size_kb <= max_size_kb:
                buffer.seek(0)
                return base64.b64encode(buffer.read()).decode('utf-8')
            
            quality -= 10
        
        # Final attempt with lowest quality
        buffer = BytesIO()
        new_img.save(buffer, format='WEBP', quality=10, optimize=True)
        buffer.seek(0)
        return base64.b64encode(buffer.read()).decode('utf-8')
        
    except Exception as e:
        logger.error(f"Image processing error: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Image processing failed: {str(e)}")

# ============== AUTH ROUTES ==============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({'email': user_data.email}, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    user_doc = {
        'id': user_id,
        'email': user_data.email,
        'password': hash_password(user_data.password),
        'name': user_data.name,
        'role': user_data.role,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, user_data.email, user_data.role)
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user_id,
            email=user_data.email,
            name=user_data.name,
            role=user_data.role,
            created_at=now
        )
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({'email': credentials.email}, {'_id': 0})
    if not user or not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    token = create_token(user['id'], user['email'], user['role'])
    
    created_at = user['created_at']
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user['id'],
            email=user['email'],
            name=user['name'],
            role=user['role'],
            created_at=created_at
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    created_at = current_user['created_at']
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    
    return UserResponse(
        id=current_user['id'],
        email=current_user['email'],
        name=current_user['name'],
        role=current_user['role'],
        permissions=current_user.get('permissions'),
        created_at=created_at
    )

# ============== DOCTOR ROUTES ==============

@api_router.post("/doctors", response_model=DoctorResponse)
async def create_doctor(doctor_data: DoctorCreate, current_user: dict = Depends(get_current_user)):
    doctor_id = str(uuid.uuid4())
    customer_code = await generate_customer_code()
    now = datetime.now(timezone.utc)
    
    # Get transport name if transport_id provided
    transport_name = None
    if doctor_data.transport_id:
        transport = await db.transports.find_one({'id': doctor_data.transport_id}, {'_id': 0, 'name': 1})
        transport_name = transport.get('name') if transport else None
    
    doctor_doc = {
        'id': doctor_id,
        'customer_code': customer_code,
        'name': doctor_data.name,
        'reg_no': doctor_data.reg_no,
        'address': doctor_data.address,
        'address_line_1': doctor_data.address_line_1,
        'address_line_2': doctor_data.address_line_2,
        'district': doctor_data.district,
        'state': doctor_data.state,
        'pincode': doctor_data.pincode,
        'delivery_station': doctor_data.delivery_station,
        'transport_id': doctor_data.transport_id,
        'email': doctor_data.email,
        'phone': doctor_data.phone,
        'lead_status': doctor_data.lead_status,
        'dob': doctor_data.dob,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': current_user['id']
    }
    
    await db.doctors.insert_one(doctor_doc)
    
    return DoctorResponse(
        id=doctor_id,
        customer_code=customer_code,
        name=doctor_data.name,
        reg_no=doctor_data.reg_no,
        address=doctor_data.address,
        address_line_1=doctor_data.address_line_1,
        address_line_2=doctor_data.address_line_2,
        district=doctor_data.district,
        state=doctor_data.state,
        pincode=doctor_data.pincode,
        delivery_station=doctor_data.delivery_station,
        transport_id=doctor_data.transport_id,
        transport_name=transport_name,
        email=doctor_data.email,
        phone=doctor_data.phone,
        lead_status=doctor_data.lead_status,
        dob=doctor_data.dob,
        created_at=now,
        updated_at=now
    )

@api_router.get("/doctors", response_model=List[DoctorResponse])
async def get_doctors(
    search: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'customer_code': {'$regex': search, '$options': 'i'}},
            {'email': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}},
            {'reg_no': {'$regex': search, '$options': 'i'}}
        ]
    
    if status and status != 'all':
        query['lead_status'] = status
    
    doctors = await db.doctors.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    
    # Get all transport IDs to fetch names in one query
    transport_ids = [doc.get('transport_id') for doc in doctors if doc.get('transport_id')]
    transport_map = {}
    if transport_ids:
        transports = await db.transports.find({'id': {'$in': transport_ids}}, {'_id': 0, 'id': 1, 'name': 1}).to_list(100)
        transport_map = {t['id']: t['name'] for t in transports}
    
    result = []
    for doc in doctors:
        created_at = doc.get('created_at')
        updated_at = doc.get('updated_at') or doc.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
        
        result.append(DoctorResponse(
            id=doc['id'],
            customer_code=doc['customer_code'],
            name=doc['name'],
            reg_no=doc.get('reg_no', ''),
            address=doc.get('address', ''),
            address_line_1=doc.get('address_line_1'),
            address_line_2=doc.get('address_line_2'),
            district=doc.get('district'),
            state=doc.get('state'),
            pincode=doc.get('pincode'),
            delivery_station=doc.get('delivery_station'),
            transport_id=doc.get('transport_id'),
            transport_name=transport_map.get(doc.get('transport_id')),
            email=doc.get('email', ''),
            phone=doc['phone'],
            lead_status=doc.get('lead_status', 'Pipeline'),
            dob=doc.get('dob'),
            priority=doc.get('priority'),
            last_contact_date=doc.get('last_contact_date'),
            follow_up_date=doc.get('follow_up_date'),
            is_portal_customer=doc.get('is_portal_customer', False),
            portal_customer_id=doc.get('portal_customer_id'),
            created_at=created_at,
            updated_at=updated_at
        ))
    
    return result

@api_router.get("/doctors/{doctor_id}", response_model=DoctorResponse)
async def get_doctor(doctor_id: str, current_user: dict = Depends(get_current_user)):
    doctor = await db.doctors.find_one({'id': doctor_id}, {'_id': 0})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    # Get transport name if transport_id exists
    transport_name = None
    if doctor.get('transport_id'):
        transport = await db.transports.find_one({'id': doctor['transport_id']}, {'_id': 0, 'name': 1})
        transport_name = transport.get('name') if transport else None
    
    created_at = doctor.get('created_at')
    updated_at = doctor.get('updated_at') or doctor.get('created_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    return DoctorResponse(
        id=doctor['id'],
        customer_code=doctor['customer_code'],
        name=doctor['name'],
        reg_no=doctor.get('reg_no', ''),
        address=doctor.get('address', ''),
        address_line_1=doctor.get('address_line_1'),
        address_line_2=doctor.get('address_line_2'),
        district=doctor.get('district'),
        state=doctor.get('state'),
        pincode=doctor.get('pincode'),
        delivery_station=doctor.get('delivery_station'),
        transport_id=doctor.get('transport_id'),
        transport_name=transport_name,
        email=doctor.get('email', ''),
        phone=doctor['phone'],
        lead_status=doctor.get('lead_status', 'Pipeline'),
        dob=doctor.get('dob'),
        priority=doctor.get('priority'),
        last_contact_date=doctor.get('last_contact_date'),
        follow_up_date=doctor.get('follow_up_date'),
        is_portal_customer=doctor.get('is_portal_customer', False),
        portal_customer_id=doctor.get('portal_customer_id'),
        created_at=created_at,
        updated_at=updated_at
    )

@api_router.put("/doctors/{doctor_id}", response_model=DoctorResponse)
async def update_doctor(doctor_id: str, doctor_data: DoctorUpdate, current_user: dict = Depends(get_current_user)):
    doctor = await db.doctors.find_one({'id': doctor_id}, {'_id': 0})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    update_data = {k: v for k, v in doctor_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.doctors.update_one({'id': doctor_id}, {'$set': update_data})
    
    updated_doctor = await db.doctors.find_one({'id': doctor_id}, {'_id': 0})
    
    # Get transport name if transport_id exists
    transport_name = None
    if updated_doctor.get('transport_id'):
        transport = await db.transports.find_one({'id': updated_doctor['transport_id']}, {'_id': 0, 'name': 1})
        transport_name = transport.get('name') if transport else None
    
    created_at = updated_doctor.get('created_at')
    updated_at = updated_doctor.get('updated_at') or updated_doctor.get('created_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    return DoctorResponse(
        id=updated_doctor['id'],
        customer_code=updated_doctor['customer_code'],
        name=updated_doctor['name'],
        reg_no=updated_doctor.get('reg_no', ''),
        address=updated_doctor.get('address', ''),
        address_line_1=updated_doctor.get('address_line_1'),
        address_line_2=updated_doctor.get('address_line_2'),
        district=updated_doctor.get('district'),
        state=updated_doctor.get('state'),
        pincode=updated_doctor.get('pincode'),
        delivery_station=updated_doctor.get('delivery_station'),
        transport_id=updated_doctor.get('transport_id'),
        transport_name=transport_name,
        email=updated_doctor.get('email', ''),
        phone=updated_doctor['phone'],
        lead_status=updated_doctor.get('lead_status', 'Pipeline'),
        dob=updated_doctor.get('dob'),
        priority=updated_doctor.get('priority'),
        last_contact_date=updated_doctor.get('last_contact_date'),
        follow_up_date=updated_doctor.get('follow_up_date'),
        is_portal_customer=updated_doctor.get('is_portal_customer', False),
        portal_customer_id=updated_doctor.get('portal_customer_id'),
        created_at=created_at,
        updated_at=updated_at
    )

@api_router.delete("/doctors/{doctor_id}")
async def delete_doctor(doctor_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.doctors.delete_one({'id': doctor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Doctor not found")
    # Also delete related notes and tasks
    await db.doctor_notes.delete_many({'doctor_id': doctor_id})
    await db.tasks.delete_many({'doctor_id': doctor_id})
    return {"message": "Doctor deleted successfully"}

@api_router.post("/doctors/bulk-delete")
async def bulk_delete_doctors(ids: List[str], current_user: dict = Depends(get_current_user)):
    """Bulk delete multiple doctors"""
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    
    # Delete doctors
    result = await db.doctors.delete_many({'id': {'$in': ids}})
    
    # Also delete related notes and tasks
    await db.doctor_notes.delete_many({'doctor_id': {'$in': ids}})
    await db.tasks.delete_many({'doctor_id': {'$in': ids}})
    
    return {"message": f"{result.deleted_count} doctor(s) deleted successfully", "deleted_count": result.deleted_count}

# ============== DOCTOR NOTES ROUTES ==============

@api_router.get("/doctors/{doctor_id}/notes", response_model=List[DoctorNoteResponse])
async def get_doctor_notes(doctor_id: str, current_user: dict = Depends(get_current_user)):
    """Get all notes for a doctor"""
    notes = await db.doctor_notes.find({'doctor_id': doctor_id}, {'_id': 0}).sort('created_at', -1).to_list(100)
    
    result = []
    for note in notes:
        created_at = note['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append(DoctorNoteResponse(
            id=note['id'],
            doctor_id=note['doctor_id'],
            note=note['note'],
            created_by=note['created_by'],
            created_at=created_at
        ))
    
    return result

@api_router.post("/doctors/{doctor_id}/notes", response_model=DoctorNoteResponse)
async def add_doctor_note(doctor_id: str, note_data: DoctorNoteCreate, current_user: dict = Depends(get_current_user)):
    """Add a note to a doctor"""
    doctor = await db.doctors.find_one({'id': doctor_id}, {'_id': 0})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    now = datetime.now(timezone.utc)
    note_doc = {
        'id': str(uuid.uuid4()),
        'doctor_id': doctor_id,
        'note': note_data.note,
        'created_by': current_user['name'],
        'created_at': now.isoformat()
    }
    
    await db.doctor_notes.insert_one(note_doc)
    
    return DoctorNoteResponse(
        id=note_doc['id'],
        doctor_id=doctor_id,
        note=note_doc['note'],
        created_by=note_doc['created_by'],
        created_at=now
    )

@api_router.delete("/doctors/{doctor_id}/notes/{note_id}")
async def delete_doctor_note(doctor_id: str, note_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a note"""
    result = await db.doctor_notes.delete_one({'id': note_id, 'doctor_id': doctor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    return {"message": "Note deleted successfully"}

# ============== MEDICAL ROUTES ==============

@api_router.post("/medicals", response_model=MedicalResponse)
async def create_medical(medical_data: MedicalCreate, current_user: dict = Depends(get_current_user)):
    medical_id = str(uuid.uuid4())
    customer_code = await generate_medical_code()
    now = datetime.now(timezone.utc)
    
    # Get transport name if transport_id provided
    transport_name = None
    if medical_data.transport_id:
        transport = await db.transports.find_one({'id': medical_data.transport_id}, {'_id': 0, 'name': 1})
        transport_name = transport.get('name') if transport else None
    
    medical_doc = {
        'id': medical_id,
        'customer_code': customer_code,
        'name': medical_data.name,
        'proprietor_name': medical_data.proprietor_name,
        'gst_number': medical_data.gst_number,
        'drug_license': medical_data.drug_license,
        'address': medical_data.address,
        'address_line_1': medical_data.address_line_1,
        'address_line_2': medical_data.address_line_2,
        'state': medical_data.state,
        'district': medical_data.district,
        'pincode': medical_data.pincode,
        'delivery_station': medical_data.delivery_station,
        'transport_id': medical_data.transport_id,
        'email': medical_data.email,
        'phone': medical_data.phone,
        'alternate_phone': medical_data.alternate_phone,
        'lead_status': medical_data.lead_status,
        'birthday': medical_data.birthday,
        'anniversary': medical_data.anniversary,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': current_user['id']
    }
    
    await db.medicals.insert_one(medical_doc)
    
    return MedicalResponse(
        id=medical_id,
        customer_code=customer_code,
        name=medical_data.name,
        proprietor_name=medical_data.proprietor_name,
        gst_number=medical_data.gst_number,
        drug_license=medical_data.drug_license,
        address=medical_data.address,
        address_line_1=medical_data.address_line_1,
        address_line_2=medical_data.address_line_2,
        state=medical_data.state,
        district=medical_data.district,
        pincode=medical_data.pincode,
        delivery_station=medical_data.delivery_station,
        transport_id=medical_data.transport_id,
        transport_name=transport_name,
        email=medical_data.email,
        phone=medical_data.phone,
        alternate_phone=medical_data.alternate_phone,
        lead_status=medical_data.lead_status,
        birthday=medical_data.birthday,
        anniversary=medical_data.anniversary,
        created_at=now,
        updated_at=now
    )

@api_router.get("/medicals", response_model=List[MedicalResponse])
async def get_medicals(
    search: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'customer_code': {'$regex': search, '$options': 'i'}},
            {'proprietor_name': {'$regex': search, '$options': 'i'}},
            {'email': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}},
            {'gst_number': {'$regex': search, '$options': 'i'}}
        ]
    
    if status and status != 'all':
        query['lead_status'] = status
    
    medicals = await db.medicals.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    
    # Get all transport IDs to fetch names in one query
    transport_ids = [doc.get('transport_id') for doc in medicals if doc.get('transport_id')]
    transport_map = {}
    if transport_ids:
        transports = await db.transports.find({'id': {'$in': transport_ids}}, {'_id': 0, 'id': 1, 'name': 1}).to_list(100)
        transport_map = {t['id']: t['name'] for t in transports}
    
    result = []
    for doc in medicals:
        created_at = doc.get('created_at')
        updated_at = doc.get('updated_at') or doc.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
        
        result.append(MedicalResponse(
            id=doc['id'],
            customer_code=doc['customer_code'],
            name=doc['name'],
            proprietor_name=doc.get('proprietor_name'),
            gst_number=doc.get('gst_number'),
            drug_license=doc.get('drug_license'),
            address=doc.get('address'),
            address_line_1=doc.get('address_line_1'),
            address_line_2=doc.get('address_line_2'),
            state=doc.get('state'),
            district=doc.get('district'),
            pincode=doc.get('pincode'),
            delivery_station=doc.get('delivery_station'),
            transport_id=doc.get('transport_id'),
            transport_name=transport_map.get(doc.get('transport_id')),
            email=doc.get('email'),
            phone=doc['phone'],
            alternate_phone=doc.get('alternate_phone'),
            lead_status=doc.get('lead_status', 'Pipeline'),
            priority=doc.get('priority'),
            last_contact_date=doc.get('last_contact_date'),
            follow_up_date=doc.get('follow_up_date'),
            birthday=doc.get('birthday'),
            anniversary=doc.get('anniversary'),
            is_portal_customer=doc.get('is_portal_customer', False),
            portal_customer_id=doc.get('portal_customer_id'),
            created_at=created_at,
            updated_at=updated_at
        ))
    
    return result

@api_router.get("/medicals/{medical_id}", response_model=MedicalResponse)
async def get_medical(medical_id: str, current_user: dict = Depends(get_current_user)):
    medical = await db.medicals.find_one({'id': medical_id}, {'_id': 0})
    if not medical:
        raise HTTPException(status_code=404, detail="Medical not found")
    
    created_at = medical.get('created_at')
    updated_at = medical.get('updated_at') or medical.get('created_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    return MedicalResponse(
        id=medical['id'],
        customer_code=medical['customer_code'],
        name=medical['name'],
        proprietor_name=medical.get('proprietor_name'),
        gst_number=medical.get('gst_number'),
        drug_license=medical.get('drug_license'),
        address=medical.get('address'),
        address_line_1=medical.get('address_line_1'),
        address_line_2=medical.get('address_line_2'),
        state=medical.get('state'),
        district=medical.get('district'),
        pincode=medical.get('pincode'),
        delivery_station=medical.get('delivery_station'),
        transport_id=medical.get('transport_id'),
        transport_name=None,
        email=medical.get('email'),
        phone=medical['phone'],
        alternate_phone=medical.get('alternate_phone'),
        lead_status=medical.get('lead_status', 'Pipeline'),
        priority=medical.get('priority'),
        last_contact_date=medical.get('last_contact_date'),
        follow_up_date=medical.get('follow_up_date'),
        birthday=medical.get('birthday'),
        anniversary=medical.get('anniversary'),
        is_portal_customer=medical.get('is_portal_customer', False),
        portal_customer_id=medical.get('portal_customer_id'),
        created_at=created_at,
        updated_at=updated_at
    )

@api_router.put("/medicals/{medical_id}", response_model=MedicalResponse)
async def update_medical(medical_id: str, medical_data: MedicalUpdate, current_user: dict = Depends(get_current_user)):
    medical = await db.medicals.find_one({'id': medical_id}, {'_id': 0})
    if not medical:
        raise HTTPException(status_code=404, detail="Medical not found")
    
    update_data = {k: v for k, v in medical_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.medicals.update_one({'id': medical_id}, {'$set': update_data})
    
    updated = await db.medicals.find_one({'id': medical_id}, {'_id': 0})
    created_at = updated.get('created_at')
    updated_at = updated.get('updated_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    return MedicalResponse(
        id=updated['id'],
        customer_code=updated['customer_code'],
        name=updated['name'],
        proprietor_name=updated.get('proprietor_name'),
        gst_number=updated.get('gst_number'),
        drug_license=updated.get('drug_license'),
        address=updated.get('address'),
        address_line_1=updated.get('address_line_1'),
        address_line_2=updated.get('address_line_2'),
        state=updated.get('state'),
        district=updated.get('district'),
        pincode=updated.get('pincode'),
        delivery_station=updated.get('delivery_station'),
        transport_id=updated.get('transport_id'),
        transport_name=None,
        email=updated.get('email'),
        phone=updated['phone'],
        alternate_phone=updated.get('alternate_phone'),
        lead_status=updated.get('lead_status', 'Pipeline'),
        priority=updated.get('priority'),
        last_contact_date=updated.get('last_contact_date'),
        follow_up_date=updated.get('follow_up_date'),
        birthday=updated.get('birthday'),
        anniversary=updated.get('anniversary'),
        is_portal_customer=updated.get('is_portal_customer', False),
        portal_customer_id=updated.get('portal_customer_id'),
        created_at=created_at,
        updated_at=updated_at
    )

@api_router.delete("/medicals/{medical_id}")
async def delete_medical(medical_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.medicals.delete_one({'id': medical_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Medical not found")
    # Also delete related notes
    await db.medical_notes.delete_many({'medical_id': medical_id})
    return {"message": "Medical deleted successfully"}

@api_router.post("/medicals/bulk-delete")
async def bulk_delete_medicals(ids: List[str], current_user: dict = Depends(get_current_user)):
    """Bulk delete multiple medicals"""
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    
    # Delete medicals
    result = await db.medicals.delete_many({'id': {'$in': ids}})
    
    # Also delete related notes
    await db.medical_notes.delete_many({'medical_id': {'$in': ids}})
    
    return {"message": f"{result.deleted_count} medical(s) deleted successfully", "deleted_count": result.deleted_count}

@api_router.put("/medicals/{medical_id}/contact")
async def update_medical_last_contact(medical_id: str, current_user: dict = Depends(get_current_user)):
    """Update last contact date and set follow-up to 25 days from now"""
    medical = await db.medicals.find_one({'id': medical_id}, {'_id': 0})
    if not medical:
        raise HTTPException(status_code=404, detail="Medical not found")
    
    now = datetime.now(timezone.utc)
    follow_up = now + timedelta(days=25)
    
    await db.medicals.update_one(
        {'id': medical_id},
        {'$set': {
            'last_contact_date': now.strftime('%Y-%m-%d'),
            'follow_up_date': follow_up.strftime('%Y-%m-%d'),
            'updated_at': now.isoformat()
        }}
    )
    
    return {"message": "Contact updated successfully", "follow_up_date": follow_up.strftime('%Y-%m-%d')}

# Medical Notes
@api_router.get("/medicals/{medical_id}/notes", response_model=List[MedicalNoteResponse])
async def get_medical_notes(medical_id: str, current_user: dict = Depends(get_current_user)):
    notes = await db.medical_notes.find({'medical_id': medical_id}, {'_id': 0}).sort('created_at', -1).to_list(100)
    
    result = []
    for note in notes:
        created_at = note['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append(MedicalNoteResponse(
            id=note['id'],
            medical_id=note['medical_id'],
            note=note['note'],
            created_by=note['created_by'],
            created_at=created_at
        ))
    
    return result

@api_router.post("/medicals/{medical_id}/notes", response_model=MedicalNoteResponse)
async def add_medical_note(medical_id: str, note_data: MedicalNoteCreate, current_user: dict = Depends(get_current_user)):
    medical = await db.medicals.find_one({'id': medical_id}, {'_id': 0})
    if not medical:
        raise HTTPException(status_code=404, detail="Medical not found")
    
    now = datetime.now(timezone.utc)
    note_doc = {
        'id': str(uuid.uuid4()),
        'medical_id': medical_id,
        'note': note_data.note,
        'created_by': current_user['name'],
        'created_at': now.isoformat()
    }
    
    await db.medical_notes.insert_one(note_doc)
    
    return MedicalNoteResponse(
        id=note_doc['id'],
        medical_id=medical_id,
        note=note_doc['note'],
        created_by=note_doc['created_by'],
        created_at=now
    )

@api_router.delete("/medicals/{medical_id}/notes/{note_id}")
async def delete_medical_note(medical_id: str, note_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.medical_notes.delete_one({'id': note_id, 'medical_id': medical_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    return {"message": "Note deleted successfully"}

# ============== AGENCY ROUTES ==============

@api_router.post("/agencies", response_model=AgencyResponse)
async def create_agency(agency_data: AgencyCreate, current_user: dict = Depends(get_current_user)):
    agency_id = str(uuid.uuid4())
    customer_code = await generate_agency_code()
    now = datetime.now(timezone.utc)
    
    agency_doc = {
        'id': agency_id,
        'customer_code': customer_code,
        'name': agency_data.name,
        'proprietor_name': agency_data.proprietor_name,
        'gst_number': agency_data.gst_number,
        'drug_license': agency_data.drug_license,
        'address': agency_data.address,
        'address_line_1': agency_data.address_line_1,
        'address_line_2': agency_data.address_line_2,
        'state': agency_data.state,
        'district': agency_data.district,
        'pincode': agency_data.pincode,
        'delivery_station': agency_data.delivery_station,
        'transport_id': agency_data.transport_id,
        'email': agency_data.email,
        'phone': agency_data.phone,
        'alternate_phone': agency_data.alternate_phone,
        'lead_status': agency_data.lead_status,
        'birthday': agency_data.birthday,
        'anniversary': agency_data.anniversary,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': current_user['id']
    }
    
    await db.agencies.insert_one(agency_doc)
    
    # Get transport name if transport_id is provided
    transport_name = None
    if agency_data.transport_id:
        transport = await db.transports.find_one({'id': agency_data.transport_id})
        transport_name = transport['name'] if transport else None
    
    return AgencyResponse(
        id=agency_id,
        customer_code=customer_code,
        name=agency_data.name,
        proprietor_name=agency_data.proprietor_name,
        gst_number=agency_data.gst_number,
        drug_license=agency_data.drug_license,
        address=agency_data.address,
        address_line_1=agency_data.address_line_1,
        address_line_2=agency_data.address_line_2,
        state=agency_data.state,
        district=agency_data.district,
        pincode=agency_data.pincode,
        delivery_station=agency_data.delivery_station,
        transport_id=agency_data.transport_id,
        transport_name=transport_name,
        email=agency_data.email,
        phone=agency_data.phone,
        alternate_phone=agency_data.alternate_phone,
        lead_status=agency_data.lead_status,
        birthday=agency_data.birthday,
        anniversary=agency_data.anniversary,
        created_at=now,
        updated_at=now
    )

@api_router.get("/agencies", response_model=List[AgencyResponse])
async def get_agencies(
    search: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'customer_code': {'$regex': search, '$options': 'i'}},
            {'proprietor_name': {'$regex': search, '$options': 'i'}},
            {'email': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}},
            {'gst_number': {'$regex': search, '$options': 'i'}}
        ]
    
    if status and status != 'all':
        query['lead_status'] = status
    
    agencies = await db.agencies.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    
    # Get all transport ids to lookup names
    transport_ids = list(set([doc.get('transport_id') for doc in agencies if doc.get('transport_id')]))
    transports_map = {}
    if transport_ids:
        transports = await db.transports.find({'id': {'$in': transport_ids}}).to_list(100)
        transports_map = {t['id']: t['name'] for t in transports}
    
    result = []
    for doc in agencies:
        created_at = doc.get('created_at')
        updated_at = doc.get('updated_at') or doc.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
        
        result.append(AgencyResponse(
            id=doc['id'],
            customer_code=doc['customer_code'],
            name=doc['name'],
            proprietor_name=doc.get('proprietor_name'),
            gst_number=doc.get('gst_number'),
            drug_license=doc.get('drug_license'),
            address=doc.get('address'),
            address_line_1=doc.get('address_line_1'),
            address_line_2=doc.get('address_line_2'),
            state=doc.get('state'),
            district=doc.get('district'),
            pincode=doc.get('pincode'),
            delivery_station=doc.get('delivery_station'),
            transport_id=doc.get('transport_id'),
            transport_name=transports_map.get(doc.get('transport_id')),
            email=doc.get('email'),
            phone=doc['phone'],
            alternate_phone=doc.get('alternate_phone'),
            lead_status=doc.get('lead_status', 'Pipeline'),
            priority=doc.get('priority'),
            last_contact_date=doc.get('last_contact_date'),
            follow_up_date=doc.get('follow_up_date'),
            birthday=doc.get('birthday'),
            anniversary=doc.get('anniversary'),
            is_portal_customer=doc.get('is_portal_customer', False),
            portal_customer_id=doc.get('portal_customer_id'),
            created_at=created_at,
            updated_at=updated_at
        ))
    
    return result

@api_router.get("/agencies/{agency_id}", response_model=AgencyResponse)
async def get_agency(agency_id: str, current_user: dict = Depends(get_current_user)):
    agency = await db.agencies.find_one({'id': agency_id}, {'_id': 0})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    
    created_at = agency.get('created_at')
    updated_at = agency.get('updated_at') or agency.get('created_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    # Get transport name if transport_id exists
    transport_name = None
    if agency.get('transport_id'):
        transport = await db.transports.find_one({'id': agency['transport_id']})
        transport_name = transport['name'] if transport else None
    
    return AgencyResponse(
        id=agency['id'],
        customer_code=agency['customer_code'],
        name=agency['name'],
        proprietor_name=agency.get('proprietor_name'),
        gst_number=agency.get('gst_number'),
        drug_license=agency.get('drug_license'),
        address=agency.get('address'),
        address_line_1=agency.get('address_line_1'),
        address_line_2=agency.get('address_line_2'),
        state=agency.get('state'),
        district=agency.get('district'),
        pincode=agency.get('pincode'),
        delivery_station=agency.get('delivery_station'),
        transport_id=agency.get('transport_id'),
        transport_name=transport_name,
        email=agency.get('email'),
        phone=agency['phone'],
        alternate_phone=agency.get('alternate_phone'),
        lead_status=agency.get('lead_status', 'Pipeline'),
        priority=agency.get('priority'),
        last_contact_date=agency.get('last_contact_date'),
        follow_up_date=agency.get('follow_up_date'),
        birthday=agency.get('birthday'),
        anniversary=agency.get('anniversary'),
        is_portal_customer=agency.get('is_portal_customer', False),
        portal_customer_id=agency.get('portal_customer_id'),
        created_at=created_at,
        updated_at=updated_at
    )

@api_router.put("/agencies/{agency_id}", response_model=AgencyResponse)
async def update_agency(agency_id: str, agency_data: AgencyUpdate, current_user: dict = Depends(get_current_user)):
    agency = await db.agencies.find_one({'id': agency_id}, {'_id': 0})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    
    update_data = {k: v for k, v in agency_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.agencies.update_one({'id': agency_id}, {'$set': update_data})
    
    updated = await db.agencies.find_one({'id': agency_id}, {'_id': 0})
    created_at = updated.get('created_at')
    updated_at = updated.get('updated_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    # Get transport name if transport_id exists
    transport_name = None
    if updated.get('transport_id'):
        transport = await db.transports.find_one({'id': updated['transport_id']})
        transport_name = transport['name'] if transport else None
    
    return AgencyResponse(
        id=updated['id'],
        customer_code=updated['customer_code'],
        name=updated['name'],
        proprietor_name=updated.get('proprietor_name'),
        gst_number=updated.get('gst_number'),
        drug_license=updated.get('drug_license'),
        address=updated.get('address'),
        address_line_1=updated.get('address_line_1'),
        address_line_2=updated.get('address_line_2'),
        state=updated.get('state'),
        district=updated.get('district'),
        pincode=updated.get('pincode'),
        delivery_station=updated.get('delivery_station'),
        transport_id=updated.get('transport_id'),
        transport_name=transport_name,
        email=updated.get('email'),
        phone=updated['phone'],
        alternate_phone=updated.get('alternate_phone'),
        lead_status=updated.get('lead_status', 'Pipeline'),
        priority=updated.get('priority'),
        last_contact_date=updated.get('last_contact_date'),
        follow_up_date=updated.get('follow_up_date'),
        birthday=updated.get('birthday'),
        anniversary=updated.get('anniversary'),
        is_portal_customer=updated.get('is_portal_customer', False),
        portal_customer_id=updated.get('portal_customer_id'),
        created_at=created_at,
        updated_at=updated_at
    )

@api_router.delete("/agencies/{agency_id}")
async def delete_agency(agency_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.agencies.delete_one({'id': agency_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agency not found")
    # Also delete related notes
    await db.agency_notes.delete_many({'agency_id': agency_id})
    return {"message": "Agency deleted successfully"}

@api_router.post("/agencies/bulk-delete")
async def bulk_delete_agencies(ids: List[str], current_user: dict = Depends(get_current_user)):
    """Bulk delete multiple agencies"""
    if not ids:
        raise HTTPException(status_code=400, detail="No IDs provided")
    
    # Delete agencies
    result = await db.agencies.delete_many({'id': {'$in': ids}})
    
    # Also delete related notes
    await db.agency_notes.delete_many({'agency_id': {'$in': ids}})
    
    return {"message": f"{result.deleted_count} agency(ies) deleted successfully", "deleted_count": result.deleted_count}

@api_router.put("/agencies/{agency_id}/contact")
async def update_agency_last_contact(agency_id: str, current_user: dict = Depends(get_current_user)):
    """Update last contact date and set follow-up to 25 days from now"""
    agency = await db.agencies.find_one({'id': agency_id}, {'_id': 0})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    
    now = datetime.now(timezone.utc)
    follow_up = now + timedelta(days=25)
    
    await db.agencies.update_one(
        {'id': agency_id},
        {'$set': {
            'last_contact_date': now.strftime('%Y-%m-%d'),
            'follow_up_date': follow_up.strftime('%Y-%m-%d'),
            'updated_at': now.isoformat()
        }}
    )
    
    return {"message": "Contact updated successfully", "follow_up_date": follow_up.strftime('%Y-%m-%d')}

# Agency Notes
@api_router.get("/agencies/{agency_id}/notes", response_model=List[AgencyNoteResponse])
async def get_agency_notes(agency_id: str, current_user: dict = Depends(get_current_user)):
    notes = await db.agency_notes.find({'agency_id': agency_id}, {'_id': 0}).sort('created_at', -1).to_list(100)
    
    result = []
    for note in notes:
        created_at = note['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append(AgencyNoteResponse(
            id=note['id'],
            agency_id=note['agency_id'],
            note=note['note'],
            created_by=note['created_by'],
            created_at=created_at
        ))
    
    return result

@api_router.post("/agencies/{agency_id}/notes", response_model=AgencyNoteResponse)
async def add_agency_note(agency_id: str, note_data: AgencyNoteCreate, current_user: dict = Depends(get_current_user)):
    agency = await db.agencies.find_one({'id': agency_id}, {'_id': 0})
    if not agency:
        raise HTTPException(status_code=404, detail="Agency not found")
    
    now = datetime.now(timezone.utc)
    note_doc = {
        'id': str(uuid.uuid4()),
        'agency_id': agency_id,
        'note': note_data.note,
        'created_by': current_user['name'],
        'created_at': now.isoformat()
    }
    
    await db.agency_notes.insert_one(note_doc)
    
    return AgencyNoteResponse(
        id=note_doc['id'],
        agency_id=agency_id,
        note=note_doc['note'],
        created_by=note_doc['created_by'],
        created_at=now
    )

@api_router.delete("/agencies/{agency_id}/notes/{note_id}")
async def delete_agency_note(agency_id: str, note_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.agency_notes.delete_one({'id': note_id, 'agency_id': agency_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    return {"message": "Note deleted successfully"}

# ============== TASKS ROUTES ==============

@api_router.get("/tasks", response_model=List[TaskResponse])
async def get_all_tasks(
    status: Optional[str] = None, 
    doctor_id: Optional[str] = None, 
    medical_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all tasks with optional filters"""
    query = {}
    if status:
        query['status'] = status
    if doctor_id:
        query['doctor_id'] = doctor_id
    if medical_id:
        query['medical_id'] = medical_id
    if agency_id:
        query['agency_id'] = agency_id
    
    tasks = await db.tasks.find(query, {'_id': 0}).sort('due_date', 1).to_list(1000)
    
    result = []
    for task in tasks:
        created_at = task['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        # Get entity name (doctor, medical, or agency)
        entity_name = None
        if task.get('doctor_id'):
            entity = await db.doctors.find_one({'id': task['doctor_id']}, {'_id': 0, 'name': 1})
            entity_name = entity['name'] if entity else None
        elif task.get('medical_id'):
            entity = await db.medicals.find_one({'id': task['medical_id']}, {'_id': 0, 'name': 1})
            entity_name = entity['name'] if entity else None
        elif task.get('agency_id'):
            entity = await db.agencies.find_one({'id': task['agency_id']}, {'_id': 0, 'name': 1})
            entity_name = entity['name'] if entity else None
        
        result.append(TaskResponse(
            id=task['id'],
            doctor_id=task.get('doctor_id', ''),
            doctor_name=entity_name,
            title=task['title'],
            description=task.get('description'),
            due_date=task.get('due_date'),
            priority=task.get('priority', 'moderate'),
            status=task.get('status', 'pending'),
            created_at=created_at
        ))
    
    return result

@api_router.post("/tasks", response_model=TaskResponse)
async def create_task(task_data: TaskCreate, current_user: dict = Depends(get_current_user)):
    """Create a new task"""
    entity_name = None
    entity_id = None
    entity_type = None
    
    if task_data.doctor_id:
        entity = await db.doctors.find_one({'id': task_data.doctor_id}, {'_id': 0})
        if not entity:
            raise HTTPException(status_code=404, detail="Doctor not found")
        entity_name = entity['name']
        entity_id = task_data.doctor_id
        entity_type = 'doctor_id'
    elif task_data.medical_id:
        entity = await db.medicals.find_one({'id': task_data.medical_id}, {'_id': 0})
        if not entity:
            raise HTTPException(status_code=404, detail="Medical not found")
        entity_name = entity['name']
        entity_id = task_data.medical_id
        entity_type = 'medical_id'
    elif task_data.agency_id:
        entity = await db.agencies.find_one({'id': task_data.agency_id}, {'_id': 0})
        if not entity:
            raise HTTPException(status_code=404, detail="Agency not found")
        entity_name = entity['name']
        entity_id = task_data.agency_id
        entity_type = 'agency_id'
    else:
        raise HTTPException(status_code=400, detail="Either doctor_id, medical_id, or agency_id is required")
    
    now = datetime.now(timezone.utc)
    task_doc = {
        'id': str(uuid.uuid4()),
        entity_type: entity_id,
        'title': task_data.title,
        'description': task_data.description,
        'due_date': task_data.due_date,
        'priority': task_data.priority or 'moderate',
        'status': 'pending',
        'created_at': now.isoformat()
    }
    
    await db.tasks.insert_one(task_doc)
    
    return TaskResponse(
        id=task_doc['id'],
        doctor_id=task_doc.get('doctor_id', ''),
        doctor_name=entity_name,
        title=task_doc['title'],
        description=task_doc['description'],
        due_date=task_doc['due_date'],
        priority=task_doc['priority'],
        status=task_doc['status'],
        created_at=now
    )

@api_router.put("/tasks/{task_id}", response_model=TaskResponse)
async def update_task(task_id: str, task_data: TaskUpdate, current_user: dict = Depends(get_current_user)):
    """Update a task"""
    task = await db.tasks.find_one({'id': task_id}, {'_id': 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    update_data = {k: v for k, v in task_data.model_dump().items() if v is not None}
    
    if update_data:
        await db.tasks.update_one({'id': task_id}, {'$set': update_data})
    
    updated_task = await db.tasks.find_one({'id': task_id}, {'_id': 0})
    
    # Find entity name - check doctor_id, medical_id, or agency_id
    entity_name = None
    entity_id = updated_task.get('doctor_id') or updated_task.get('medical_id') or updated_task.get('agency_id') or ''
    if updated_task.get('doctor_id'):
        entity = await db.doctors.find_one({'id': updated_task['doctor_id']}, {'_id': 0, 'name': 1})
        entity_name = entity['name'] if entity else None
    elif updated_task.get('medical_id'):
        entity = await db.medicals.find_one({'id': updated_task['medical_id']}, {'_id': 0, 'name': 1})
        entity_name = entity['name'] if entity else None
    elif updated_task.get('agency_id'):
        entity = await db.agencies.find_one({'id': updated_task['agency_id']}, {'_id': 0, 'name': 1})
        entity_name = entity['name'] if entity else None
    
    created_at = updated_task['created_at']
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    
    return TaskResponse(
        id=updated_task['id'],
        doctor_id=entity_id,
        doctor_name=entity_name,
        title=updated_task['title'],
        description=updated_task.get('description'),
        due_date=updated_task.get('due_date'),
        priority=updated_task.get('priority', 'moderate'),
        status=updated_task.get('status', 'pending'),
        created_at=created_at
    )

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a task"""
    result = await db.tasks.delete_one({'id': task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted successfully"}

# ============== FOLLOW-UP REMINDERS ==============

@api_router.get("/reminders")
async def get_follow_up_reminders(current_user: dict = Depends(get_current_user)):
    """Get leads that need follow-up (based on 25 days rule or set follow-up date)"""
    today = datetime.now(timezone.utc).date()
    today_str = today.isoformat()
    
    # Calculate 25 days ago
    days_25_ago = (today - timedelta(days=25)).isoformat()
    
    # Get all doctors except "Not Interested" and "Closed"
    doctors = await db.doctors.find(
        {'lead_status': {'$nin': ['Not Interested', 'Closed']}},
        {'_id': 0}
    ).to_list(1000)
    
    reminders = []
    
    for doc in doctors:
        needs_follow_up = False
        follow_up_reason = ""
        days_overdue = 0
        
        follow_up_date = doc.get('follow_up_date')
        last_contact = doc.get('last_contact_date')
        
        # Check if follow-up date is set and due
        if follow_up_date:
            if follow_up_date <= today_str:
                needs_follow_up = True
                follow_up_reason = "Scheduled follow-up"
                try:
                    fu_date = datetime.fromisoformat(follow_up_date).date()
                    days_overdue = (today - fu_date).days
                except:
                    days_overdue = 0
        # If no follow-up date, check 25 days rule from last contact
        elif last_contact:
            if last_contact <= days_25_ago:
                needs_follow_up = True
                follow_up_reason = "25 days since last contact"
                try:
                    lc_date = datetime.fromisoformat(last_contact).date()
                    days_overdue = (today - lc_date).days - 25
                except:
                    days_overdue = 0
        # If no last contact and no follow-up, use created_at
        else:
            created_at = doc.get('created_at', '')
            if isinstance(created_at, str) and created_at:
                try:
                    created_date = datetime.fromisoformat(created_at.replace('Z', '+00:00')).date()
                    if created_date <= datetime.fromisoformat(days_25_ago).date():
                        needs_follow_up = True
                        follow_up_reason = "25 days since added (no contact)"
                        days_overdue = (today - created_date).days - 25
                except:
                    pass
        
        if needs_follow_up:
            reminders.append({
                'doctor_id': doc['id'],
                'doctor_name': doc['name'],
                'customer_code': doc['customer_code'],
                'phone': doc['phone'],
                'lead_status': doc['lead_status'],
                'priority': doc.get('priority', 'moderate'),
                'last_contact_date': last_contact,
                'follow_up_date': follow_up_date,
                'reason': follow_up_reason,
                'days_overdue': max(0, days_overdue)
            })
    
    # Sort by priority (high first) then by days overdue
    priority_order = {'high': 0, 'moderate': 1, 'low': 2}
    reminders.sort(key=lambda x: (priority_order.get(x['priority'], 1), -x['days_overdue']))
    
    return reminders

@api_router.put("/doctors/{doctor_id}/contact")
async def update_last_contact(doctor_id: str, current_user: dict = Depends(get_current_user)):
    """Update last contact date to today and auto-set follow-up to 25 days"""
    doctor = await db.doctors.find_one({'id': doctor_id}, {'_id': 0})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    today = datetime.now(timezone.utc).date()
    follow_up = today + timedelta(days=25)
    
    # Only auto-set follow-up if status is not "Not Interested"
    update_data = {
        'last_contact_date': today.isoformat(),
        'updated_at': datetime.now(timezone.utc).isoformat()
    }
    
    if doctor.get('lead_status') != 'Not Interested':
        update_data['follow_up_date'] = follow_up.isoformat()
    
    await db.doctors.update_one({'id': doctor_id}, {'$set': update_data})
    
    return {
        "message": "Last contact updated",
        "last_contact_date": today.isoformat(),
        "follow_up_date": update_data.get('follow_up_date')
    }

# ============== SMTP CONFIG ROUTES ==============

@api_router.post("/smtp-config", response_model=SMTPConfigResponse)
async def create_smtp_config(config_data: SMTPConfigCreate, current_user: dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can configure SMTP")
    
    config_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Delete existing config (only one allowed)
    await db.smtp_config.delete_many({})
    
    config_doc = {
        'id': config_id,
        'smtp_server': config_data.smtp_server,
        'smtp_port': config_data.smtp_port,
        'smtp_username': config_data.smtp_username,
        'smtp_password': config_data.smtp_password,
        'from_email': config_data.from_email,
        'from_name': config_data.from_name,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat()
    }
    
    await db.smtp_config.insert_one(config_doc)
    
    return SMTPConfigResponse(
        id=config_id,
        smtp_server=config_data.smtp_server,
        smtp_port=config_data.smtp_port,
        smtp_username=config_data.smtp_username,
        from_email=config_data.from_email,
        from_name=config_data.from_name,
        created_at=now,
        updated_at=now
    )

@api_router.get("/smtp-config", response_model=Optional[SMTPConfigResponse])
async def get_smtp_config(current_user: dict = Depends(get_current_user)):
    config = await db.smtp_config.find_one({}, {'_id': 0, 'smtp_password': 0})
    if not config:
        return None
    
    created_at = config['created_at']
    updated_at = config['updated_at']
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    return SMTPConfigResponse(
        id=config['id'],
        smtp_server=config['smtp_server'],
        smtp_port=config['smtp_port'],
        smtp_username=config['smtp_username'],
        from_email=config['from_email'],
        from_name=config['from_name'],
        created_at=created_at,
        updated_at=updated_at
    )

# ============== EMAIL ROUTES ==============

async def send_email_task(smtp_config: dict, doctor: dict, subject: str, body: str, is_html: bool, log_id: str):
    try:
        msg = MIMEMultipart()
        msg['From'] = f"{smtp_config['from_name']} <{smtp_config['from_email']}>"
        msg['To'] = f"{doctor['name']} <{doctor['email']}>"
        msg['Subject'] = subject
        
        content_type = 'html' if is_html else 'plain'
        msg.attach(MIMEText(body, content_type))
        
        with smtplib.SMTP(smtp_config['smtp_server'], smtp_config['smtp_port'], timeout=30) as server:
            server.starttls()
            server.login(smtp_config['smtp_username'], smtp_config['smtp_password'])
            server.sendmail(smtp_config['from_email'], [doctor['email']], msg.as_string())
        
        await db.email_logs.update_one(
            {'id': log_id},
            {'$set': {'status': 'sent', 'sent_at': datetime.now(timezone.utc).isoformat()}}
        )
        logger.info(f"Email sent to {doctor['email']}")
    except Exception as e:
        await db.email_logs.update_one(
            {'id': log_id},
            {'$set': {'status': 'failed', 'error_message': str(e)}}
        )
        logger.error(f"Failed to send email: {str(e)}")

@api_router.post("/send-email")
async def send_email(
    email_data: SendEmailRequest,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(get_current_user)
):
    smtp_config = await db.smtp_config.find_one({}, {'_id': 0})
    if not smtp_config:
        raise HTTPException(status_code=400, detail="SMTP not configured. Please configure SMTP settings first.")
    
    doctor = await db.doctors.find_one({'id': email_data.doctor_id}, {'_id': 0})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    log_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    email_log = {
        'id': log_id,
        'doctor_id': doctor['id'],
        'doctor_name': doctor['name'],
        'doctor_email': doctor['email'],
        'subject': email_data.subject,
        'body': email_data.body,
        'status': 'pending',
        'created_at': now.isoformat(),
        'sent_by': current_user['id']
    }
    
    await db.email_logs.insert_one(email_log)
    
    background_tasks.add_task(
        send_email_task,
        smtp_config,
        doctor,
        email_data.subject,
        email_data.body,
        email_data.is_html,
        log_id
    )
    
    return {"message": "Email queued for sending", "log_id": log_id}

@api_router.get("/email-logs", response_model=List[EmailLogResponse])
async def get_email_logs(current_user: dict = Depends(get_current_user)):
    logs = await db.email_logs.find({}, {'_id': 0}).sort('created_at', -1).to_list(100)
    
    result = []
    for log in logs:
        sent_at = log.get('sent_at')
        if sent_at and isinstance(sent_at, str):
            sent_at = datetime.fromisoformat(sent_at.replace('Z', '+00:00'))
        
        result.append(EmailLogResponse(
            id=log['id'],
            doctor_id=log['doctor_id'],
            doctor_name=log['doctor_name'],
            doctor_email=log['doctor_email'],
            subject=log['subject'],
            status=log['status'],
            sent_at=sent_at,
            error_message=log.get('error_message')
        ))
    
    return result

# ============== DASHBOARD ROUTES ==============

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    total_doctors = await db.doctors.count_documents({})
    
    pipeline = [
        {'$group': {'_id': '$lead_status', 'count': {'$sum': 1}}}
    ]
    status_counts = await db.doctors.aggregate(pipeline).to_list(100)
    by_status = {item['_id']: item['count'] for item in status_counts}
    
    # Ensure all statuses are represented
    for status in ['Customer', 'Contacted', 'Pipeline', 'Not Interested', 'Closed']:
        if status not in by_status:
            by_status[status] = 0
    
    # Recent emails count (last 7 days)
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    recent_emails = await db.email_logs.count_documents({'created_at': {'$gte': seven_days_ago}})
    
    # Recent doctors
    recent_docs = await db.doctors.find({}, {'_id': 0}).sort('created_at', -1).to_list(5)
    recent_doctors = []
    for doc in recent_docs:
        created_at = doc.get('created_at', datetime.now(timezone.utc))
        updated_at = doc.get('updated_at', created_at)  # Fallback to created_at if missing
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
        
        recent_doctors.append(DoctorResponse(
            id=doc['id'],
            customer_code=doc['customer_code'],
            name=doc['name'],
            reg_no=doc['reg_no'],
            address=doc['address'],
            email=doc['email'],
            phone=doc['phone'],
            lead_status=doc['lead_status'],
            dob=doc.get('dob'),
            created_at=created_at,
            updated_at=updated_at
        ))
    
    return DashboardStats(
        total_doctors=total_doctors,
        by_status=by_status,
        recent_emails=recent_emails,
        recent_doctors=recent_doctors
    )


@api_router.get("/dashboard/comprehensive-stats")
async def get_comprehensive_dashboard_stats(current_user: dict = Depends(get_current_user)):
    """Get comprehensive dashboard statistics for all entities"""
    
    # ============== CUSTOMERS STATS (Doctors, Medicals, Agencies) ==============
    # Lead statuses for all entities
    lead_statuses = ['Customer', 'Contacted', 'Pipeline', 'Not Interested', 'Closed']
    
    # Doctors stats
    total_doctors = await db.doctors.count_documents({})
    doctor_status_pipeline = [{'$group': {'_id': '$lead_status', 'count': {'$sum': 1}}}]
    doctor_status_counts = await db.doctors.aggregate(doctor_status_pipeline).to_list(100)
    doctors_by_status = {item['_id']: item['count'] for item in doctor_status_counts if item['_id']}
    for status in lead_statuses:
        if status not in doctors_by_status:
            doctors_by_status[status] = 0
    
    # Medicals stats
    total_medicals = await db.medicals.count_documents({})
    medical_status_pipeline = [{'$group': {'_id': '$lead_status', 'count': {'$sum': 1}}}]
    medical_status_counts = await db.medicals.aggregate(medical_status_pipeline).to_list(100)
    medicals_by_status = {item['_id']: item['count'] for item in medical_status_counts if item['_id']}
    for status in lead_statuses:
        if status not in medicals_by_status:
            medicals_by_status[status] = 0
    
    # Agencies stats
    total_agencies = await db.agencies.count_documents({})
    agency_status_pipeline = [{'$group': {'_id': '$lead_status', 'count': {'$sum': 1}}}]
    agency_status_counts = await db.agencies.aggregate(agency_status_pipeline).to_list(100)
    agencies_by_status = {item['_id']: item['count'] for item in agency_status_counts if item['_id']}
    for status in lead_statuses:
        if status not in agencies_by_status:
            agencies_by_status[status] = 0
    
    # Combined lead status stats
    combined_by_status = {}
    for status in lead_statuses:
        combined_by_status[status] = doctors_by_status.get(status, 0) + medicals_by_status.get(status, 0) + agencies_by_status.get(status, 0)
    
    # ============== ORDERS STATS ==============
    order_statuses = ['pending', 'confirmed', 'ready_to_despatch', 'shipped', 'delivered', 'cancelled']
    order_status_pipeline = [{'$group': {'_id': '$status', 'count': {'$sum': 1}}}]
    order_status_counts = await db.orders.aggregate(order_status_pipeline).to_list(100)
    orders_by_status = {item['_id']: item['count'] for item in order_status_counts if item['_id']}
    for status in order_statuses:
        if status not in orders_by_status:
            orders_by_status[status] = 0
    total_orders = sum(orders_by_status.values())
    
    # Recent orders count (last 7 days)
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    recent_orders = await db.orders.count_documents({'created_at': {'$gte': seven_days_ago}})
    
    # ============== PENDING ITEMS STATS ==============
    pending_items = await db.pending_items.find({}, {'_id': 0}).to_list(1000)
    
    # Helper to safely parse quantity (handles expressions like '10+5')
    def safe_parse_qty(qty_str):
        try:
            if isinstance(qty_str, (int, float)):
                return int(qty_str)
            qty_str = str(qty_str).strip()
            if '+' in qty_str:
                return sum(int(x.strip()) for x in qty_str.split('+') if x.strip().isdigit())
            return int(qty_str) if qty_str.isdigit() else 1
        except:
            return 1
    
    total_pending_qty = sum(safe_parse_qty(item.get('quantity', 1)) for item in pending_items)
    
    # Group pending items by item name
    pending_by_item = {}
    for item in pending_items:
        item_name = item.get('item_name', 'Unknown')
        qty = safe_parse_qty(item.get('quantity', 1))
        if item_name in pending_by_item:
            pending_by_item[item_name] += qty
        else:
            pending_by_item[item_name] = qty
    
    # Sort by quantity descending
    pending_by_item_sorted = dict(sorted(pending_by_item.items(), key=lambda x: x[1], reverse=True)[:10])
    
    # ============== EXPENSES STATS ==============
    now = datetime.now(timezone.utc)
    current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
    
    # Current month expenses
    current_month_expenses = await db.expenses.find({
        'date': {'$gte': current_month_start.isoformat()[:10]}
    }, {'_id': 0}).to_list(1000)
    current_month_total = sum(float(e.get('amount', 0)) for e in current_month_expenses)
    
    # Previous month expenses
    previous_month_expenses = await db.expenses.find({
        'date': {'$gte': previous_month_start.isoformat()[:10], '$lt': current_month_start.isoformat()[:10]}
    }, {'_id': 0}).to_list(1000)
    previous_month_total = sum(float(e.get('amount', 0)) for e in previous_month_expenses)
    
    # By category
    category_pipeline = [
        {'$group': {'_id': '$category', 'total': {'$sum': '$amount'}}}
    ]
    expense_by_category = await db.expenses.aggregate(category_pipeline).to_list(100)
    expenses_by_category = {item['_id']: round(item['total'], 2) for item in expense_by_category if item['_id']}
    
    # By payment type
    payment_pipeline = [
        {'$group': {'_id': '$payment_type', 'total': {'$sum': '$amount'}}}
    ]
    expense_by_payment = await db.expenses.aggregate(payment_pipeline).to_list(100)
    expenses_by_payment = {item['_id']: round(item['total'], 2) for item in expense_by_payment if item['_id']}
    
    # ============== ITEMS STATS ==============
    total_items = await db.items.count_documents({})
    
    # Items by subcategory
    subcategory_pipeline = [
        {'$unwind': {'path': '$subcategories', 'preserveNullAndEmptyArrays': True}},
        {'$group': {'_id': '$subcategories', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}}
    ]
    items_by_subcategory = await db.items.aggregate(subcategory_pipeline).to_list(100)
    subcategory_stats = {item['_id'] or 'Uncategorized': item['count'] for item in items_by_subcategory}
    
    # Items by main category
    main_category_pipeline = [
        {'$unwind': {'path': '$main_categories', 'preserveNullAndEmptyArrays': True}},
        {'$group': {'_id': '$main_categories', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}}
    ]
    items_by_main_category = await db.items.aggregate(main_category_pipeline).to_list(100)
    main_category_stats = {item['_id'] or 'Uncategorized': item['count'] for item in items_by_main_category}
    
    # Most ordered items (from order items)
    order_items_pipeline = [
        {'$unwind': '$items'},
        {'$group': {'_id': '$items.item_name', 'total_ordered': {'$sum': 1}}},
        {'$sort': {'total_ordered': -1}},
        {'$limit': 10}
    ]
    most_ordered = await db.orders.aggregate(order_items_pipeline).to_list(10)
    most_ordered_items = [{'item_name': item['_id'], 'order_count': item['total_ordered']} for item in most_ordered if item['_id']]
    
    # Least ordered items (items that exist but have fewer orders)
    all_items = await db.items.find({}, {'_id': 0, 'id': 1, 'item_name': 1, 'item_code': 1}).to_list(1000)
    item_order_counts = {item['_id']: item['total_ordered'] for item in most_ordered if item['_id']}
    
    # Get all item names and find least ordered
    all_item_names = [item.get('item_name') for item in all_items]
    least_ordered_items = []
    for item in all_items:
        item_name = item.get('item_name')
        order_count = item_order_counts.get(item_name, 0)
        least_ordered_items.append({'item_name': item_name, 'item_code': item.get('item_code', ''), 'order_count': order_count})
    
    least_ordered_items.sort(key=lambda x: x['order_count'])
    least_ordered_items = least_ordered_items[:10]
    
    # Items with no orders in 30+ days
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    recent_order_items = await db.orders.aggregate([
        {'$match': {'created_at': {'$gte': thirty_days_ago}}},
        {'$unwind': '$items'},
        {'$group': {'_id': '$items.item_name'}}
    ]).to_list(1000)
    recently_ordered_names = {item['_id'] for item in recent_order_items if item['_id']}
    
    stale_items = []
    for item in all_items:
        item_name = item.get('item_name')
        if item_name and item_name not in recently_ordered_names:
            stale_items.append({'item_name': item_name, 'item_code': item.get('item_code', '')})
    
    # ============== SUPPORT TICKETS STATS ==============
    ticket_statuses = ['open', 'in_progress', 'resolved', 'closed']
    ticket_status_pipeline = [{'$group': {'_id': '$status', 'count': {'$sum': 1}}}]
    ticket_status_counts = await db.support_tickets.aggregate(ticket_status_pipeline).to_list(100)
    tickets_by_status = {item['_id']: item['count'] for item in ticket_status_counts if item['_id']}
    for status in ticket_statuses:
        if status not in tickets_by_status:
            tickets_by_status[status] = 0
    total_tickets = sum(tickets_by_status.values())
    
    # Recent tickets (last 7 days)
    recent_tickets = await db.support_tickets.count_documents({'created_at': {'$gte': seven_days_ago}})
    
    return {
        'customers': {
            'doctors': {
                'total': total_doctors,
                'by_status': doctors_by_status
            },
            'medicals': {
                'total': total_medicals,
                'by_status': medicals_by_status
            },
            'agencies': {
                'total': total_agencies,
                'by_status': agencies_by_status
            },
            'combined_by_status': combined_by_status,
            'total_all': total_doctors + total_medicals + total_agencies
        },
        'orders': {
            'total': total_orders,
            'by_status': orders_by_status,
            'recent_7_days': recent_orders
        },
        'pending_items': {
            'total_items': len(pending_items),
            'total_quantity': total_pending_qty,
            'by_item': pending_by_item_sorted
        },
        'expenses': {
            'current_month_total': round(current_month_total, 2),
            'previous_month_total': round(previous_month_total, 2),
            'change_percent': round(((current_month_total - previous_month_total) / previous_month_total * 100) if previous_month_total > 0 else 0, 1),
            'by_category': expenses_by_category,
            'by_payment_type': expenses_by_payment
        },
        'items': {
            'total': total_items,
            'by_main_category': main_category_stats,
            'by_subcategory': subcategory_stats,
            'most_ordered': most_ordered_items,
            'least_ordered': least_ordered_items,
            'no_orders_30_days': stale_items[:20],
            'stale_count': len(stale_items)
        },
        'support_tickets': {
            'total': total_tickets,
            'by_status': tickets_by_status,
            'recent_7_days': recent_tickets
        }
    }


# ============== ITEM ROUTES ==============

@api_router.get("/item-categories", response_model=List[CategoryResponse])
async def get_item_categories(current_user: dict = Depends(get_current_user)):
    """Get all unique item categories with counts"""
    pipeline = [
        {'$match': {'category': {'$ne': None, '$ne': '', '$exists': True}}},
        {'$group': {'_id': '$category', 'count': {'$sum': 1}}},
        {'$match': {'_id': {'$ne': None}}},
        {'$sort': {'_id': 1}}
    ]
    categories = await db.items.aggregate(pipeline).to_list(100)
    return [CategoryResponse(name=cat['_id'], count=cat['count']) for cat in categories if cat['_id']]

@api_router.post("/items", response_model=ItemResponse)
async def create_item(item_data: ItemCreate, current_user: dict = Depends(get_current_user)):
    item_id = str(uuid.uuid4())
    
    # Use custom item_code if provided, otherwise auto-generate
    if item_data.item_code and item_data.item_code.strip():
        item_code = item_data.item_code.strip()
        # Check if code already exists
        existing = await db.items.find_one({'item_code': item_code}, {'_id': 0})
        if existing:
            raise HTTPException(status_code=400, detail="Item code already exists")
    else:
        item_code = await generate_item_code()
    
    now = datetime.now(timezone.utc)
    
    # Process image if provided
    processed_image = None
    if item_data.image_base64:
        try:
            # Decode base64 and process
            image_bytes = base64.b64decode(item_data.image_base64)
            processed_image = process_image_to_webp(image_bytes)
        except Exception as e:
            logger.error(f"Image processing error: {str(e)}")
    
    # Set default role-based pricing from rate if not provided
    rate = item_data.rate or 0
    
    item_doc = {
        'id': item_id,
        'item_code': item_code,
        'item_name': item_data.item_name,
        'main_categories': item_data.main_categories or [],
        'subcategories': item_data.subcategories or [],
        'composition': item_data.composition,
        'mrp': item_data.mrp,
        'gst': item_data.gst,
        # Role-based pricing - default to rate if not provided
        'rate_doctors': item_data.rate_doctors if item_data.rate_doctors is not None else rate,
        'offer_doctors': item_data.offer_doctors or item_data.offer,
        'special_offer_doctors': item_data.special_offer_doctors or item_data.special_offer,
        'rate_medicals': item_data.rate_medicals if item_data.rate_medicals is not None else rate,
        'offer_medicals': item_data.offer_medicals or item_data.offer,
        'special_offer_medicals': item_data.special_offer_medicals or item_data.special_offer,
        'rate_agencies': item_data.rate_agencies if item_data.rate_agencies is not None else rate,
        'offer_agencies': item_data.offer_agencies or item_data.offer,
        'special_offer_agencies': item_data.special_offer_agencies or item_data.special_offer,
        # Legacy fields
        'rate': rate,
        'offer': item_data.offer,
        'special_offer': item_data.special_offer,
        'custom_fields': [cf.model_dump() for cf in (item_data.custom_fields or [])],
        'image_webp': processed_image,
        'has_image': processed_image is not None,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': current_user['id']
    }
    
    await db.items.insert_one(item_doc)
    
    return ItemResponse(
        id=item_id,
        item_code=item_code,
        item_name=item_data.item_name,
        main_categories=item_data.main_categories or [],
        subcategories=item_data.subcategories or [],
        composition=item_data.composition,
        mrp=item_data.mrp,
        gst=item_data.gst,
        rate_doctors=item_doc['rate_doctors'],
        offer_doctors=item_doc['offer_doctors'],
        special_offer_doctors=item_doc['special_offer_doctors'],
        rate_medicals=item_doc['rate_medicals'],
        offer_medicals=item_doc['offer_medicals'],
        special_offer_medicals=item_doc['special_offer_medicals'],
        rate_agencies=item_doc['rate_agencies'],
        offer_agencies=item_doc['offer_agencies'],
        special_offer_agencies=item_doc['special_offer_agencies'],
        rate=rate,
        offer=item_data.offer,
        special_offer=item_data.special_offer,
        custom_fields=item_data.custom_fields or [],
        image_url=f"/api/items/{item_id}/image" if processed_image else None,
        created_at=now
    )

@api_router.get("/items/{item_id}/image")
async def get_item_image(item_id: str):
    """Get item image as WebP"""
    item = await db.items.find_one({'id': item_id}, {'image_webp': 1})
    if not item or not item.get('image_webp'):
        raise HTTPException(status_code=404, detail="Image not found")
    
    image_data = base64.b64decode(item['image_webp'])
    return Response(content=image_data, media_type="image/webp")

@api_router.get("/items", response_model=List[ItemResponse])
async def get_items(
    search: Optional[str] = None,
    main_category: Optional[str] = None,
    subcategory: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if search:
        query['$or'] = [
            {'item_name': {'$regex': search, '$options': 'i'}},
            {'item_code': {'$regex': search, '$options': 'i'}},
            {'composition': {'$regex': search, '$options': 'i'}}
        ]
    
    if main_category:
        query['main_categories'] = main_category
    
    if subcategory:
        query['subcategories'] = subcategory
    
    items = await db.items.find(query, {'_id': 0, 'image_webp': 0}).sort('created_at', -1).to_list(1000)
    
    result = []
    for item in items:
        created_at = item['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        custom_fields = [CustomField(**cf) for cf in item.get('custom_fields', [])]
        
        # Check if image exists
        has_image = await db.items.find_one({'id': item['id'], 'image_webp': {'$ne': None}}, {'_id': 1})
        
        result.append(ItemResponse(
            id=item['id'],
            item_code=item['item_code'],
            item_name=item['item_name'],
            main_categories=item.get('main_categories', []) or item.get('main_category', []) if isinstance(item.get('main_category'), list) else ([item.get('main_category')] if item.get('main_category') else []),
            subcategories=item.get('subcategories', []),
            composition=item.get('composition'),
            mrp=item['mrp'],
            gst=item.get('gst', 0),
            # Role-based pricing
            rate_doctors=item.get('rate_doctors') or item.get('rate', 0),
            offer_doctors=item.get('offer_doctors') or item.get('offer'),
            special_offer_doctors=item.get('special_offer_doctors') or item.get('special_offer'),
            rate_medicals=item.get('rate_medicals') or item.get('rate', 0),
            offer_medicals=item.get('offer_medicals') or item.get('offer'),
            special_offer_medicals=item.get('special_offer_medicals') or item.get('special_offer'),
            rate_agencies=item.get('rate_agencies') or item.get('rate', 0),
            offer_agencies=item.get('offer_agencies') or item.get('offer'),
            special_offer_agencies=item.get('special_offer_agencies') or item.get('special_offer'),
            # Legacy
            rate=item.get('rate', 0),
            offer=item.get('offer'),
            special_offer=item.get('special_offer'),
            custom_fields=custom_fields,
            image_url=f"/api/items/{item['id']}/image" if has_image else None,
            created_at=created_at
        ))
    
    return result

@api_router.get("/items/{item_id}", response_model=ItemResponse)
async def get_item(item_id: str, current_user: dict = Depends(get_current_user)):
    item = await db.items.find_one({'id': item_id}, {'_id': 0, 'image_webp': 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    created_at = item['created_at']
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    
    custom_fields = [CustomField(**cf) for cf in item.get('custom_fields', [])]
    
    # Check if image exists
    has_image = await db.items.find_one({'id': item_id, 'image_webp': {'$ne': None}}, {'_id': 1})
    
    return ItemResponse(
        id=item['id'],
        item_code=item['item_code'],
        item_name=item['item_name'],
        main_categories=item.get('main_categories', []) or ([item.get('main_category')] if item.get('main_category') else []),
        subcategories=item.get('subcategories', []),
        composition=item.get('composition'),
        offer=item.get('offer'),
        special_offer=item.get('special_offer'),
        mrp=item['mrp'],
        rate=item['rate'],
        gst=item.get('gst', 0),
        # Role-based pricing - Doctors
        rate_doctors=item.get('rate_doctors'),
        offer_doctors=item.get('offer_doctors'),
        special_offer_doctors=item.get('special_offer_doctors'),
        # Role-based pricing - Medicals
        rate_medicals=item.get('rate_medicals'),
        offer_medicals=item.get('offer_medicals'),
        special_offer_medicals=item.get('special_offer_medicals'),
        # Role-based pricing - Agencies
        rate_agencies=item.get('rate_agencies'),
        offer_agencies=item.get('offer_agencies'),
        special_offer_agencies=item.get('special_offer_agencies'),
        custom_fields=custom_fields,
        image_url=f"/api/items/{item['id']}/image" if has_image else None,
        created_at=created_at
    )

@api_router.put("/items/{item_id}", response_model=ItemResponse)
async def update_item(item_id: str, item_data: ItemUpdate, current_user: dict = Depends(get_current_user)):
    item = await db.items.find_one({'id': item_id}, {'_id': 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    update_data = {}
    for k, v in item_data.model_dump().items():
        if v is not None:
            if k == 'custom_fields':
                update_data[k] = [cf.model_dump() if hasattr(cf, 'model_dump') else cf for cf in v]
            elif k == 'image_base64':
                # Process and update image
                try:
                    image_bytes = base64.b64decode(v)
                    update_data['image_webp'] = process_image_to_webp(image_bytes)
                except Exception as e:
                    logger.error(f"Image processing error: {str(e)}")
            elif k == 'item_code':
                # Check if new code already exists (excluding current item)
                if v != item.get('item_code'):
                    existing = await db.items.find_one({'item_code': v, 'id': {'$ne': item_id}}, {'_id': 0})
                    if existing:
                        raise HTTPException(status_code=400, detail="Item code already exists")
                update_data[k] = v
            else:
                update_data[k] = v
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.items.update_one({'id': item_id}, {'$set': update_data})
    
    updated_item = await db.items.find_one({'id': item_id}, {'_id': 0, 'image_webp': 0})
    
    created_at = updated_item['created_at']
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    
    custom_fields = [CustomField(**cf) for cf in updated_item.get('custom_fields', [])]
    
    # Check if image exists
    has_image = await db.items.find_one({'id': item_id, 'image_webp': {'$ne': None}}, {'_id': 1})
    
    return ItemResponse(
        id=updated_item['id'],
        item_code=updated_item['item_code'],
        item_name=updated_item['item_name'],
        main_categories=updated_item.get('main_categories', []) or ([updated_item.get('main_category')] if updated_item.get('main_category') else []),
        subcategories=updated_item.get('subcategories', []),
        composition=updated_item.get('composition'),
        offer=updated_item.get('offer'),
        special_offer=updated_item.get('special_offer'),
        mrp=updated_item['mrp'],
        rate=updated_item['rate'],
        gst=updated_item.get('gst', 0),
        # Role-based pricing - Doctors
        rate_doctors=updated_item.get('rate_doctors'),
        offer_doctors=updated_item.get('offer_doctors'),
        special_offer_doctors=updated_item.get('special_offer_doctors'),
        # Role-based pricing - Medicals
        rate_medicals=updated_item.get('rate_medicals'),
        offer_medicals=updated_item.get('offer_medicals'),
        special_offer_medicals=updated_item.get('special_offer_medicals'),
        # Role-based pricing - Agencies
        rate_agencies=updated_item.get('rate_agencies'),
        offer_agencies=updated_item.get('offer_agencies'),
        special_offer_agencies=updated_item.get('special_offer_agencies'),
        custom_fields=custom_fields,
        image_url=f"/api/items/{item_id}/image" if has_image else None,
        created_at=created_at
    )

@api_router.delete("/items/{item_id}/image")
async def delete_item_image(item_id: str, current_user: dict = Depends(get_current_user)):
    """Delete item image"""
    result = await db.items.update_one({'id': item_id}, {'$set': {'image_webp': None}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Image deleted successfully"}

@api_router.delete("/items/{item_id}")
async def delete_item(item_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.items.delete_one({'id': item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted successfully"}

# ============== BULK IMPORT ENDPOINTS ==============

@api_router.get("/items/import/template")
async def get_import_template(current_user: dict = Depends(get_current_user)):
    """Generate and return Excel template for bulk item import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Import"
    
    # Define headers
    headers = [
        "Item Code*", "Item Name*", "Main Categories", "Subcategories", 
        "Composition", "MRP*", "Rate*", "GST %", "Offer", "Special Offer"
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add sample data rows
    sample_data = [
        ["ITM-0001", "Paracetamol 500mg", "Large Animals, Poultry", "Injection, Powder", "Paracetamol IP 500mg", "50.00", "35.00", "12", "Buy 10 Get 1 Free", "Limited Offer"],
        ["ITM-0002", "Vitamin D3 1000IU", "Pets", "Liquids", "Cholecalciferol 1000IU", "250.00", "180.00", "5", "", ""],
        ["", "Sample Item 3", "Large Animals", "Bolus", "Sample composition", "100.00", "75.00", "18", "", ""],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    instructions = wb.create_sheet("Instructions")
    instructions_data = [
        ["BULK ITEM IMPORT INSTRUCTIONS"],
        [""],
        ["Required Fields (marked with *):"],
        ["- Item Code: Unique code (leave empty for auto-generation)"],
        ["- Item Name: Product name (required)"],
        ["- MRP: Maximum Retail Price (required, number)"],
        ["- Rate: Selling rate (required, number)"],
        [""],
        ["Optional Fields:"],
        ["- Main Categories: Comma-separated (e.g., 'Large Animals, Poultry, Pets')"],
        ["- Subcategories: Comma-separated (e.g., 'Injection, Powder, Liquids')"],
        ["- Composition: Product composition text"],
        ["- GST %: GST percentage (number, default 0)"],
        ["- Offer: Offer text (e.g., 'Buy 10 Get 1 Free')"],
        ["- Special Offer: Special offer text"],
        [""],
        ["Notes:"],
        ["- Images can be added manually after import"],
        ["- Default logo will be used for items without images"],
        ["- Delete sample rows before importing"],
        ["- Maximum 500 items per import"],
    ]
    
    for row_num, row_data in enumerate(instructions_data, 1):
        cell = instructions.cell(row=row_num, column=1, value=row_data[0])
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
    
    # Adjust column widths
    column_widths = [15, 25, 30, 30, 35, 12, 12, 10, 25, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items_import_template.xlsx"}
    )

@api_router.post("/items/import")
async def bulk_import_items(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """Bulk import items from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents), sheet_name=0)
        
        # Normalize column names (remove * and extra spaces)
        df.columns = [col.replace('*', '').strip() for col in df.columns]
        
        # Expected columns mapping
        column_mapping = {
            'Item Code': 'item_code',
            'Item Name': 'item_name',
            'Main Categories': 'main_categories',
            'Subcategories': 'subcategories',
            'Composition': 'composition',
            'MRP': 'mrp',
            'Rate': 'rate',
            'GST %': 'gst',
            'Offer': 'offer',
            'Special Offer': 'special_offer'
        }
        
        # Rename columns
        df = df.rename(columns=column_mapping)
        
        # Validate required columns
        required_cols = ['item_name', 'mrp', 'rate']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing_cols)}")
        
        # Remove rows where item_name is empty
        df = df.dropna(subset=['item_name'])
        df = df[df['item_name'].astype(str).str.strip() != '']
        
        if len(df) == 0:
            raise HTTPException(status_code=400, detail="No valid items found in the file")
        
        if len(df) > 500:
            raise HTTPException(status_code=400, detail="Maximum 500 items per import. Please split your file.")
        
        # Get company logo for default image
        company_settings = await db.company_settings.find_one({}, {'_id': 0})
        default_image = company_settings.get('logo_base64') if company_settings else None
        
        now = datetime.now(timezone.utc)
        items_created = 0
        items_updated = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                item_name = str(row.get('item_name', '')).strip()
                if not item_name:
                    continue
                
                # Parse item_code
                item_code = str(row.get('item_code', '')).strip() if pd.notna(row.get('item_code')) else ''
                
                # Parse main_categories (comma-separated)
                main_cats_str = str(row.get('main_categories', '')) if pd.notna(row.get('main_categories')) else ''
                main_categories = [cat.strip() for cat in main_cats_str.split(',') if cat.strip()]
                
                # Parse subcategories (comma-separated)
                sub_cats_str = str(row.get('subcategories', '')) if pd.notna(row.get('subcategories')) else ''
                subcategories = [cat.strip() for cat in sub_cats_str.split(',') if cat.strip()]
                
                # Parse numeric fields
                mrp = float(row.get('mrp', 0)) if pd.notna(row.get('mrp')) else 0
                rate = float(row.get('rate', 0)) if pd.notna(row.get('rate')) else 0
                gst = float(row.get('gst', 0)) if pd.notna(row.get('gst')) else 0
                
                # Parse text fields
                composition = str(row.get('composition', '')) if pd.notna(row.get('composition')) else ''
                offer = str(row.get('offer', '')) if pd.notna(row.get('offer')) else ''
                special_offer = str(row.get('special_offer', '')) if pd.notna(row.get('special_offer')) else ''
                
                # Check if item with same code exists
                existing_item = None
                if item_code:
                    existing_item = await db.items.find_one({'item_code': item_code}, {'_id': 0})
                
                if existing_item:
                    # Update existing item
                    update_data = {
                        'item_name': item_name,
                        'main_categories': main_categories,
                        'subcategories': subcategories,
                        'composition': composition,
                        'mrp': mrp,
                        'rate': rate,
                        'gst': gst,
                        'offer': offer,
                        'special_offer': special_offer,
                        'updated_at': now.isoformat()
                    }
                    await db.items.update_one({'item_code': item_code}, {'$set': update_data})
                    items_updated += 1
                else:
                    # Generate item_code if not provided
                    if not item_code:
                        item_code = await generate_item_code()
                    
                    # Create new item
                    item_doc = {
                        'id': str(uuid.uuid4()),
                        'item_code': item_code,
                        'item_name': item_name,
                        'main_categories': main_categories,
                        'subcategories': subcategories,
                        'composition': composition,
                        'mrp': mrp,
                        'rate': rate,
                        'gst': gst,
                        'offer': offer,
                        'special_offer': special_offer,
                        'image_base64': default_image,  # Use company logo as default
                        'custom_fields': [],
                        'created_at': now.isoformat(),
                        'updated_at': now.isoformat()
                    }
                    await db.items.insert_one(item_doc)
                    items_created += 1
                    
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        result_message = f"Import completed: {items_created} items created, {items_updated} items updated"
        if errors:
            result_message += f". {len(errors)} errors occurred."
        
        return {
            "message": result_message,
            "items_created": items_created,
            "items_updated": items_updated,
            "errors": errors[:10] if errors else []  # Return first 10 errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# ============== COMPANY SETTINGS ROUTES ==============

@api_router.post("/company-settings", response_model=CompanySettingsResponse)
async def save_company_settings(settings: CompanySettingsCreate, current_user: dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can update company settings")
    
    settings_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Process logo if provided
    processed_logo = None
    if settings.logo_base64:
        try:
            image_bytes = base64.b64decode(settings.logo_base64)
            processed_logo = process_image_to_webp(image_bytes, max_size_kb=50, target_size=(200, 200))
        except Exception as e:
            logger.error(f"Logo processing error: {str(e)}")
    
    # Process login background image if provided
    processed_bg = None
    if settings.login_background_image:
        try:
            image_bytes = base64.b64decode(settings.login_background_image)
            processed_bg = process_image_to_webp(image_bytes, max_size_kb=200, target_size=(1920, 1080))
        except Exception as e:
            logger.error(f"Background image processing error: {str(e)}")
    
    # Delete existing settings (only one allowed)
    await db.company_settings.delete_many({})
    
    settings_doc = {
        'id': settings_id,
        'company_name': settings.company_name,
        'address': settings.address,
        'email': settings.email,
        'phone': settings.phone,
        'gst_number': settings.gst_number,
        'drug_license': settings.drug_license,
        'logo_webp': processed_logo,
        'terms_conditions': settings.terms_conditions,
        'login_tagline': settings.login_tagline,
        'login_background_color': settings.login_background_color,
        'login_background_webp': processed_bg,
        'updated_at': now.isoformat()
    }
    
    await db.company_settings.insert_one(settings_doc)
    
    return CompanySettingsResponse(
        id=settings_id,
        company_name=settings.company_name,
        address=settings.address,
        email=settings.email,
        phone=settings.phone,
        gst_number=settings.gst_number,
        drug_license=settings.drug_license,
        logo_url="/api/company-settings/logo" if processed_logo else None,
        terms_conditions=settings.terms_conditions,
        login_tagline=settings.login_tagline,
        login_background_color=settings.login_background_color,
        login_background_image_url="/api/company-settings/login-background" if processed_bg else None,
        updated_at=now
    )

@api_router.get("/company-settings", response_model=Optional[CompanySettingsResponse])
async def get_company_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.company_settings.find_one({}, {'_id': 0, 'logo_webp': 0, 'login_background_webp': 0})
    if not settings:
        return None
    
    updated_at = settings['updated_at']
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    has_logo = await db.company_settings.find_one({'logo_webp': {'$ne': None}}, {'_id': 1})
    has_bg = await db.company_settings.find_one({'login_background_webp': {'$ne': None}}, {'_id': 1})
    
    return CompanySettingsResponse(
        id=settings['id'],
        company_name=settings['company_name'],
        address=settings['address'],
        email=settings['email'],
        phone=settings.get('phone'),
        gst_number=settings['gst_number'],
        drug_license=settings['drug_license'],
        logo_url="/api/company-settings/logo" if has_logo else None,
        terms_conditions=settings.get('terms_conditions'),
        login_tagline=settings.get('login_tagline'),
        login_background_color=settings.get('login_background_color'),
        login_background_image_url="/api/company-settings/login-background" if has_bg else None,
        updated_at=updated_at
    )

@api_router.get("/company-settings/logo")
async def get_company_logo():
    settings = await db.company_settings.find_one({}, {'logo_webp': 1})
    if not settings or not settings.get('logo_webp'):
        raise HTTPException(status_code=404, detail="Logo not found")
    
    image_data = base64.b64decode(settings['logo_webp'])
    return Response(content=image_data, media_type="image/webp")

@api_router.get("/company-settings/login-background")
async def get_login_background():
    settings = await db.company_settings.find_one({}, {'login_background_webp': 1})
    if not settings or not settings.get('login_background_webp'):
        raise HTTPException(status_code=404, detail="Login background not found")
    
    image_data = base64.b64decode(settings['login_background_webp'])
    return Response(content=image_data, media_type="image/webp")

# ============== PUBLIC SHOWCASE ROUTES (NO AUTH) ==============

@api_router.get("/public/company-settings")
async def get_public_company_settings():
    settings = await db.company_settings.find_one({}, {'_id': 0, 'logo_webp': 0, 'login_background_webp': 0})
    if not settings:
        return None
    
    has_logo = await db.company_settings.find_one({'logo_webp': {'$ne': None}}, {'_id': 1})
    has_bg = await db.company_settings.find_one({'login_background_webp': {'$ne': None}}, {'_id': 1})
    
    return {
        'company_name': settings['company_name'],
        'address': settings['address'],
        'email': settings['email'],
        'phone': settings.get('phone'),
        'gst_number': settings['gst_number'],
        'drug_license': settings['drug_license'],
        'logo_url': "/api/company-settings/logo" if has_logo else None,
        'terms_conditions': settings.get('terms_conditions'),
        'login_tagline': settings.get('login_tagline'),
        'login_background_color': settings.get('login_background_color'),
        'login_background_image_url': "/api/company-settings/login-background" if has_bg else None
    }

@api_router.get("/public/items")
async def get_public_items(main_category: Optional[str] = None, subcategory: Optional[str] = None):
    """Get all items with category filters for public showcase"""
    query = {}
    if main_category:
        query['main_categories'] = main_category
    if subcategory:
        query['subcategories'] = subcategory
    
    items = await db.items.find(query, {'_id': 0, 'image_webp': 0, 'created_by': 0}).sort('item_name', 1).to_list(1000)
    
    result = []
    for item in items:
        has_image = await db.items.find_one({'id': item['id'], 'image_webp': {'$ne': None}}, {'_id': 1})
        item_data = {
            'id': item['id'],
            'item_code': item['item_code'],
            'item_name': item['item_name'],
            'main_categories': item.get('main_categories', []) or ([item.get('main_category')] if item.get('main_category') else []),
            'subcategories': item.get('subcategories', []),
            'composition': item.get('composition'),
            'offer': item.get('offer'),
            'special_offer': item.get('special_offer'),
            'mrp': item['mrp'],
            'rate': item['rate'],
            'gst': item.get('gst', 0),
            'image_url': f"/api/items/{item['id']}/image" if has_image else None
        }
        result.append(item_data)
    
    return result

@api_router.get("/public/categories")
async def get_public_categories():
    """Get all main categories and subcategories for filters"""
    items = await db.items.find({}, {'main_categories': 1, 'main_category': 1, 'subcategories': 1, '_id': 0}).to_list(1000)
    
    main_categories = set()
    subcategories_map = {}
    
    for item in items:
        # Handle both old (main_category) and new (main_categories) field names
        main_cats = item.get('main_categories', [])
        if not main_cats and item.get('main_category'):
            main_cats = [item.get('main_category')]
        
        for main_cat in main_cats:
            if main_cat:
                main_categories.add(main_cat)
                if main_cat not in subcategories_map:
                    subcategories_map[main_cat] = set()
                for sub in item.get('subcategories', []):
                    subcategories_map[main_cat].add(sub)
    
    return {
        'main_categories': sorted(list(main_categories)),
        'subcategories': {k: sorted(list(v)) for k, v in subcategories_map.items()}
    }

# ============== LOCATION DATA ROUTES (PUBLIC) ==============

@api_router.get("/public/states")
async def get_states():
    """Get list of all Indian states and union territories"""
    return {"states": get_all_states()}

@api_router.get("/public/districts/{state}")
async def get_districts(state: str):
    """Get list of districts for a given state"""
    districts = get_districts_by_state(state)
    if not districts:
        return {"districts": [], "message": "No districts found for the given state"}
    return {"districts": districts}

@api_router.get("/public/doctor/{mobile}")
async def get_doctor_by_mobile(mobile: str):
    """Get doctor details by mobile number for auto-fill"""
    # Clean mobile number (remove spaces, dashes, etc.)
    clean_mobile = ''.join(filter(str.isdigit, mobile))
    
    # Search with various formats
    doctor = await db.doctors.find_one(
        {'$or': [
            {'phone': {'$regex': clean_mobile[-10:], '$options': 'i'}},
            {'phone': clean_mobile},
            {'phone': f"+91{clean_mobile[-10:]}"},
            {'phone': f"91{clean_mobile[-10:]}"}
        ]},
        {'_id': 0}
    )
    
    if not doctor:
        return None
    
    return {
        'id': doctor['id'],
        'name': doctor['name'],
        'phone': doctor['phone'],
        'email': doctor['email'],
        'address': doctor['address'],
        'customer_code': doctor['customer_code']
    }

# ============== CUSTOMER PORTAL ROUTES ==============

# Store OTPs in memory (in production, use Redis)
customer_otp_store = {}

async def generate_customer_code(role: str):
    """Generate unique customer code based on role"""
    prefix = "CUS"
    if role == "doctor":
        prefix = "DOC"
    elif role == "medical":
        prefix = "MED"
    elif role == "agency":
        prefix = "AGY"
    
    count = await db.portal_customers.count_documents({'role': role})
    return f"{prefix}-{str(count + 1).zfill(4)}"

async def generate_ticket_number():
    """Generate unique ticket number"""
    today = datetime.now(timezone.utc).strftime('%Y%m%d')
    count = await db.support_tickets.count_documents({
        'created_at': {'$regex': f'^{today[:4]}-{today[4:6]}-{today[6:8]}'}
    })
    return f"TKT-{today}-{str(count + 1).zfill(4)}"

def create_customer_token(customer_id: str, role: str):
    """Create JWT token for customer"""
    payload = {
        'customer_id': customer_id,
        'role': role,
        'type': 'customer',
        'exp': datetime.now(timezone.utc) + timedelta(days=30)
    }
    return jwt.encode(payload, os.environ.get('JWT_SECRET', 'vmp-crm-secret-key-2024'), algorithm='HS256')

async def get_current_customer(credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer())):
    """Validate customer JWT token"""
    try:
        payload = jwt.decode(credentials.credentials, os.environ.get('JWT_SECRET', 'vmp-crm-secret-key-2024'), algorithms=['HS256'])
        if payload.get('type') != 'customer':
            raise HTTPException(status_code=401, detail="Invalid token type")
        
        customer = await db.portal_customers.find_one({'id': payload['customer_id']}, {'_id': 0})
        if not customer:
            raise HTTPException(status_code=401, detail="Customer not found")
        if customer['status'] != 'approved':
            raise HTTPException(status_code=403, detail=f"Account is {customer['status']}")
        return customer
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

@api_router.post("/customer/send-otp")
async def customer_send_otp(request: CustomerOTPRequest):
    """Send OTP to customer for registration or password reset"""
    clean_phone = ''.join(filter(str.isdigit, request.phone))
    if len(clean_phone) < 10:
        raise HTTPException(status_code=400, detail="Invalid phone number")
    
    # Check if customer exists for password reset
    if request.purpose == "reset_password":
        customer = await db.portal_customers.find_one({'phone': clean_phone}, {'_id': 0})
        if not customer:
            raise HTTPException(status_code=404, detail="No account found with this phone number")
    
    # Check for duplicate registration - prevent if phone already registered
    if request.purpose == "register":
        # Check portal_customers first
        existing_portal = await db.portal_customers.find_one({'phone': clean_phone}, {'_id': 0})
        if existing_portal:
            status = existing_portal.get('status', '')
            if status == 'pending_approval':
                raise HTTPException(status_code=400, detail="This phone number has a pending registration. Please wait for approval or contact support.")
            elif status == 'approved':
                raise HTTPException(status_code=400, detail="This phone number is already registered. Please login instead.")
            elif status == 'rejected':
                raise HTTPException(status_code=400, detail="Registration was rejected. Please contact support for assistance.")
            elif status == 'suspended':
                raise HTTPException(status_code=400, detail="This account is suspended. Please contact support.")
            else:
                raise HTTPException(status_code=400, detail="This phone number is already registered.")
        
        # Check doctors, medicals, agencies collections (admin-created customers)
        existing_doctor = await db.doctors.find_one({'phone': clean_phone}, {'_id': 0})
        if existing_doctor:
            raise HTTPException(status_code=400, detail="This phone number is already registered as a Doctor. Please contact admin for portal access.")
        
        existing_medical = await db.medicals.find_one({'phone': clean_phone}, {'_id': 0})
        if existing_medical:
            raise HTTPException(status_code=400, detail="This phone number is already registered as a Medical Store. Please contact admin for portal access.")
        
        existing_agency = await db.agencies.find_one({'phone': clean_phone}, {'_id': 0})
        if existing_agency:
            raise HTTPException(status_code=400, detail="This phone number is already registered as an Agency. Please contact admin for portal access.")
    
    # Generate 4-digit OTP
    otp = str(random.randint(1000, 9999))
    
    # Store OTP with expiry (5 minutes)
    customer_otp_store[f"{clean_phone}_{request.purpose}"] = {
        'otp': otp,
        'expires': datetime.now(timezone.utc) + timedelta(minutes=5)
    }
    
    # Send OTP via WhatsApp using BotMasterSender API
    config = await get_whatsapp_config()
    if config.get('api_url') and config.get('auth_token') and config.get('sender_id'):
        try:
            purpose_text = "registration" if request.purpose == "register" else "password reset"
            message = f"Your VMP CRM verification code for {purpose_text} is: *{otp}*\n\nThis code expires in 5 minutes. Do not share this code with anyone."
            
            # Ensure mobile has 91 prefix for India
            wa_mobile = clean_phone if clean_phone.startswith('91') else f"91{clean_phone[-10:]}"
            
            params = {
                'action': 'send',
                'senderId': config['sender_id'],
                'authToken': config['auth_token'],
                'messageText': message,
                'receiverId': wa_mobile
            }
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(config['api_url'], params=params)
                if response.status_code == 200:
                    await log_whatsapp_message(wa_mobile, 'otp', message, 'success')
                    logger.info(f"OTP sent successfully to {wa_mobile}")
                else:
                    await log_whatsapp_message(wa_mobile, 'otp', message, 'failed', error_message=f"Status: {response.status_code}")
                    logger.error(f"WhatsApp OTP failed: {response.status_code} - {response.text}")
        except Exception as e:
            logger.error(f"WhatsApp OTP error: {str(e)}")
            await log_whatsapp_message(clean_phone, 'otp', f"OTP: {otp}", 'failed', error_message=str(e))
    
    return {"message": "OTP sent successfully", "phone": clean_phone}

@api_router.post("/customer/verify-otp")
async def customer_verify_otp(request: CustomerOTPVerify):
    """Verify OTP for customer"""
    clean_phone = ''.join(filter(str.isdigit, request.phone))
    otp_key = f"{clean_phone}_{request.purpose}"
    
    stored = customer_otp_store.get(otp_key)
    
    # First check if it's a fallback OTP (admin-managed)
    fallback_otp = await db.fallback_otps.find_one({'otp': request.otp, 'is_active': True}, {'_id': 0})
    if fallback_otp:
        # Increment usage count
        await db.fallback_otps.update_one(
            {'id': fallback_otp['id']},
            {'$inc': {'used_count': 1}}
        )
        logger.info(f"Fallback OTP used for phone {clean_phone}")
        
        # Store as verified for registration flow
        customer_otp_store[otp_key] = {
            'otp': request.otp,
            'expires': datetime.now(timezone.utc) + timedelta(minutes=30),
            'verified': True,
            'is_fallback': True
        }
        return {"message": "OTP verified successfully", "verified": True}
    
    # Check regular OTP
    if not stored:
        raise HTTPException(status_code=400, detail="OTP not found or expired. Please request a new OTP.")
    
    if datetime.now(timezone.utc) > stored['expires']:
        del customer_otp_store[otp_key]
        raise HTTPException(status_code=400, detail="OTP expired. Please request a new OTP.")
    
    if stored['otp'] != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    # Mark OTP as verified
    customer_otp_store[otp_key]['verified'] = True
    
    return {"message": "OTP verified successfully", "verified": True}

@api_router.post("/customer/register", response_model=CustomerResponse)
async def customer_register(request: CustomerRegister):
    """Register new customer (requires OTP verification)"""
    clean_phone = ''.join(filter(str.isdigit, request.phone))
    
    # Verify OTP was verified
    otp_key = f"{clean_phone}_register"
    stored = customer_otp_store.get(otp_key)
    if not stored or not stored.get('verified'):
        raise HTTPException(status_code=400, detail="Please verify OTP first")
    
    # Check if phone already exists
    existing = await db.portal_customers.find_one({'phone': clean_phone}, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="Phone number already registered")
    
    # Validate role
    if request.role not in ['doctor', 'medical', 'agency']:
        raise HTTPException(status_code=400, detail="Invalid role. Must be doctor, medical, or agency")
    
    customer_id = str(uuid.uuid4())
    customer_code = await generate_customer_code(request.role)
    now = datetime.now(timezone.utc)
    
    # Hash password
    password_hash = bcrypt.hashpw(request.password.encode(), bcrypt.gensalt()).decode()
    
    customer_doc = {
        'id': customer_id,
        'customer_code': customer_code,
        'name': request.name,
        'phone': clean_phone,
        'email': request.email,
        'password_hash': password_hash,
        'role': request.role,
        'status': 'pending_approval',
        'reg_no': request.reg_no,
        'dob': request.dob,
        'proprietor_name': request.proprietor_name,
        'gst_number': request.gst_number,
        'drug_license': request.drug_license,
        'alternate_phone': request.alternate_phone,
        'birthday': request.birthday,
        'anniversary': request.anniversary,
        'address_line_1': request.address_line_1,
        'address_line_2': request.address_line_2,
        'state': request.state,
        'district': request.district,
        'pincode': request.pincode,
        'delivery_station': request.delivery_station,
        'transport_id': request.transport_id,
        'created_at': now.isoformat(),
        'approved_at': None,
        'approved_by': None
    }
    
    await db.portal_customers.insert_one(customer_doc)
    
    # Clear OTP
    del customer_otp_store[otp_key]
    
    # Get transport name if exists
    transport_name = None
    if request.transport_id:
        transport = await db.transports.find_one({'id': request.transport_id}, {'_id': 0, 'name': 1})
        transport_name = transport.get('name') if transport else None
    
    return CustomerResponse(
        id=customer_id,
        customer_code=customer_code,
        name=request.name,
        phone=clean_phone,
        email=request.email,
        role=request.role,
        status='pending_approval',
        reg_no=request.reg_no,
        dob=request.dob,
        proprietor_name=request.proprietor_name,
        gst_number=request.gst_number,
        drug_license=request.drug_license,
        alternate_phone=request.alternate_phone,
        birthday=request.birthday,
        anniversary=request.anniversary,
        address_line_1=request.address_line_1,
        address_line_2=request.address_line_2,
        state=request.state,
        district=request.district,
        pincode=request.pincode,
        delivery_station=request.delivery_station,
        transport_id=request.transport_id,
        transport_name=transport_name,
        created_at=now
    )

@api_router.post("/customer/login")
async def customer_login(request: CustomerLogin):
    """Customer login"""
    clean_phone = ''.join(filter(str.isdigit, request.phone))
    
    customer = await db.portal_customers.find_one({'phone': clean_phone}, {'_id': 0})
    if not customer:
        raise HTTPException(status_code=401, detail="Invalid phone or password")
    
    if not bcrypt.checkpw(request.password.encode(), customer['password_hash'].encode()):
        raise HTTPException(status_code=401, detail="Invalid phone or password")
    
    if customer['status'] == 'pending_approval':
        raise HTTPException(status_code=403, detail="Your account is pending approval. Please wait for admin approval.")
    
    if customer['status'] == 'rejected':
        raise HTTPException(status_code=403, detail="Your registration was rejected. Please contact support.")
    
    if customer['status'] == 'suspended':
        raise HTTPException(status_code=403, detail="Your account has been suspended. Please contact support.")
    
    token = create_customer_token(customer['id'], customer['role'])
    
    return {
        "access_token": token,
        "customer": {
            "id": customer['id'],
            "name": customer['name'],
            "phone": customer['phone'],
            "role": customer['role'],
            "customer_code": customer['customer_code']
        }
    }

@api_router.post("/customer/login-otp-send")
async def customer_login_otp_send(request: CustomerOTPRequest):
    """Send OTP for customer login (passwordless)"""
    clean_phone = ''.join(filter(str.isdigit, request.phone))
    if len(clean_phone) < 10:
        raise HTTPException(status_code=400, detail="Invalid phone number")
    
    customer = await db.portal_customers.find_one({'phone': clean_phone}, {'_id': 0})
    if not customer:
        raise HTTPException(status_code=404, detail="No account found with this phone number. Please register first.")
    
    if customer['status'] == 'pending_approval':
        raise HTTPException(status_code=403, detail="Your account is pending approval.")
    if customer['status'] in ('rejected', 'suspended'):
        raise HTTPException(status_code=403, detail="Your account is not active. Please contact support.")
    
    otp = str(random.randint(1000, 9999))
    customer_otp_store[f"{clean_phone}_login"] = {
        'otp': otp,
        'expires': datetime.now(timezone.utc) + timedelta(minutes=5)
    }
    
    config = await get_whatsapp_config()
    if config.get('api_url') and config.get('auth_token') and config.get('sender_id'):
        try:
            await send_whatsapp_otp(clean_phone, otp)
        except Exception as e:
            logger.error(f"Failed to send login OTP via WhatsApp: {e}")
    
    return {"message": "OTP sent to your WhatsApp", "phone": clean_phone}

@api_router.post("/customer/login-otp-verify")
async def customer_login_otp_verify(request: CustomerOTPVerify):
    """Verify OTP and login customer (passwordless)"""
    clean_phone = ''.join(filter(str.isdigit, request.phone))
    otp_key = f"{clean_phone}_login"
    
    stored = customer_otp_store.get(otp_key)
    
    # Check fallback OTP
    if not stored:
        fallback_otp = await db.fallback_otps.find_one({'otp': request.otp, 'is_active': True}, {'_id': 0})
        if fallback_otp:
            customer_otp_store[otp_key] = {'otp': request.otp, 'expires': datetime.now(timezone.utc) + timedelta(minutes=5)}
            stored = customer_otp_store[otp_key]
    
    if not stored:
        raise HTTPException(status_code=400, detail="OTP not found or expired. Please request a new one.")
    
    if datetime.now(timezone.utc) > stored['expires']:
        del customer_otp_store[otp_key]
        raise HTTPException(status_code=400, detail="OTP expired. Please request a new one.")
    
    if stored['otp'] != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    del customer_otp_store[otp_key]
    
    customer = await db.portal_customers.find_one({'phone': clean_phone}, {'_id': 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    if customer['status'] != 'approved':
        raise HTTPException(status_code=403, detail="Your account is not active.")
    
    token = create_customer_token(customer['id'], customer['role'])
    
    return {
        "access_token": token,
        "customer": {
            "id": customer['id'],
            "name": customer['name'],
            "phone": customer['phone'],
            "role": customer['role'],
            "customer_code": customer['customer_code']
        }
    }

@api_router.post("/customer/reset-password")
async def customer_reset_password(request: CustomerResetPassword):
    """Reset customer password"""
    clean_phone = ''.join(filter(str.isdigit, request.phone))
    
    # Verify OTP
    otp_key = f"{clean_phone}_reset_password"
    stored = customer_otp_store.get(otp_key)
    if not stored:
        raise HTTPException(status_code=400, detail="OTP not found or expired")
    
    if stored['otp'] != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    # Update password
    password_hash = bcrypt.hashpw(request.new_password.encode(), bcrypt.gensalt()).decode()
    result = await db.portal_customers.update_one(
        {'phone': clean_phone},
        {'$set': {'password_hash': password_hash, 'updated_at': datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Clear OTP
    del customer_otp_store[otp_key]
    
    return {"message": "Password reset successfully"}

@api_router.get("/customer/profile", response_model=CustomerResponse)
async def get_customer_profile(customer: dict = Depends(get_current_customer)):
    """Get current customer profile"""
    transport_name = None
    if customer.get('transport_id'):
        transport = await db.transports.find_one({'id': customer['transport_id']}, {'_id': 0, 'name': 1})
        transport_name = transport.get('name') if transport else None
    
    created_at = customer.get('created_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    
    approved_at = customer.get('approved_at')
    if approved_at and isinstance(approved_at, str):
        approved_at = datetime.fromisoformat(approved_at.replace('Z', '+00:00'))
    
    return CustomerResponse(
        id=customer['id'],
        customer_code=customer['customer_code'],
        name=customer['name'],
        phone=customer['phone'],
        email=customer.get('email'),
        role=customer['role'],
        status=customer['status'],
        reg_no=customer.get('reg_no'),
        proprietor_name=customer.get('proprietor_name'),
        gst_number=customer.get('gst_number'),
        drug_license=customer.get('drug_license'),
        address_line_1=customer.get('address_line_1'),
        address_line_2=customer.get('address_line_2'),
        state=customer.get('state'),
        district=customer.get('district'),
        pincode=customer.get('pincode'),
        delivery_station=customer.get('delivery_station'),
        transport_id=customer.get('transport_id'),
        transport_name=transport_name,
        created_at=created_at,
        approved_at=approved_at,
        approved_by=customer.get('approved_by')
    )

@api_router.put("/customer/profile", response_model=CustomerResponse)
async def update_customer_profile(update_data: CustomerProfileUpdate, customer: dict = Depends(get_current_customer)):
    """Update customer profile"""
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.portal_customers.update_one({'id': customer['id']}, {'$set': update_dict})
    
    # Get updated customer
    updated_customer = await db.portal_customers.find_one({'id': customer['id']}, {'_id': 0})
    return await get_customer_profile(updated_customer)

@api_router.get("/customer/items")
async def get_customer_items(
    main_category: Optional[str] = None,
    subcategory: Optional[str] = None,
    customer: dict = Depends(get_current_customer)
):
    """Get items with role-based pricing for logged-in customer"""
    query = {}
    if main_category:
        query['main_categories'] = main_category
    if subcategory:
        query['subcategories'] = subcategory
    
    items = await db.items.find(query, {'_id': 0, 'image_webp': 0, 'created_by': 0}).sort('item_name', 1).to_list(1000)
    
    role = customer['role']
    result = []
    
    for item in items:
        # Get role-specific pricing
        if role == 'doctor':
            rate = item.get('rate_doctors') or item.get('rate', 0)
            offer = item.get('offer_doctors') or item.get('offer')
            special_offer = item.get('special_offer_doctors') or item.get('special_offer')
        elif role == 'medical':
            rate = item.get('rate_medicals') or item.get('rate', 0)
            offer = item.get('offer_medicals') or item.get('offer')
            special_offer = item.get('special_offer_medicals') or item.get('special_offer')
        else:  # agency
            rate = item.get('rate_agencies') or item.get('rate', 0)
            offer = item.get('offer_agencies') or item.get('offer')
            special_offer = item.get('special_offer_agencies') or item.get('special_offer')
        
        result.append({
            'id': item['id'],
            'item_code': item['item_code'],
            'item_name': item['item_name'],
            'main_categories': item.get('main_categories', []),
            'subcategories': item.get('subcategories', []),
            'composition': item.get('composition'),
            'mrp': item['mrp'],
            'rate': rate,
            'offer': offer,
            'special_offer': special_offer,
            'gst': item.get('gst', 0),
            'image_url': f"/api/items/{item['id']}/image" if item.get('has_image') else None
        })
    
    return result

@api_router.get("/customer/orders")
async def get_customer_orders(customer: dict = Depends(get_current_customer)):
    """Get order history for logged-in customer"""
    # Find orders linked to this customer's phone
    orders = await db.orders.find(
        {'doctor_phone': customer['phone']},
        {'_id': 0}
    ).sort('created_at', -1).to_list(100)
    
    result = []
    for order in orders:
        created_at = order.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append({
            'id': order['id'],
            'order_number': order['order_number'],
            'items': order.get('items', []),
            'status': order['status'],
            'transport_name': order.get('transport_name'),
            'tracking_number': order.get('tracking_number'),
            'tracking_url': order.get('tracking_url'),
            'delivery_station': order.get('delivery_station'),
            'payment_mode': order.get('payment_mode'),
            'boxes_count': order.get('boxes_count'),
            'cans_count': order.get('cans_count'),
            'bags_count': order.get('bags_count'),
            'invoice_number': order.get('invoice_number'),
            'invoice_date': order.get('invoice_date'),
            'invoice_value': order.get('invoice_value'),
            'notes': order.get('notes'),
            'created_at': created_at
        })
    
    return result

@api_router.post("/customer/orders")
async def create_customer_order(request: Request, background_tasks: BackgroundTasks, customer: dict = Depends(get_current_customer)):
    """Create a new order from the customer portal cart"""
    body = await request.json()
    items = body.get('items', [])
    notes = body.get('notes', '')
    
    if not items:
        raise HTTPException(status_code=400, detail="Cart is empty")
    
    # Generate order number
    today_str = datetime.now(timezone.utc).strftime('%Y%m%d')
    today_orders = await db.orders.count_documents({
        'order_number': {'$regex': f'^ORD-{today_str}'}
    })
    order_number = f"ORD-{today_str}-{(today_orders + 1):04d}"
    
    order_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    clean_phone = customer.get('phone', '')
    customer_name = customer.get('name', '')
    customer_role = customer.get('role', 'doctor')
    
    # Try to find linked entity
    entity_id = None
    collection_map = {
        'doctor': db.doctors,
        'medical': db.medicals,
        'agency': db.agencies
    }
    collection = collection_map.get(customer_role, db.doctors)
    linked = await collection.find_one({'phone': clean_phone}, {'_id': 0, 'id': 1})
    if linked:
        entity_id = linked['id']
    
    # Build order items
    order_items = []
    for item in items:
        order_items.append({
            'item_id': item.get('item_id', ''),
            'item_code': item.get('item_code', ''),
            'item_name': item.get('item_name', ''),
            'quantity': str(item.get('quantity', '1')),
            'rate': float(item.get('rate', 0)),
            'mrp': float(item.get('rate', 0)),
        })
    
    order_doc = {
        'id': order_id,
        'order_number': order_number,
        'doctor_id': entity_id or customer.get('id', ''),
        'customer_type': customer_role,
        'doctor_name': customer_name,
        'doctor_phone': clean_phone,
        'doctor_email': customer.get('email', ''),
        'doctor_address': customer.get('address', ''),
        'items': order_items,
        'status': 'pending',
        'notes': notes,
        'created_by': customer_name,
        'created_by_id': customer.get('id', ''),
        'source': 'customer_portal',
        'created_at': now.isoformat()
    }
    
    await db.orders.insert_one(order_doc)
    
    # Send WhatsApp confirmation in background
    try:
        order_item_models = [OrderItem(**oi) for oi in order_items]
        background_tasks.add_task(
            send_whatsapp_order,
            clean_phone,
            order_item_models,
            order_number,
            customer_name
        )
    except Exception as e:
        logger.error(f"Failed to queue WhatsApp for customer order: {e}")
    
    # Send email confirmation in background
    try:
        background_tasks.add_task(
            send_order_confirmation_email,
            order_doc,
            order_items
        )
    except Exception as e:
        logger.error(f"Failed to queue email for customer order: {e}")
    
    return {
        "message": "Order placed successfully",
        "order_number": order_number,
        "order_id": order_id
    }

@api_router.get("/customer/tasks")
async def get_customer_tasks(customer: dict = Depends(get_current_customer)):
    """Get tasks assigned to customer (created by admin/staff for them)"""
    tasks = []
    
    # Strategy 1: Find linked entity by phone
    collection_map = {
        'doctor': ('doctors', 'doctor_id'),
        'medical': ('medicals', 'medical_id'),
        'agency': ('agencies', 'agency_id')
    }
    
    role = customer.get('role', 'doctor')
    coll_name, id_field = collection_map.get(role, ('doctors', 'doctor_id'))
    collection = db[coll_name]
    
    # Try finding by phone first
    linked = await collection.find_one({'phone': customer['phone']}, {'_id': 0, 'id': 1})
    
    # Fallback: try finding by portal_customer_id
    if not linked:
        linked = await collection.find_one({'portal_customer_id': customer['id']}, {'_id': 0, 'id': 1})
    
    if linked:
        tasks = await db.tasks.find({id_field: linked['id']}, {'_id': 0}).sort('created_at', -1).to_list(50)
    
    result = []
    for task in tasks:
        created_at = task.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append({
            'id': task['id'],
            'title': task['title'],
            'description': task.get('description'),
            'due_date': task.get('due_date'),
            'priority': task.get('priority', 'moderate'),
            'status': task.get('status', 'pending'),
            'created_at': created_at
        })
    
    return result

# Support Ticket Routes for Customers

@api_router.post("/customer/tickets", response_model=TicketResponse)
async def create_ticket(ticket_data: TicketCreate, customer: dict = Depends(get_current_customer)):
    """Create support ticket"""
    ticket_id = str(uuid.uuid4())
    ticket_number = await generate_ticket_number()
    now = datetime.now(timezone.utc)
    
    # Get order details if linked
    order_number = None
    if ticket_data.order_id:
        order = await db.orders.find_one({'id': ticket_data.order_id}, {'_id': 0, 'order_number': 1})
        order_number = order.get('order_number') if order else None
    
    ticket_doc = {
        'id': ticket_id,
        'ticket_number': ticket_number,
        'customer_id': customer['id'],
        'customer_name': customer['name'],
        'customer_phone': customer['phone'],
        'customer_role': customer['role'],
        'subject': ticket_data.subject,
        'description': ticket_data.description,
        'order_id': ticket_data.order_id,
        'order_number': order_number,
        'priority': ticket_data.priority,
        'status': 'open',
        'replies': [],
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'resolved_at': None
    }
    
    await db.support_tickets.insert_one(ticket_doc)
    
    return TicketResponse(
        id=ticket_id,
        ticket_number=ticket_number,
        customer_id=customer['id'],
        customer_name=customer['name'],
        customer_phone=customer['phone'],
        customer_role=customer['role'],
        subject=ticket_data.subject,
        description=ticket_data.description,
        order_id=ticket_data.order_id,
        order_number=order_number,
        priority=ticket_data.priority,
        status='open',
        replies=[],
        created_at=now,
        updated_at=now
    )

@api_router.get("/customer/tickets")
async def get_customer_tickets(customer: dict = Depends(get_current_customer)):
    """Get customer's support tickets"""
    tickets = await db.support_tickets.find(
        {'customer_id': customer['id']},
        {'_id': 0}
    ).sort('created_at', -1).to_list(100)
    
    result = []
    for ticket in tickets:
        created_at = ticket.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        updated_at = ticket.get('updated_at')
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
        
        result.append({
            'id': ticket['id'],
            'ticket_number': ticket['ticket_number'],
            'subject': ticket['subject'],
            'description': ticket['description'],
            'order_id': ticket.get('order_id'),
            'order_number': ticket.get('order_number'),
            'priority': ticket['priority'],
            'status': ticket['status'],
            'replies': ticket.get('replies', []),
            'created_at': created_at,
            'updated_at': updated_at
        })
    
    return result

@api_router.post("/customer/tickets/{ticket_id}/reply")
async def add_customer_ticket_reply(ticket_id: str, reply: TicketReply, customer: dict = Depends(get_current_customer)):
    """Add reply to ticket from customer"""
    ticket = await db.support_tickets.find_one({'id': ticket_id, 'customer_id': customer['id']}, {'_id': 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    now = datetime.now(timezone.utc)
    reply_doc = {
        'id': str(uuid.uuid4()),
        'message': reply.message,
        'sender_type': 'customer',
        'sender_name': customer['name'],
        'created_at': now.isoformat()
    }
    
    await db.support_tickets.update_one(
        {'id': ticket_id},
        {
            '$push': {'replies': reply_doc},
            '$set': {'updated_at': now.isoformat()}
        }
    )
    
    return {"message": "Reply added successfully", "reply": reply_doc}

# Admin routes for customer management

@api_router.get("/customers")
async def get_all_customers(
    status: Optional[str] = None,
    role: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all portal customers (admin only)"""
    query = {}
    if status:
        query['status'] = status
    if role:
        query['role'] = role
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}},
            {'customer_code': {'$regex': search, '$options': 'i'}},
            {'email': {'$regex': search, '$options': 'i'}}
        ]
    
    customers = await db.portal_customers.find(query, {'_id': 0, 'password_hash': 0}).sort('created_at', -1).to_list(500)
    
    # Get transport names
    transport_ids = [c.get('transport_id') for c in customers if c.get('transport_id')]
    transport_map = {}
    if transport_ids:
        transports = await db.transports.find({'id': {'$in': transport_ids}}, {'_id': 0, 'id': 1, 'name': 1}).to_list(100)
        transport_map = {t['id']: t['name'] for t in transports}
    
    result = []
    for c in customers:
        created_at = c.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append({
            **c,
            'transport_name': transport_map.get(c.get('transport_id')),
            'created_at': created_at
        })
    
    return result

@api_router.put("/customers/{customer_id}/approve")
async def approve_customer(customer_id: str, approval: CustomerApproval, current_user: dict = Depends(get_current_user)):
    """Approve or reject customer registration"""
    customer = await db.portal_customers.find_one({'id': customer_id}, {'_id': 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    now = datetime.now(timezone.utc)
    update_data = {
        'status': approval.status,
        'updated_at': now.isoformat()
    }
    
    if approval.status == 'approved':
        update_data['approved_at'] = now.isoformat()
        update_data['approved_by'] = current_user['name']
        
        # Create record in respective collection (doctors/medicals/agencies)
        linked_record_id = await create_linked_customer_record(customer, current_user['name'])
        if linked_record_id:
            update_data['linked_record_id'] = linked_record_id
            
    elif approval.status == 'rejected':
        update_data['rejection_reason'] = approval.rejection_reason
    
    await db.portal_customers.update_one({'id': customer_id}, {'$set': update_data})
    
    # Send WhatsApp notification using BotMasterSender API
    config = await get_whatsapp_config()
    if config.get('api_url') and config.get('auth_token') and config.get('sender_id'):
        try:
            if approval.status == 'approved':
                message = f"Great news, {customer['name']}!\n\nYour VMP CRM account has been *APPROVED*!\n\nYou can now login to view products and place orders.\n\nCustomer Code: {customer['customer_code']}"
            else:
                reason = approval.rejection_reason or "Please contact support for more details."
                message = f"Hello {customer['name']},\n\nUnfortunately, your VMP CRM registration has been declined.\n\nReason: {reason}\n\nPlease contact support for assistance."
            
            # Ensure mobile has 91 prefix for India
            wa_mobile = customer['phone'] if customer['phone'].startswith('91') else f"91{customer['phone'][-10:]}"
            
            params = {
                'action': 'send',
                'senderId': config['sender_id'],
                'authToken': config['auth_token'],
                'messageText': message,
                'receiverId': wa_mobile
            }
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(config['api_url'], params=params)
                if response.status_code == 200:
                    await log_whatsapp_message(wa_mobile, 'account_status', message, 'success', recipient_name=customer['name'])
                else:
                    logger.error(f"WhatsApp notification failed: {response.status_code}")
        except Exception as e:
            logger.error(f"WhatsApp notification error: {str(e)}")
    
    return {"message": f"Customer {approval.status} successfully"}


@api_router.post("/customers/{customer_id}/send-new-password")
async def send_new_password_to_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    """Generate and send new password to customer via WhatsApp (admin only)"""
    # Check if it's a portal customer first
    customer = await db.portal_customers.find_one({'id': customer_id}, {'_id': 0})
    customer_type = 'portal'
    
    # If not found in portal_customers, check doctors/medicals/agencies
    if not customer:
        customer = await db.doctors.find_one({'id': customer_id}, {'_id': 0})
        customer_type = 'doctor'
    if not customer:
        customer = await db.medicals.find_one({'id': customer_id}, {'_id': 0})
        customer_type = 'medical'
    if not customer:
        customer = await db.agencies.find_one({'id': customer_id}, {'_id': 0})
        customer_type = 'agency'
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Generate new random password (8 chars alphanumeric)
    new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    password_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
    
    phone = customer.get('phone', '')
    clean_phone = ''.join(filter(str.isdigit, phone))
    
    if len(clean_phone) < 10:
        raise HTTPException(status_code=400, detail="Customer has no valid phone number")
    
    # Update password based on customer type
    if customer_type == 'portal':
        await db.portal_customers.update_one(
            {'id': customer_id},
            {'$set': {'password_hash': password_hash, 'password_sent_at': datetime.now(timezone.utc).isoformat()}}
        )
    else:
        # For admin-created customers, we need to create or update portal access
        existing_portal = await db.portal_customers.find_one({'phone': clean_phone}, {'_id': 0})
        if existing_portal:
            # Update existing portal account
            await db.portal_customers.update_one(
                {'phone': clean_phone},
                {'$set': {'password_hash': password_hash, 'password_sent_at': datetime.now(timezone.utc).isoformat()}}
            )
        else:
            # Create new portal account for admin-created customer
            now = datetime.now(timezone.utc)
            role_map = {'doctor': 'doctor', 'medical': 'medical', 'agency': 'agency'}
            portal_customer = {
                'id': str(uuid.uuid4()),
                'customer_code': customer.get('customer_code', await generate_customer_code(role_map.get(customer_type, 'doctor'))),
                'name': customer.get('name', ''),
                'phone': clean_phone,
                'email': customer.get('email'),
                'password_hash': password_hash,
                'role': role_map.get(customer_type, 'doctor'),
                'status': 'approved',  # Auto-approved since created by admin
                'reg_no': customer.get('reg_no'),
                'dob': customer.get('dob'),
                'proprietor_name': customer.get('proprietor_name'),
                'gst_number': customer.get('gst_number'),
                'drug_license': customer.get('drug_license'),
                'alternate_phone': customer.get('alternate_phone'),
                'birthday': customer.get('birthday'),
                'anniversary': customer.get('anniversary'),
                'address_line_1': customer.get('address_line_1'),
                'address_line_2': customer.get('address_line_2'),
                'state': customer.get('state'),
                'district': customer.get('district'),
                'pincode': customer.get('pincode'),
                'delivery_station': customer.get('delivery_station'),
                'transport_id': customer.get('transport_id'),
                'created_at': now.isoformat(),
                'approved_at': now.isoformat(),
                'approved_by': current_user.get('name', 'Admin'),
                'linked_record_id': customer_id,
                'password_sent_at': now.isoformat()
            }
            await db.portal_customers.insert_one(portal_customer)
            logger.info(f"Created portal account for admin customer: {customer.get('name')}")
    
    # Send new password via WhatsApp
    config = await get_whatsapp_config()
    if config.get('api_url') and config.get('auth_token') and config.get('sender_id'):
        try:
            message = f"Hello {customer.get('name', 'Customer')}!\n\n" \
                      f"Your VMP CRM portal login credentials:\n\n" \
                      f"*Phone:* {clean_phone}\n" \
                      f"*Password:* {new_password}\n\n" \
                      f"Login at the customer portal to view products and place orders.\n\n" \
                      f"Please change your password after first login for security."
            
            wa_mobile = clean_phone if clean_phone.startswith('91') else f"91{clean_phone[-10:]}"
            
            params = {
                'action': 'send',
                'senderId': config['sender_id'],
                'authToken': config['auth_token'],
                'messageText': message,
                'receiverId': wa_mobile
            }
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(config['api_url'], params=params)
                if response.status_code == 200:
                    await log_whatsapp_message(wa_mobile, 'password_reset', message, 'success', recipient_name=customer.get('name'))
                    return {"message": "New password sent via WhatsApp successfully", "password_sent": True}
                else:
                    logger.error(f"WhatsApp password send failed: {response.status_code}")
                    return {"message": "Password updated but WhatsApp delivery failed. Please share password manually.", "password": new_password, "password_sent": False}
        except Exception as e:
            logger.error(f"WhatsApp password send error: {str(e)}")
            return {"message": "Password updated but WhatsApp delivery failed. Please share password manually.", "password": new_password, "password_sent": False}
    else:
        # No WhatsApp configured - return password for manual sharing
        return {"message": "Password updated. WhatsApp not configured - please share password manually.", "password": new_password, "password_sent": False}


async def create_linked_customer_record(customer: dict, approved_by: str) -> Optional[str]:
    """Create a record in doctors/medicals/agencies when portal customer is approved"""
    now = datetime.now(timezone.utc)
    role = customer.get('role', '').lower()
    
    # Base record data
    record_id = str(uuid.uuid4())
    base_data = {
        'id': record_id,
        'customer_code': customer['customer_code'],
        'name': customer['name'],
        'phone': customer['phone'],
        'email': customer.get('email') or '',
        'address': customer.get('address') or '',
        'address_line_1': customer.get('address_line_1') or '',
        'address_line_2': customer.get('address_line_2') or '',
        'district': customer.get('district') or '',
        'state': customer.get('state') or '',
        'pincode': customer.get('pincode') or '',
        'delivery_station': customer.get('delivery_station') or '',
        'transport_id': customer.get('transport_id'),
        'lead_status': 'Customer',  # Mark as Customer (converted from portal)
        'is_portal_customer': True,  # Flag to identify portal customers
        'portal_customer_id': customer['id'],  # Link back to portal_customers
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': approved_by
    }
    
    try:
        if role == 'doctor':
            # Add doctor-specific fields
            base_data['reg_no'] = customer.get('reg_no') or ''
            base_data['dob'] = customer.get('dob')
            base_data['priority'] = 'moderate'
            await db.doctors.insert_one(base_data)
            logger.info(f"Created doctor record for portal customer: {customer['name']}")
            
        elif role == 'medical':
            # Add medical-specific fields
            base_data['proprietor_name'] = customer.get('proprietor_name') or ''
            base_data['gst_number'] = customer.get('gst_number') or ''
            base_data['drug_license'] = customer.get('drug_license') or ''
            base_data['priority'] = 'moderate'
            await db.medicals.insert_one(base_data)
            logger.info(f"Created medical record for portal customer: {customer['name']}")
            
        elif role == 'agency':
            # Add agency-specific fields
            base_data['proprietor_name'] = customer.get('proprietor_name') or ''
            base_data['gst_number'] = customer.get('gst_number') or ''
            base_data['drug_license'] = customer.get('drug_license') or ''
            base_data['priority'] = 'moderate'
            await db.agencies.insert_one(base_data)
            logger.info(f"Created agency record for portal customer: {customer['name']}")
        else:
            logger.warning(f"Unknown role '{role}' for portal customer: {customer['name']}")
            return None
            
        return record_id
    except Exception as e:
        logger.error(f"Failed to create linked record for portal customer: {str(e)}")
        return None

# Admin routes for support tickets

@api_router.get("/support/tickets")
async def get_all_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all support tickets (admin/staff)"""
    query = {}
    if status:
        query['status'] = status
    if priority:
        query['priority'] = priority
    
    tickets = await db.support_tickets.find(query, {'_id': 0}).sort('created_at', -1).to_list(500)
    
    result = []
    for ticket in tickets:
        created_at = ticket.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        updated_at = ticket.get('updated_at')
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
        
        result.append({
            **ticket,
            'created_at': created_at,
            'updated_at': updated_at
        })
    
    return result

@api_router.put("/support/tickets/{ticket_id}/status")
async def update_ticket_status(ticket_id: str, status: str, current_user: dict = Depends(get_current_user)):
    """Update ticket status and notify customer via WhatsApp"""
    if status not in ['open', 'in_progress', 'resolved', 'closed']:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    # Get ticket details first
    ticket = await db.support_tickets.find_one({'id': ticket_id}, {'_id': 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    old_status = ticket.get('status')
    
    now = datetime.now(timezone.utc)
    update_data = {'status': status, 'updated_at': now.isoformat()}
    
    if status == 'resolved':
        update_data['resolved_at'] = now.isoformat()
        update_data['resolved_by'] = current_user['name']
    
    result = await db.support_tickets.update_one({'id': ticket_id}, {'$set': update_data})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Send WhatsApp notification if status actually changed
    if old_status != status and ticket.get('customer_phone'):
        config = await get_whatsapp_config()
        if config.get('api_url') and config.get('auth_token') and config.get('sender_id'):
            try:
                # Get company name
                company = await db.company_settings.find_one({}, {'_id': 0})
                company_name = company.get('company_name', 'VETMECH PHARMA') if company else 'VETMECH PHARMA'
                
                # Status-specific messages
                ticket_num = ticket.get('ticket_number', ticket_id[:8])
                customer_name = ticket.get('customer_name', 'Customer')
                
                status_messages = {
                    'open': f"Dear {customer_name},\n\nYour support ticket #{ticket_num} has been *REOPENED*.\n\nSubject: {ticket.get('subject')}\n\nOur team will look into your issue shortly.\n\nRegards,\n{company_name}",
                    
                    'in_progress': f"Dear {customer_name},\n\nGood news! Your support ticket #{ticket_num} is now *IN PROGRESS*.\n\nSubject: {ticket.get('subject')}\n\nOur team is actively working on resolving your issue. We'll update you soon.\n\nRegards,\n{company_name}",
                    
                    'resolved': f"Dear {customer_name},\n\nGreat news! Your support ticket #{ticket_num} has been *RESOLVED*.\n\nSubject: {ticket.get('subject')}\n\nWe hope this resolves your concern. If you have any further questions, please don't hesitate to reach out.\n\nThank you for your patience!\n\nRegards,\n{company_name}",
                    
                    'closed': f"Dear {customer_name},\n\nYour support ticket #{ticket_num} has been *CLOSED*.\n\nSubject: {ticket.get('subject')}\n\nIf you need any further assistance, please create a new ticket or contact us directly.\n\nThank you for choosing {company_name}!\n\nRegards,\n{company_name}"
                }
                
                message = status_messages.get(status, f"Your ticket #{ticket_num} status has been updated to: {status.upper()}")
                
                # Send WhatsApp
                wa_mobile = ticket['customer_phone'] if ticket['customer_phone'].startswith('91') else f"91{ticket['customer_phone'][-10:]}"
                
                params = {
                    'action': 'send',
                    'senderId': config['sender_id'],
                    'authToken': config['auth_token'],
                    'messageText': message,
                    'receiverId': wa_mobile
                }
                
                async with httpx.AsyncClient(timeout=30.0) as client:
                    response = await client.get(config['api_url'], params=params)
                    if response.status_code == 200:
                        await log_whatsapp_message(wa_mobile, 'ticket_status', message, 'success', recipient_name=customer_name)
                        logger.info(f"Ticket status notification sent to {wa_mobile}")
                    else:
                        logger.error(f"WhatsApp notification failed: {response.status_code}")
                        
            except Exception as e:
                logger.error(f"WhatsApp notification error: {str(e)}")
    
    return {"message": "Ticket status updated"}

@api_router.post("/support/tickets/{ticket_id}/reply")
async def add_admin_ticket_reply(ticket_id: str, reply: TicketReply, current_user: dict = Depends(get_current_user)):
    """Add reply to ticket from admin/staff"""
    ticket = await db.support_tickets.find_one({'id': ticket_id}, {'_id': 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    now = datetime.now(timezone.utc)
    reply_doc = {
        'id': str(uuid.uuid4()),
        'message': reply.message,
        'sender_type': 'admin',
        'sender_name': current_user['name'],
        'created_at': now.isoformat()
    }
    
    await db.support_tickets.update_one(
        {'id': ticket_id},
        {
            '$push': {'replies': reply_doc},
            '$set': {'updated_at': now.isoformat(), 'status': 'in_progress'}
        }
    )
    
    # Notify customer via WhatsApp using BotMasterSender API
    config = await get_whatsapp_config()
    if config.get('api_url') and config.get('auth_token') and config.get('sender_id'):
        try:
            # Get company name
            company = await db.company_settings.find_one({}, {'_id': 0})
            company_name = company.get('company_name', 'VETMECH PHARMA') if company else 'VETMECH PHARMA'
            
            customer_name = ticket.get('customer_name', 'Customer')
            ticket_num = ticket.get('ticket_number', ticket_id[:8])
            
            message = f"Dear {customer_name},\n\nYou have received a new reply on your support ticket #{ticket_num}:\n\n\"{reply.message[:300]}{'...' if len(reply.message) > 300 else ''}\"\n\n- {current_user['name']}\n\nRegards,\n{company_name}"
            
            # Ensure mobile has 91 prefix for India
            wa_mobile = ticket['customer_phone'] if ticket['customer_phone'].startswith('91') else f"91{ticket['customer_phone'][-10:]}"
            
            params = {
                'action': 'send',
                'senderId': config['sender_id'],
                'authToken': config['auth_token'],
                'messageText': message,
                'receiverId': wa_mobile
            }
            
            async with httpx.AsyncClient(timeout=30.0) as client:
                response = await client.get(config['api_url'], params=params)
                if response.status_code == 200:
                    await log_whatsapp_message(wa_mobile, 'ticket_reply', message, 'success', recipient_name=customer_name)
                    logger.info(f"Ticket reply notification sent to {wa_mobile}")
        except Exception as e:
            logger.error(f"WhatsApp notification error: {str(e)}")
    
    return {"message": "Reply added successfully", "reply": reply_doc}

# Admin ticket creation model
class AdminTicketCreate(BaseModel):
    subject: str
    description: str
    category: str = "general"
    priority: str = "medium"
    customer_name: str
    customer_phone: str
    customer_email: Optional[str] = None

class TicketUpdate(BaseModel):
    subject: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[str] = None

@api_router.post("/support/tickets/admin")
async def create_admin_ticket(ticket_data: AdminTicketCreate, current_user: dict = Depends(get_current_user)):
    """Create a support ticket from admin panel (for walk-in/phone customers)"""
    now = datetime.now(timezone.utc)
    
    # Generate ticket number
    count = await db.support_tickets.count_documents({})
    ticket_number = f"TKT-{count + 1:05d}"
    
    ticket_doc = {
        'id': str(uuid.uuid4()),
        'ticket_number': ticket_number,
        'subject': ticket_data.subject,
        'description': ticket_data.description,
        'category': ticket_data.category,
        'priority': ticket_data.priority,
        'status': 'open',
        'customer_id': None,  # No portal customer linked
        'customer_name': ticket_data.customer_name,
        'customer_phone': ticket_data.customer_phone,
        'customer_email': ticket_data.customer_email,
        'replies': [],
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': current_user['name']
    }
    
    await db.support_tickets.insert_one(ticket_doc)
    del ticket_doc['_id']
    return ticket_doc

@api_router.put("/support/tickets/{ticket_id}")
async def update_ticket(ticket_id: str, ticket_data: TicketUpdate, current_user: dict = Depends(get_current_user)):
    """Update ticket details"""
    update_fields = {'updated_at': datetime.now(timezone.utc).isoformat()}
    
    if ticket_data.subject:
        update_fields['subject'] = ticket_data.subject
    if ticket_data.description:
        update_fields['description'] = ticket_data.description
    if ticket_data.category:
        update_fields['category'] = ticket_data.category
    if ticket_data.priority:
        update_fields['priority'] = ticket_data.priority
    
    result = await db.support_tickets.update_one(
        {'id': ticket_id},
        {'$set': update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    return {"message": "Ticket updated successfully"}

@api_router.delete("/support/tickets/{ticket_id}")
async def delete_ticket(ticket_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a support ticket"""
    result = await db.support_tickets.delete_one({'id': ticket_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    return {"message": "Ticket deleted successfully"}

# ============== OTP & ORDER ROUTES ==============

async def get_whatsapp_config():
    """Get WhatsApp config from database"""
    config = await db.whatsapp_config.find_one({}, {'_id': 0})
    if not config:
        # Default config
        return {
            'api_url': 'https://api.botmastersender.com/api/v1/',
            'auth_token': '1d97fa5b-b9f8-4b1c-9479-cae962594d5f',
            'sender_id': '919944472488'
        }
    return config

async def log_whatsapp_message(
    recipient_phone: str,
    message_type: str,
    message_preview: str,
    status: str,
    recipient_name: str = None,
    error_message: str = None
):
    """Log WhatsApp message to database"""
    try:
        log_doc = {
            'id': str(uuid.uuid4()),
            'recipient_phone': recipient_phone,
            'recipient_name': recipient_name,
            'message_type': message_type,
            'message_preview': message_preview[:500] if message_preview else '',  # Truncate long messages
            'status': status,
            'error_message': error_message,
            'created_at': datetime.now(timezone.utc).isoformat()
        }
        await db.whatsapp_logs.insert_one(log_doc)
    except Exception as e:
        logger.error(f"Failed to log WhatsApp message: {str(e)}")

async def send_whatsapp_otp(mobile: str, otp: str):
    """Send OTP via WhatsApp API"""
    config = await get_whatsapp_config()
    
    if not config:
        logger.warning("WhatsApp config not found, OTP not sent via WhatsApp")
        await log_whatsapp_message(mobile, 'otp', f"OTP: {otp}", 'failed', error_message='WhatsApp not configured')
        return True  # Return True to allow order flow to continue (OTP is logged)
    
    # Ensure mobile has 91 prefix
    clean_mobile = ''.join(filter(str.isdigit, mobile))
    if not clean_mobile.startswith('91'):
        clean_mobile = f"91{clean_mobile[-10:]}"
    
    message = f"Your VMP CRM verification code is: {otp}. Valid for 5 minutes."
    
    params = {
        'action': 'send',
        'senderId': config['sender_id'],
        'authToken': config['auth_token'],
        'messageText': message,
        'receiverId': clean_mobile
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(config['api_url'], params=params, timeout=30)
            logger.info(f"WhatsApp OTP sent to {clean_mobile}: {response.status_code}")
            status = 'success' if response.status_code == 200 else 'failed'
            await log_whatsapp_message(clean_mobile, 'otp', message, status)
            return response.status_code == 200
    except Exception as e:
        logger.error(f"WhatsApp OTP error: {str(e)}")
        await log_whatsapp_message(clean_mobile, 'otp', message, 'failed', error_message=str(e))
        return True  # Return True to allow flow to continue even if WhatsApp fails


# ============== FALLBACK OTP MANAGEMENT ==============

@api_router.get("/fallback-otps")
async def get_fallback_otps(current_user: dict = Depends(get_current_user)):
    """Get all fallback OTPs (admin only)"""
    otps = await db.fallback_otps.find({}, {'_id': 0}).sort('created_at', -1).to_list(100)
    return otps

@api_router.post("/fallback-otps")
async def create_fallback_otp(otp_data: FallbackOTPCreate, current_user: dict = Depends(get_current_user)):
    """Create a new fallback OTP (admin only)"""
    # Validate OTP format (4 digits)
    if not otp_data.otp.isdigit() or len(otp_data.otp) != 4:
        raise HTTPException(status_code=400, detail="OTP must be exactly 4 digits")
    
    # Check if OTP already exists
    existing = await db.fallback_otps.find_one({'otp': otp_data.otp}, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="This OTP already exists")
    
    otp_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    otp_doc = {
        'id': otp_id,
        'otp': otp_data.otp,
        'is_active': True,
        'used_count': 0,
        'created_at': now.isoformat(),
        'created_by': current_user['name']
    }
    
    await db.fallback_otps.insert_one(otp_doc)
    
    return FallbackOTPResponse(
        id=otp_id,
        otp=otp_data.otp,
        is_active=True,
        used_count=0,
        created_at=now,
        created_by=current_user['name']
    )

@api_router.put("/fallback-otps/{otp_id}/toggle")
async def toggle_fallback_otp(otp_id: str, current_user: dict = Depends(get_current_user)):
    """Toggle fallback OTP active status (admin only)"""
    otp = await db.fallback_otps.find_one({'id': otp_id}, {'_id': 0})
    if not otp:
        raise HTTPException(status_code=404, detail="OTP not found")
    
    new_status = not otp.get('is_active', True)
    await db.fallback_otps.update_one(
        {'id': otp_id},
        {'$set': {'is_active': new_status}}
    )
    
    return {"message": f"OTP {'activated' if new_status else 'deactivated'} successfully", "is_active": new_status}

@api_router.delete("/fallback-otps/{otp_id}")
async def delete_fallback_otp(otp_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a fallback OTP (admin only)"""
    result = await db.fallback_otps.delete_one({'id': otp_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="OTP not found")
    
    return {"message": "OTP deleted successfully"}

@api_router.post("/fallback-otps/seed")
async def seed_default_otps(current_user: dict = Depends(get_current_user)):
    """Seed default fallback OTPs (admin only)"""
    default_otps = ['0101', '3355', '8796', '5896', '5478', '9635', '1596', '1478', '9630', '8520', '7410']
    now = datetime.now(timezone.utc)
    added = 0
    
    for otp in default_otps:
        existing = await db.fallback_otps.find_one({'otp': otp}, {'_id': 0})
        if not existing:
            otp_doc = {
                'id': str(uuid.uuid4()),
                'otp': otp,
                'is_active': True,
                'used_count': 0,
                'created_at': now.isoformat(),
                'created_by': current_user['name']
            }
            await db.fallback_otps.insert_one(otp_doc)
            added += 1
    
    return {"message": f"Added {added} new fallback OTPs", "total_default": len(default_otps)}


# ============== MARKETING MODULE ==============

def generate_reference_number():
    """Generate random 7-digit reference number for anti-ban"""
    return str(random.randint(1000000, 9999999))

@api_router.get("/marketing/templates")
async def get_marketing_templates(current_user: dict = Depends(get_current_user)):
    """Get all marketing templates"""
    templates = await db.marketing_templates.find({}, {'_id': 0}).sort('created_at', -1).to_list(100)
    return templates

@api_router.post("/marketing/templates")
async def create_marketing_template(template: MarketingTemplateCreate, current_user: dict = Depends(get_current_user)):
    """Create a new marketing template"""
    template_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    template_doc = {
        'id': template_id,
        'name': template.name,
        'category': template.category,
        'message': template.message,
        'is_active': template.is_active,
        'created_at': now.isoformat(),
        'created_by': current_user['name']
    }
    
    await db.marketing_templates.insert_one(template_doc)
    # Return without _id
    del template_doc['_id']
    return template_doc

@api_router.put("/marketing/templates/{template_id}")
async def update_marketing_template(template_id: str, template: MarketingTemplateCreate, current_user: dict = Depends(get_current_user)):
    """Update a marketing template"""
    result = await db.marketing_templates.update_one(
        {'id': template_id},
        {'$set': {
            'name': template.name,
            'category': template.category,
            'message': template.message,
            'is_active': template.is_active,
            'updated_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template updated successfully"}

@api_router.delete("/marketing/templates/{template_id}")
async def delete_marketing_template(template_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a marketing template"""
    result = await db.marketing_templates.delete_one({'id': template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template deleted successfully"}

@api_router.get("/marketing/recipients")
async def get_marketing_recipients(
    entity_type: str = "all",
    status: str = "all",
    current_user: dict = Depends(get_current_user)
):
    """Get potential recipients for marketing based on filters"""
    recipients = []
    
    # Build status filter
    status_filter = {}
    if status != "all":
        status_map = {
            "pipeline": "Pipeline",
            "customer": "Customer",
            "contacted": "Contacted",
            "not_interested": "Not Interested",
            "closed": "Closed"
        }
        if status in status_map:
            status_filter['lead_status'] = status_map[status]
    
    async def fetch_from_collection(collection_name, recipient_type):
        query = {**status_filter}
        docs = await db[collection_name].find(query, {'_id': 0}).to_list(1000)
        for doc in docs:
            if doc.get('phone'):
                recipients.append({
                    'id': doc.get('id'),
                    'name': doc.get('name'),
                    'phone': doc.get('phone'),
                    'email': doc.get('email'),
                    'type': recipient_type,
                    'lead_status': doc.get('lead_status', 'Pipeline'),
                    'customer_code': doc.get('customer_code', '')
                })
    
    if entity_type in ['all', 'doctors']:
        await fetch_from_collection('doctors', 'doctor')
    if entity_type in ['all', 'medicals']:
        await fetch_from_collection('medicals', 'medical')
    if entity_type in ['all', 'agencies']:
        await fetch_from_collection('agencies', 'agency')
    
    return recipients

@api_router.get("/marketing/campaigns")
async def get_marketing_campaigns(
    status: str = None,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get all marketing campaigns"""
    query = {}
    if status:
        query['status'] = status
    
    # Exclude image_webp binary data from list
    campaigns = await db.marketing_campaigns.find(query, {'_id': 0, 'image_webp': 0}).sort('created_at', -1).skip(skip).limit(limit).to_list(limit)
    
    # Add image_url for campaigns with images
    for campaign in campaigns:
        if campaign.get('has_image'):
            campaign['image_url'] = f"/api/marketing/campaigns/{campaign['id']}/image"
    
    total = await db.marketing_campaigns.count_documents(query)
    
    return {"campaigns": campaigns, "total": total}

@api_router.get("/marketing/campaigns/{campaign_id}")
async def get_marketing_campaign(campaign_id: str, current_user: dict = Depends(get_current_user)):
    """Get campaign details with logs"""
    campaign = await db.marketing_campaigns.find_one({'id': campaign_id}, {'_id': 0, 'image_webp': 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    # Add image_url if campaign has image
    if campaign.get('has_image'):
        campaign['image_url'] = f"/api/marketing/campaigns/{campaign_id}/image"
    
    # Get campaign logs
    logs = await db.campaign_logs.find({'campaign_id': campaign_id}, {'_id': 0}).sort('sent_at', -1).to_list(1000)
    
    return {"campaign": campaign, "logs": logs}

@api_router.get("/marketing/campaigns/{campaign_id}/image")
async def get_campaign_image(campaign_id: str):
    """Get campaign image"""
    campaign = await db.marketing_campaigns.find_one({'id': campaign_id}, {'image_webp': 1})
    if not campaign or not campaign.get('image_webp'):
        raise HTTPException(status_code=404, detail="Image not found")
    
    return Response(content=campaign['image_webp'], media_type="image/webp")

@api_router.post("/marketing/campaigns")
async def create_marketing_campaign(campaign: MarketingCampaignCreate, current_user: dict = Depends(get_current_user)):
    """Create a new marketing campaign"""
    campaign_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Get company settings for message footer
    company = await db.company_settings.find_one({}, {'_id': 0})
    company_name = company.get('company_name', 'VETMECH PHARMA') if company else 'VETMECH PHARMA'
    
    # Build message with item details if product promo
    final_message = campaign.message
    item_details = []
    
    if campaign.campaign_type == 'product_promo' and campaign.item_ids:
        # Fetch items without image_webp binary for storage, but check if they have images
        items = await db.items.find({'id': {'$in': campaign.item_ids}}, {'_id': 0, 'image_webp': 0}).to_list(len(campaign.item_ids))
        
        # Add has_image flag to each item
        for item in items:
            has_img = await db.items.find_one({'id': item['id'], 'image_webp': {'$ne': None}}, {'_id': 1})
            item['has_image'] = has_img is not None
        
        item_details = items
    
    # Handle image upload - store in MongoDB as webp
    processed_image = None
    if campaign.image_base64:
        try:
            image_data = campaign.image_base64.split(',')[1] if ',' in campaign.image_base64 else campaign.image_base64
            image_bytes = base64.b64decode(image_data)
            processed_image = process_image_to_webp(image_bytes, max_size_kb=200, target_size=(800, 800))
        except Exception as e:
            logger.error(f"Failed to process campaign image: {e}")
    
    # Determine initial status
    status = 'scheduled' if campaign.scheduled_at else 'draft'
    
    campaign_doc = {
        'id': campaign_id,
        'name': campaign.name,
        'campaign_type': campaign.campaign_type,
        'target_entity': campaign.target_entity,
        'target_status': campaign.target_status,
        'recipient_ids': campaign.recipient_ids,
        'total_recipients': len(campaign.recipient_ids),
        'sent_count': 0,
        'failed_count': 0,
        'pending_count': len(campaign.recipient_ids),
        'message': campaign.message,
        'message_preview': campaign.message[:100] + '...' if len(campaign.message) > 100 else campaign.message,
        'item_ids': campaign.item_ids or [],
        'item_details': item_details,
        'image_webp': processed_image,
        'has_image': processed_image is not None,
        'batch_size': campaign.batch_size,
        'batch_delay_seconds': campaign.batch_delay_seconds,
        'status': status,
        'scheduled_at': campaign.scheduled_at,
        'started_at': None,
        'completed_at': None,
        'created_at': now.isoformat(),
        'created_by': current_user['name'],
        'company_name': company_name
    }
    
    await db.marketing_campaigns.insert_one(campaign_doc)
    
    # Create pending logs for each recipient
    for recipient_id in campaign.recipient_ids:
        log_doc = {
            'id': str(uuid.uuid4()),
            'campaign_id': campaign_id,
            'recipient_id': recipient_id,
            'reference_number': generate_reference_number(),
            'status': 'pending',
            'error_message': None,
            'sent_at': None
        }
        await db.campaign_logs.insert_one(log_doc)
    
    # Return without _id
    del campaign_doc['_id']
    return campaign_doc

@api_router.post("/marketing/campaigns/{campaign_id}/send")
async def send_marketing_campaign(campaign_id: str, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Start sending a marketing campaign"""
    campaign = await db.marketing_campaigns.find_one({'id': campaign_id}, {'_id': 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    if campaign['status'] in ['sending', 'completed']:
        raise HTTPException(status_code=400, detail=f"Campaign is already {campaign['status']}")
    
    # Update status to sending
    await db.marketing_campaigns.update_one(
        {'id': campaign_id},
        {'$set': {
            'status': 'sending',
            'started_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Start background task to send messages
    background_tasks.add_task(process_marketing_campaign, campaign_id)
    
    return {"message": "Campaign sending started", "status": "sending"}

async def process_marketing_campaign(campaign_id: str):
    """Background task to process and send marketing messages in batches"""
    try:
        campaign = await db.marketing_campaigns.find_one({'id': campaign_id}, {'_id': 0})
        if not campaign:
            return
        
        config = await get_whatsapp_config()
        if not config.get('api_url') or not config.get('auth_token'):
            logger.error("WhatsApp not configured for marketing campaign")
            await db.marketing_campaigns.update_one(
                {'id': campaign_id},
                {'$set': {'status': 'failed', 'error': 'WhatsApp not configured'}}
            )
            return
        
        batch_size = campaign.get('batch_size', 10)
        batch_delay = campaign.get('batch_delay_seconds', 60)
        company_name = campaign.get('company_name', 'VETMECH PHARMA')
        
        # Get pending logs
        pending_logs = await db.campaign_logs.find({
            'campaign_id': campaign_id,
            'status': 'pending'
        }, {'_id': 0}).to_list(1000)
        
        # Process in batches
        for i in range(0, len(pending_logs), batch_size):
            batch = pending_logs[i:i + batch_size]
            
            for log in batch:
                try:
                    # Get recipient details
                    recipient = None
                    for collection in ['doctors', 'medicals', 'agencies']:
                        recipient = await db[collection].find_one({'id': log['recipient_id']}, {'_id': 0})
                        if recipient:
                            recipient['type'] = collection[:-1] if collection != 'agencies' else 'agency'
                            break
                    
                    if not recipient or not recipient.get('phone'):
                        await db.campaign_logs.update_one(
                            {'id': log['id']},
                            {'$set': {'status': 'failed', 'error_message': 'Recipient not found or no phone'}}
                        )
                        await db.marketing_campaigns.update_one(
                            {'id': campaign_id},
                            {'$inc': {'failed_count': 1, 'pending_count': -1}}
                        )
                        continue
                    
                    # Build personalized message
                    message = campaign['message']
                    
                    # Replace placeholders
                    message = message.replace('{name}', recipient.get('name', 'Customer'))
                    message = message.replace('{customer_code}', recipient.get('customer_code', ''))
                    
                    # Add item details for product promos
                    if campaign['campaign_type'] == 'product_promo' and campaign.get('item_details'):
                        items_text = "\n\n*Product Details:*\n"
                        for item in campaign['item_details']:
                            # Get role-specific pricing
                            recipient_type = recipient.get('type', 'doctor')
                            rate_field = f"rate_{recipient_type}s" if recipient_type != 'agency' else 'rate_agencies'
                            offer_field = f"offer_{recipient_type}s" if recipient_type != 'agency' else 'offer_agencies'
                            special_offer_field = f"special_offer_{recipient_type}s" if recipient_type != 'agency' else 'special_offer_agencies'
                            
                            # Use role-specific rate, fallback to default rate
                            rate = item.get(rate_field) or item.get('rate', 0)
                            offer = item.get(offer_field) or item.get('offer', '')
                            special_offer = item.get(special_offer_field) or item.get('special_offer', '')
                            mrp = item.get('mrp', 0)
                            item_name = item.get('name') or item.get('item_code', 'Product')
                            
                            items_text += f"\n*{item_name}*"
                            items_text += f"\nMRP: Rs.{mrp}"
                            items_text += f"\nYour Rate: Rs.{rate}"
                            if offer:
                                items_text += f"\nOffer: {offer}"
                            if special_offer:
                                items_text += f"\nSpecial: {special_offer}"
                            items_text += "\n"
                        
                        message += items_text
                    
                    # Add footer with reference number
                    ref_number = log['reference_number']
                    message += f"\n\nRegards,\n*{company_name}*\n(#Server Ref: {ref_number})"
                    
                    # Send WhatsApp message
                    phone = recipient['phone']
                    wa_mobile = phone if phone.startswith('91') else f"91{phone[-10:]}"
                    
                    # Determine image URL to use
                    image_url_to_send = None
                    
                    # Check if campaign has uploaded image
                    if campaign.get('has_image', False):
                        image_url_to_send = f"https://distributor-hub-17.preview.emergentagent.com/api/marketing/campaigns/{campaign_id}/image"
                    # For product promotions, use first item's image as default
                    elif campaign['campaign_type'] == 'product_promo' and campaign.get('item_details'):
                        for item in campaign['item_details']:
                            if item.get('has_image'):
                                image_url_to_send = f"https://distributor-hub-17.preview.emergentagent.com/api/items/{item['id']}/image"
                                break
                    
                    if image_url_to_send:
                        # Send image with caption
                        # BotMasterSender uses 'sendFile' action for media
                        params = {
                            'action': 'sendFile',
                            'senderId': config['sender_id'],
                            'authToken': config['auth_token'],
                            'fileUrl': image_url_to_send,
                            'fileCaption': message,
                            'receiverId': wa_mobile
                        }
                    else:
                        # Send text-only message
                        params = {
                            'action': 'send',
                            'senderId': config['sender_id'],
                            'authToken': config['auth_token'],
                            'messageText': message,
                            'receiverId': wa_mobile
                        }
                    
                    async with httpx.AsyncClient(timeout=30.0) as client:
                        response = await client.get(config['api_url'], params=params)
                        
                        if response.status_code == 200:
                            await db.campaign_logs.update_one(
                                {'id': log['id']},
                                {'$set': {
                                    'status': 'sent',
                                    'sent_at': datetime.now(timezone.utc).isoformat(),
                                    'recipient_name': recipient.get('name'),
                                    'recipient_phone': wa_mobile,
                                    'recipient_type': recipient.get('type')
                                }}
                            )
                            await db.marketing_campaigns.update_one(
                                {'id': campaign_id},
                                {'$inc': {'sent_count': 1, 'pending_count': -1}}
                            )
                            
                            # Log to WhatsApp logs
                            await log_whatsapp_message(
                                wa_mobile, 'marketing', message[:200], 'success',
                                recipient_name=recipient.get('name')
                            )
                        else:
                            error_msg = f"Status {response.status_code}"
                            await db.campaign_logs.update_one(
                                {'id': log['id']},
                                {'$set': {
                                    'status': 'failed',
                                    'error_message': error_msg,
                                    'recipient_name': recipient.get('name'),
                                    'recipient_phone': wa_mobile,
                                    'recipient_type': recipient.get('type')
                                }}
                            )
                            await db.marketing_campaigns.update_one(
                                {'id': campaign_id},
                                {'$inc': {'failed_count': 1, 'pending_count': -1}}
                            )
                    
                    # Small delay between messages in same batch
                    await asyncio.sleep(2)
                    
                except Exception as e:
                    logger.error(f"Error sending to recipient {log['recipient_id']}: {e}")
                    await db.campaign_logs.update_one(
                        {'id': log['id']},
                        {'$set': {'status': 'failed', 'error_message': str(e)}}
                    )
                    await db.marketing_campaigns.update_one(
                        {'id': campaign_id},
                        {'$inc': {'failed_count': 1, 'pending_count': -1}}
                    )
            
            # Delay between batches to avoid ban
            if i + batch_size < len(pending_logs):
                logger.info(f"Campaign {campaign_id}: Batch complete, waiting {batch_delay}s before next batch")
                await asyncio.sleep(batch_delay)
        
        # Mark campaign as completed
        await db.marketing_campaigns.update_one(
            {'id': campaign_id},
            {'$set': {
                'status': 'completed',
                'completed_at': datetime.now(timezone.utc).isoformat()
            }}
        )
        logger.info(f"Marketing campaign {campaign_id} completed")
        
    except Exception as e:
        logger.error(f"Marketing campaign {campaign_id} failed: {e}")
        await db.marketing_campaigns.update_one(
            {'id': campaign_id},
            {'$set': {'status': 'failed', 'error': str(e)}}
        )

@api_router.post("/marketing/campaigns/{campaign_id}/cancel")
async def cancel_marketing_campaign(campaign_id: str, current_user: dict = Depends(get_current_user)):
    """Cancel a marketing campaign"""
    campaign = await db.marketing_campaigns.find_one({'id': campaign_id}, {'_id': 0})
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    
    if campaign['status'] == 'completed':
        raise HTTPException(status_code=400, detail="Cannot cancel completed campaign")
    
    await db.marketing_campaigns.update_one(
        {'id': campaign_id},
        {'$set': {'status': 'cancelled'}}
    )
    
    return {"message": "Campaign cancelled"}

@api_router.get("/marketing/stats")
async def get_marketing_stats(current_user: dict = Depends(get_current_user)):
    """Get marketing statistics"""
    total_campaigns = await db.marketing_campaigns.count_documents({})
    completed_campaigns = await db.marketing_campaigns.count_documents({'status': 'completed'})
    
    # Get total messages sent this month
    start_of_month = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    sent_this_month = await db.campaign_logs.count_documents({
        'status': 'sent',
        'sent_at': {'$gte': start_of_month.isoformat()}
    })
    
    return {
        'total_campaigns': total_campaigns,
        'completed_campaigns': completed_campaigns,
        'messages_sent_this_month': sent_this_month
    }


async def send_whatsapp_order(mobile: str, items: List[OrderItem], order_number: str, doctor_name: str = None, ip_address: str = None, location: str = None):
    """Send order confirmation via WhatsApp with full details"""
    config = await get_whatsapp_config()
    
    # Get company settings
    company = await db.company_settings.find_one({}, {'_id': 0})
    company_name = company.get('company_name', 'VMP CRM') if company else 'VMP CRM'
    company_phone = config.get('sender_id', '')
    
    clean_mobile = ''.join(filter(str.isdigit, mobile))
    if not clean_mobile.startswith('91'):
        clean_mobile = f"91{clean_mobile[-10:]}"
    
    # Build personalized order message
    greeting = f"Dear Dr. {doctor_name}" if doctor_name else "Dear Customer"
    
    # Build items list
    items_text = "\n".join([f"• {item.item_name} - Qty: {item.quantity}" for item in items if item.quantity])
    
    # Location info
    location_text = ""
    if ip_address:
        location_text += f"\n📍 IP: {ip_address}"
    if location:
        location_text += f"\n📍 Location: {location}"
    
    message = f"""{greeting},

We have received your order. Kindly check it once again, we will start processing your order.

📋 *Order No:* {order_number}

*Order Details:*
{items_text}
{location_text}

Thank you for your order!

Regards,
*{company_name}*
📞 +{company_phone}"""
    
    params = {
        'action': 'send',
        'senderId': config['sender_id'],
        'authToken': config['auth_token'],
        'messageText': message,
        'receiverId': clean_mobile
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(config['api_url'], params=params, timeout=30)
            logger.info(f"Order confirmation sent to {clean_mobile}")
            status = 'success' if response.status_code == 200 else 'failed'
            await log_whatsapp_message(clean_mobile, 'order_confirmation', message, status, recipient_name=doctor_name)
    except Exception as e:
        logger.error(f"WhatsApp order message error: {str(e)}")
        await log_whatsapp_message(clean_mobile, 'order_confirmation', message, 'failed', recipient_name=doctor_name, error_message=str(e))

async def send_order_confirmation_email(order_doc: dict, items: list):
    """Send order confirmation email to customer"""
    try:
        # Check if customer has email
        customer_email = order_doc.get('doctor_email')
        if not customer_email:
            logger.info(f"No email address for order {order_doc.get('order_number')}, skipping email")
            return
        
        # Get SMTP config
        smtp_config = await db.smtp_config.find_one({}, {'_id': 0})
        if not smtp_config:
            logger.warning("SMTP not configured, skipping order confirmation email")
            return
        
        # Get company settings
        company = await db.company_settings.find_one({}, {'_id': 0})
        company_name = company.get('company_name', 'VMP CRM') if company else 'VMP CRM'
        company_address = company.get('address', '') if company else ''
        company_phone = company.get('phone', '') if company else ''
        company_email = company.get('email', '') if company else ''
        
        customer_name = order_doc.get('doctor_name', 'Customer')
        customer_phone = order_doc.get('doctor_phone', '')
        customer_address = order_doc.get('doctor_address', '')
        order_number = order_doc.get('order_number', '')
        order_date = datetime.now(timezone.utc).strftime('%d %b %Y, %I:%M %p')
        
        # Build items table rows
        items_rows = ""
        total_amount = 0
        for idx, item in enumerate(items, 1):
            item_name = item.get('item_name', '') if isinstance(item, dict) else getattr(item, 'item_name', '')
            item_code = item.get('item_code', '') if isinstance(item, dict) else getattr(item, 'item_code', '')
            quantity = item.get('quantity', '1') if isinstance(item, dict) else getattr(item, 'quantity', '1')
            rate = item.get('rate', 0) if isinstance(item, dict) else getattr(item, 'rate', 0)
            mrp = item.get('mrp', 0) if isinstance(item, dict) else getattr(item, 'mrp', 0)
            
            # Calculate amount (handle scheme quantities like "10+5")
            try:
                if '+' in str(quantity):
                    qty_parts = str(quantity).split('+')
                    total_qty = sum(int(q.strip()) for q in qty_parts)
                else:
                    total_qty = int(quantity)
                amount = float(rate or 0) * total_qty
                total_amount += amount
            except:
                amount = 0
            
            items_rows += f"""
            <tr>
                <td style="padding: 12px; border-bottom: 1px solid #eee;">{idx}</td>
                <td style="padding: 12px; border-bottom: 1px solid #eee;">{item_code}</td>
                <td style="padding: 12px; border-bottom: 1px solid #eee;">{item_name}</td>
                <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: center;">{quantity}</td>
                <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right;">₹{rate:.2f}</td>
                <td style="padding: 12px; border-bottom: 1px solid #eee; text-align: right;">₹{amount:.2f}</td>
            </tr>"""
        
        # Build HTML email
        html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; font-family: Arial, sans-serif; background-color: #f5f5f5;">
    <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff;">
        <!-- Header -->
        <div style="background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%); padding: 30px; text-align: center;">
            <h1 style="color: #ffffff; margin: 0; font-size: 24px;">{company_name}</h1>
            <p style="color: #b8d4e8; margin: 10px 0 0 0; font-size: 14px;">Order Confirmation</p>
        </div>
        
        <!-- Order Status Banner -->
        <div style="background-color: #10b981; padding: 15px; text-align: center;">
            <span style="color: #ffffff; font-size: 16px; font-weight: bold;">✓ Order Received Successfully!</span>
        </div>
        
        <!-- Content -->
        <div style="padding: 30px;">
            <!-- Greeting -->
            <p style="color: #333; font-size: 16px; margin-bottom: 20px;">
                Dear <strong>{customer_name}</strong>,
            </p>
            <p style="color: #666; font-size: 14px; line-height: 1.6;">
                Thank you for your order! We have received your order and will start processing it shortly.
            </p>
            
            <!-- Order Info Box -->
            <div style="background-color: #f8fafc; border-radius: 8px; padding: 20px; margin: 20px 0;">
                <table style="width: 100%; border-collapse: collapse;">
                    <tr>
                        <td style="padding: 8px 0;">
                            <strong style="color: #64748b;">Order Number:</strong>
                        </td>
                        <td style="padding: 8px 0; text-align: right;">
                            <strong style="color: #1e3a5f;">{order_number}</strong>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0;">
                            <strong style="color: #64748b;">Order Date:</strong>
                        </td>
                        <td style="padding: 8px 0; text-align: right;">
                            <span style="color: #333;">{order_date}</span>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 0;">
                            <strong style="color: #64748b;">Phone:</strong>
                        </td>
                        <td style="padding: 8px 0; text-align: right;">
                            <span style="color: #333;">{customer_phone}</span>
                        </td>
                    </tr>
                    {f'''<tr>
                        <td style="padding: 8px 0;">
                            <strong style="color: #64748b;">Address:</strong>
                        </td>
                        <td style="padding: 8px 0; text-align: right;">
                            <span style="color: #333;">{customer_address}</span>
                        </td>
                    </tr>''' if customer_address else ''}
                </table>
            </div>
            
            <!-- Items Table -->
            <h3 style="color: #1e3a5f; margin: 25px 0 15px 0; font-size: 16px;">Order Items</h3>
            <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                <thead>
                    <tr style="background-color: #1e3a5f; color: #ffffff;">
                        <th style="padding: 12px; text-align: left;">#</th>
                        <th style="padding: 12px; text-align: left;">Code</th>
                        <th style="padding: 12px; text-align: left;">Item Name</th>
                        <th style="padding: 12px; text-align: center;">Qty</th>
                        <th style="padding: 12px; text-align: right;">Rate</th>
                        <th style="padding: 12px; text-align: right;">Amount</th>
                    </tr>
                </thead>
                <tbody>
                    {items_rows}
                </tbody>
                <tfoot>
                    <tr style="background-color: #f8fafc;">
                        <td colspan="5" style="padding: 15px; text-align: right; font-weight: bold; color: #1e3a5f;">
                            Total Amount:
                        </td>
                        <td style="padding: 15px; text-align: right; font-weight: bold; color: #10b981; font-size: 16px;">
                            ₹{total_amount:.2f}
                        </td>
                    </tr>
                </tfoot>
            </table>
            
            <!-- Note -->
            <div style="background-color: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 25px 0; border-radius: 0 8px 8px 0;">
                <p style="margin: 0; color: #92400e; font-size: 13px;">
                    <strong>Note:</strong> This is an order confirmation. The final invoice will be shared once your order is shipped.
                </p>
            </div>
        </div>
        
        <!-- Footer -->
        <div style="background-color: #1e3a5f; padding: 25px; text-align: center;">
            <p style="color: #ffffff; margin: 0 0 5px 0; font-weight: bold;">{company_name}</p>
            <p style="color: #b8d4e8; margin: 0; font-size: 12px;">{company_address}</p>
            {f'<p style="color: #b8d4e8; margin: 5px 0 0 0; font-size: 12px;">📞 {company_phone}</p>' if company_phone else ''}
            {f'<p style="color: #b8d4e8; margin: 5px 0 0 0; font-size: 12px;">✉️ {company_email}</p>' if company_email else ''}
        </div>
    </div>
</body>
</html>
"""
        
        # Create email message
        msg = MIMEMultipart()
        msg['From'] = f"{smtp_config['from_name']} <{smtp_config['from_email']}>"
        msg['To'] = f"{customer_name} <{customer_email}>"
        msg['Subject'] = f"Order Confirmation - {order_number} | {company_name}"
        msg.attach(MIMEText(html_body, 'html'))
        
        # Log entry for tracking
        log_id = str(uuid.uuid4())
        email_log = {
            'id': log_id,
            'doctor_id': order_doc.get('doctor_id'),
            'doctor_name': customer_name,
            'doctor_email': customer_email,
            'subject': f"Order Confirmation - {order_number}",
            'body': f"Order confirmation email for {order_number}",
            'status': 'pending',
            'created_at': datetime.now(timezone.utc).isoformat(),
            'sent_by': 'system'
        }
        await db.email_logs.insert_one(email_log)
        
        # Send email
        try:
            with smtplib.SMTP(smtp_config['smtp_server'], smtp_config['smtp_port'], timeout=30) as server:
                server.starttls()
                server.login(smtp_config['smtp_username'], smtp_config['smtp_password'])
                server.sendmail(smtp_config['from_email'], [customer_email], msg.as_string())
            
            logger.info(f"Order confirmation email sent to {customer_email} for order {order_number}")
            
            # Update log as sent
            await db.email_logs.update_one(
                {'id': log_id},
                {'$set': {'status': 'sent', 'sent_at': datetime.now(timezone.utc).isoformat()}}
            )
        except Exception as smtp_error:
            logger.error(f"SMTP error sending order confirmation email: {str(smtp_error)}")
            await db.email_logs.update_one(
                {'id': log_id},
                {'$set': {'status': 'failed', 'error_message': str(smtp_error)}}
            )
        
    except Exception as e:
        logger.error(f"Failed to send order confirmation email: {str(e)}")

async def send_whatsapp_status_update(order: dict, new_status: str, update_data: dict = None):
    """Send WhatsApp notification for order status changes"""
    config = await get_whatsapp_config()
    if not config:
        logger.warning("WhatsApp config not found, skipping notification")
        return
    
    # Get company settings
    company = await db.company_settings.find_one({}, {'_id': 0})
    company_name = company.get('company_name', 'VMP CRM') if company else 'VMP CRM'
    company_phone = config.get('sender_id', '')
    
    mobile = order.get('doctor_phone', '')
    clean_mobile = ''.join(filter(str.isdigit, mobile))
    if not clean_mobile.startswith('91'):
        clean_mobile = f"91{clean_mobile[-10:]}"
    
    doctor_name = order.get('doctor_name')
    greeting = f"Dear Dr. {doctor_name}" if doctor_name else "Dear Customer"
    order_number = order.get('order_number', '')
    
    # Build items list
    items = order.get('items', [])
    items_text = "\n".join([f"• {item.get('item_name', '')} - Qty: {item.get('quantity', '')}" for item in items])
    
    message = ""
    
    if new_status == 'confirmed':
        message = f"""{greeting},

✅ Your order has been *CONFIRMED*!

📋 *Order No:* {order_number}

*Order Details:*
{items_text}

We are processing your order and will ship it soon.

Regards,
*{company_name}*
📞 +{company_phone}"""

    elif new_status == 'shipped':
        # Get tracking number from update_data (new) or order (existing)
        tracking_number = update_data.get('tracking_number') if update_data else None
        if not tracking_number:
            tracking_number = order.get('tracking_number', 'N/A')
        
        # Get transport info from order (set during ready_to_despatch) or update_data
        transport_name = order.get('transport_name') or (update_data.get('transport_name') if update_data else None) or 'N/A'
        delivery_station = order.get('delivery_station') or (update_data.get('delivery_station') if update_data else None) or 'N/A'
        payment_mode = order.get('payment_mode') or (update_data.get('payment_mode') if update_data else None) or ''
        payment_text = "To Pay" if payment_mode == 'to_pay' else "Paid" if payment_mode == 'paid' else 'N/A'
        
        # Get package details from order (set during ready_to_despatch)
        boxes = order.get('boxes_count', 0) or 0
        cans = order.get('cans_count', 0) or 0
        bags = order.get('bags_count', 0) or 0
        
        package_parts = []
        if boxes: package_parts.append(f"{boxes} Box(es)")
        if cans: package_parts.append(f"{cans} Can(s)")
        if bags: package_parts.append(f"{bags} Bag(s)")
        package_text = ", ".join(package_parts) if package_parts else "N/A"
        
        # Get invoice details from order (set during ready_to_despatch)
        invoice_number = order.get('invoice_number', 'N/A') or 'N/A'
        invoice_date = order.get('invoice_date', 'N/A') or 'N/A'
        invoice_value = order.get('invoice_value', 0) or 0
        invoice_value_text = f"₹{invoice_value:,.2f}" if invoice_value else 'N/A'
        
        message = f"""{greeting},

🚚 Your order has been *SHIPPED*!

📋 *Order No:* {order_number}

*Order Details:*
{items_text}

*Shipping Information:*
🚛 Transport: {transport_name}
📦 Tracking No: {tracking_number}
📍 Delivery Station: {delivery_station}
💰 Payment: {payment_text}

*Package Details:*
📦 {package_text}

*Invoice Details:*
🧾 Invoice No: {invoice_number}
📅 Invoice Date: {invoice_date}
💵 Invoice Value: {invoice_value_text}

Your order is on its way! Thank you for your order.

Regards,
*{company_name}*
📞 +{company_phone}"""

    elif new_status == 'delivered':
        message = f"""{greeting},

🎉 Your order has been *DELIVERED*!

📋 *Order No:* {order_number}

*Order Details:*
{items_text}

Thank you for choosing us! We hope you are satisfied with your order.

For any queries, please contact us.

Regards,
*{company_name}*
📞 +{company_phone}"""

    elif new_status == 'cancelled':
        cancellation_reason = update_data.get('cancellation_reason', 'Not specified') if update_data else 'Not specified'
        
        message = f"""{greeting},

❌ Your order has been *CANCELLED*

📋 *Order No:* {order_number}

*Order Details:*
{items_text}

*Reason for Cancellation:*
{cancellation_reason}

If you have any questions, please contact us.

Regards,
*{company_name}*
📞 +{company_phone}"""
    
    if not message:
        return
    
    params = {
        'action': 'send',
        'senderId': config['sender_id'],
        'authToken': config['auth_token'],
        'messageText': message,
        'receiverId': clean_mobile
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(config['api_url'], params=params, timeout=30)
            logger.info(f"Order status update ({new_status}) sent to {clean_mobile}")
            status = 'success' if response.status_code == 200 else 'failed'
            await log_whatsapp_message(clean_mobile, f'status_{new_status}', message, status, recipient_name=doctor_name)
    except Exception as e:
        logger.error(f"WhatsApp status update error: {str(e)}")
        await log_whatsapp_message(clean_mobile, f'status_{new_status}', message, 'failed', recipient_name=doctor_name, error_message=str(e))

async def send_whatsapp_ready_to_despatch(order: dict, update_data: dict):
    """Send WhatsApp notification for Ready to Despatch status - to both transporter and customer"""
    config = await get_whatsapp_config()
    if not config:
        logger.warning("WhatsApp config not found, skipping notification")
        return
    
    # Get company settings
    company = await db.company_settings.find_one({}, {'_id': 0})
    company_name = company.get('company_name', 'VMP CRM') if company else 'VMP CRM'
    company_phone = config.get('sender_id', '')
    
    order_number = order.get('order_number', '')
    doctor_name = order.get('doctor_name', 'Customer')
    doctor_phone = order.get('doctor_phone', '')
    doctor_address = order.get('doctor_address', 'N/A')
    
    # Get transport details
    transport_name = update_data.get('transport_name', 'N/A')
    delivery_station = update_data.get('delivery_station', 'N/A')
    payment_mode = update_data.get('payment_mode', '')
    payment_text = "To Pay" if payment_mode == 'to_pay' else "Paid" if payment_mode == 'paid' else 'N/A'
    
    # Package details
    boxes = update_data.get('boxes_count', 0) or 0
    cans = update_data.get('cans_count', 0) or 0
    bags = update_data.get('bags_count', 0) or 0
    
    package_parts = []
    if boxes: package_parts.append(f"{boxes} Box(es)")
    if cans: package_parts.append(f"{cans} Can(s)")
    if bags: package_parts.append(f"{bags} Bag(s)")
    package_text = ", ".join(package_parts) if package_parts else "N/A"
    
    # Invoice details
    invoice_number = update_data.get('invoice_number', 'N/A') or 'N/A'
    invoice_date = update_data.get('invoice_date', 'N/A') or 'N/A'
    invoice_value = update_data.get('invoice_value', 0) or 0
    invoice_value_text = f"₹{invoice_value:,.2f}" if invoice_value else 'N/A'
    
    # Build items list
    items = order.get('items', [])
    items_text = "\n".join([f"• {item.get('item_name', '')} - Qty: {item.get('quantity', '')}" for item in items])
    
    # 1. Send message to TRANSPORTER
    transport_id = update_data.get('transport_id')
    if transport_id:
        transport = await db.transports.find_one({'id': transport_id}, {'_id': 0})
        if transport and transport.get('contact_number'):
            transporter_mobile = transport.get('contact_number', '')
            clean_transporter_mobile = ''.join(filter(str.isdigit, transporter_mobile))
            if not clean_transporter_mobile.startswith('91'):
                clean_transporter_mobile = f"91{clean_transporter_mobile[-10:]}"
            
            transporter_message = f"""📦 *NEW SHIPMENT READY*

*Delivery Details:*
👤 {doctor_name}
📍 {delivery_station}

*Invoice Details:*
🧾 {invoice_number}
📅 {invoice_date}
💵 {invoice_value_text}

*Package Details:*
📦 {package_text}

*Payment:* {payment_text}"""
            
            params = {
                'action': 'send',
                'senderId': config['sender_id'],
                'authToken': config['auth_token'],
                'messageText': transporter_message,
                'receiverId': clean_transporter_mobile
            }
            
            try:
                async with httpx.AsyncClient() as client:
                    response = await client.get(config['api_url'], params=params, timeout=30)
                    logger.info(f"Ready to despatch notification sent to transporter {clean_transporter_mobile}")
                    status = 'success' if response.status_code == 200 else 'failed'
                    await log_whatsapp_message(clean_transporter_mobile, 'ready_to_despatch_transporter', transporter_message, status, recipient_name=transport_name)
            except Exception as e:
                logger.error(f"WhatsApp transporter notification error: {str(e)}")
                await log_whatsapp_message(clean_transporter_mobile, 'ready_to_despatch_transporter', transporter_message, 'failed', recipient_name=transport_name, error_message=str(e))

async def send_whatsapp_out_of_stock(order: dict, out_of_stock_items: list):
    """Send WhatsApp notification for out of stock items"""
    if not out_of_stock_items:
        return
    
    config = await get_whatsapp_config()
    if not config:
        logger.warning("WhatsApp config not found, skipping out of stock notification")
        return
    
    # Get company settings
    company = await db.company_settings.find_one({}, {'_id': 0})
    company_name = company.get('company_name', 'VMP CRM') if company else 'VMP CRM'
    company_phone = config.get('sender_id', '')
    
    mobile = order.get('doctor_phone', '')
    clean_mobile = ''.join(filter(str.isdigit, mobile))
    if not clean_mobile.startswith('91'):
        clean_mobile = f"91{clean_mobile[-10:]}"
    
    doctor_name = order.get('doctor_name')
    greeting = f"Dear Dr. {doctor_name}" if doctor_name else "Dear Customer"
    order_number = order.get('order_number', '')
    
    # Build out of stock items list
    items_text = "\n".join([f"• {item.get('item_name', '')} - Qty: {item.get('quantity', '')}" for item in out_of_stock_items])
    
    message = f"""{greeting},

⚠️ *STOCK UPDATE* for Order *{order_number}*

We regret to inform you that the following item(s) are currently *OUT OF STOCK*:

{items_text}

We have noted your requirement and will update you as soon as these items are available.

We sincerely apologize for the inconvenience and thank you for your patience and cooperation. 🙏

Your remaining order items are being processed.

For any queries, please contact us.

Regards,
*{company_name}*
📞 +{company_phone}"""
    
    params = {
        'action': 'send',
        'senderId': config['sender_id'],
        'authToken': config['auth_token'],
        'messageText': message,
        'receiverId': clean_mobile
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(config['api_url'], params=params, timeout=30)
            logger.info(f"Out of stock notification sent to {clean_mobile} for order {order_number}")
            status = 'success' if response.status_code == 200 else 'failed'
            await log_whatsapp_message(clean_mobile, 'out_of_stock', message, status, recipient_name=doctor_name)
    except Exception as e:
        logger.error(f"WhatsApp out of stock notification error: {str(e)}")
        await log_whatsapp_message(clean_mobile, 'out_of_stock', message, 'failed', recipient_name=doctor_name, error_message=str(e))

async def send_whatsapp_stock_arrived(doctor_phone: str, doctor_name: str, item_name: str, item_code: str, quantity: str):
    """Send WhatsApp notification when stock arrives for a pending item"""
    config = await get_whatsapp_config()
    if not config:
        logger.warning("WhatsApp config not found, skipping stock arrived notification")
        return False
    
    # Get company settings
    company = await db.company_settings.find_one({}, {'_id': 0})
    company_name = company.get('company_name', 'VMP CRM') if company else 'VMP CRM'
    company_phone = config.get('sender_id', '')
    
    clean_mobile = ''.join(filter(str.isdigit, doctor_phone))
    if not clean_mobile.startswith('91'):
        clean_mobile = f"91{clean_mobile[-10:]}"
    
    greeting = f"Dear Dr. {doctor_name}" if doctor_name else "Dear Customer"
    
    message = f"""{greeting},

🎉 *GOOD NEWS - STOCK ARRIVED!*

We are pleased to inform you that the following item you requested is now *BACK IN STOCK*:

📦 *{item_name}* ({item_code})
   Quantity requested: {quantity}

You can now place your order or contact us to complete your previous pending order.

Thank you for your patience! 🙏

Regards,
*{company_name}*
📞 +{company_phone}"""
    
    params = {
        'action': 'send',
        'senderId': config['sender_id'],
        'authToken': config['auth_token'],
        'messageText': message,
        'receiverId': clean_mobile
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(config['api_url'], params=params, timeout=30)
            logger.info(f"Stock arrived notification sent to {clean_mobile} for item {item_code}")
            status = 'success' if response.status_code == 200 else 'failed'
            await log_whatsapp_message(clean_mobile, 'stock_arrived', message, status, recipient_name=doctor_name)
            return response.status_code == 200
    except Exception as e:
        logger.error(f"WhatsApp stock arrived notification error: {str(e)}")
        await log_whatsapp_message(clean_mobile, 'stock_arrived', message, 'failed', recipient_name=doctor_name, error_message=str(e))
        return False

@api_router.post("/public/send-otp")
async def send_otp(request: OTPRequest):
    """Send OTP to mobile number via WhatsApp"""
    clean_mobile = ''.join(filter(str.isdigit, request.mobile))
    if len(clean_mobile) < 10:
        raise HTTPException(status_code=400, detail="Invalid mobile number")
    
    # Generate 6-digit OTP
    otp = str(random.randint(100000, 999999))
    
    # Store OTP with expiry (5 minutes)
    otp_doc = {
        'mobile': clean_mobile[-10:],
        'otp': otp,
        'created_at': datetime.now(timezone.utc).isoformat(),
        'expires_at': (datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        'verified': False
    }
    
    # Delete old OTPs for this mobile
    await db.otps.delete_many({'mobile': clean_mobile[-10:]})
    await db.otps.insert_one(otp_doc)
    
    # Send OTP via WhatsApp
    sent = await send_whatsapp_otp(request.mobile, otp)
    
    # For development/testing, also log the OTP
    logger.info(f"OTP for {clean_mobile}: {otp}")
    
    return {"message": "OTP sent successfully", "sent": sent}

@api_router.post("/public/verify-otp", response_model=OrderResponse)
async def verify_otp_and_submit_order(request: OTPVerify):
    """Verify OTP and submit order"""
    clean_mobile = ''.join(filter(str.isdigit, request.mobile))[-10:]
    
    # Find OTP
    otp_doc = await db.otps.find_one({
        'mobile': clean_mobile,
        'otp': request.otp,
        'verified': False
    }, {'_id': 0})
    
    if not otp_doc:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    # Check expiry
    expires_at = datetime.fromisoformat(otp_doc['expires_at'].replace('Z', '+00:00'))
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="OTP expired")
    
    # Mark OTP as verified
    await db.otps.update_one({'mobile': clean_mobile}, {'$set': {'verified': True}})
    
    # Get doctor details if exists
    doctor = await db.doctors.find_one(
        {'$or': [
            {'phone': {'$regex': clean_mobile, '$options': 'i'}},
            {'phone': f"+91{clean_mobile}"},
            {'phone': f"91{clean_mobile}"}
        ]},
        {'_id': 0}
    )
    
    # Generate order number
    order_count = await db.orders.count_documents({})
    order_number = f"ORD-{str(order_count + 1).zfill(6)}"
    
    order_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Filter items with quantity
    valid_items = [item for item in request.items if item.quantity and item.quantity.strip()]
    
    order_doc = {
        'id': order_id,
        'order_number': order_number,
        'doctor_id': doctor['id'] if doctor else None,
        'doctor_name': doctor['name'] if doctor else None,
        'doctor_phone': request.mobile,
        'doctor_email': doctor['email'] if doctor else None,
        'doctor_address': doctor['address'] if doctor else None,
        'doctor_customer_code': doctor['customer_code'] if doctor else None,
        'items': [item.model_dump() for item in valid_items],
        'status': 'pending',
        'ip_address': request.ip_address,
        'location': request.location,
        'device_info': request.device_info,
        'created_at': now.isoformat()
    }
    
    await db.orders.insert_one(order_doc)
    
    # Send order confirmation via WhatsApp with full details
    await send_whatsapp_order(
        mobile=request.mobile, 
        items=valid_items, 
        order_number=order_number,
        doctor_name=doctor['name'] if doctor else None,
        ip_address=request.ip_address,
        location=request.location
    )
    
    # Send Email confirmation to customer (if email available)
    await send_order_confirmation_email(
        order_doc,
        [item.model_dump() for item in valid_items]
    )
    
    return OrderResponse(
        id=order_id,
        order_number=order_number,
        doctor_id=doctor['id'] if doctor else None,
        doctor_name=doctor['name'] if doctor else None,
        doctor_phone=request.mobile,
        doctor_email=doctor['email'] if doctor else None,
        doctor_address=doctor['address'] if doctor else None,
        items=valid_items,
        status='pending',
        ip_address=request.ip_address,
        location=request.location,
        device_info=request.device_info,
        created_at=now
    )

# ============== ORDERS ADMIN ROUTES ==============

@api_router.get("/customers/search")
async def search_customers(q: str, current_user: dict = Depends(get_current_user)):
    """Search across doctors, medicals, and agencies by name or phone"""
    if not q or len(q) < 2:
        return []
    
    search_regex = {'$regex': q, '$options': 'i'}
    results = []
    
    # Search doctors
    doctors = await db.doctors.find({
        '$or': [
            {'name': search_regex},
            {'phone': search_regex},
            {'customer_code': search_regex}
        ]
    }, {'_id': 0}).limit(10).to_list(10)
    
    for doc in doctors:
        results.append({
            'id': doc['id'],
            'name': doc.get('name', ''),
            'phone': doc.get('phone', ''),
            'email': doc.get('email', ''),
            'address': doc.get('address', ''),
            'customer_code': doc.get('customer_code', ''),
            'type': 'doctor',
            'type_label': 'Doctor'
        })
    
    # Search medicals
    medicals = await db.medicals.find({
        '$or': [
            {'name': search_regex},
            {'phone': search_regex},
            {'customer_code': search_regex}
        ]
    }, {'_id': 0}).limit(10).to_list(10)
    
    for med in medicals:
        results.append({
            'id': med['id'],
            'name': med.get('name', ''),
            'phone': med.get('phone', ''),
            'email': med.get('email', ''),
            'address': med.get('address', ''),
            'customer_code': med.get('customer_code', ''),
            'type': 'medical',
            'type_label': 'Medical'
        })
    
    # Search agencies
    agencies = await db.agencies.find({
        '$or': [
            {'name': search_regex},
            {'phone': search_regex},
            {'customer_code': search_regex}
        ]
    }, {'_id': 0}).limit(10).to_list(10)
    
    for agency in agencies:
        results.append({
            'id': agency['id'],
            'name': agency.get('name', ''),
            'phone': agency.get('phone', ''),
            'email': agency.get('email', ''),
            'address': agency.get('address', ''),
            'customer_code': agency.get('customer_code', ''),
            'type': 'agency',
            'type_label': 'Agency'
        })
    
    return results

@api_router.post("/orders")
async def create_manual_order(order_data: ManualOrderCreate, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Create a new order manually by admin/staff"""
    # Generate order number
    today_str = datetime.now(timezone.utc).strftime('%Y%m%d')
    today_orders = await db.orders.count_documents({
        'order_number': {'$regex': f'^ORD-{today_str}'}
    })
    order_number = f"ORD-{today_str}-{(today_orders + 1):04d}"
    
    order_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Clean phone number
    clean_phone = ''.join(filter(str.isdigit, order_data.customer_phone))
    if len(clean_phone) > 10:
        clean_phone = clean_phone[-10:]
    
    entity_id = order_data.customer_id
    entity_type = order_data.customer_type
    
    # If no customer_id provided, try to find existing or create new
    if not entity_id:
        # Search for existing entity by phone
        collection_map = {
            'doctor': db.doctors,
            'medical': db.medicals,
            'agency': db.agencies
        }
        collection = collection_map.get(entity_type, db.doctors)
        
        existing = await collection.find_one({'phone': {'$regex': clean_phone}}, {'_id': 0})
        
        if existing:
            entity_id = existing['id']
        else:
            # Create new entity
            entity_id = str(uuid.uuid4())
            
            # Generate customer code based on type
            prefix_map = {'doctor': 'VMP', 'medical': 'MED', 'agency': 'AGN'}
            prefix = prefix_map.get(entity_type, 'VMP')
            
            last_entity = await collection.find_one({}, sort=[('customer_code', -1)])
            if last_entity and last_entity.get('customer_code', '').startswith(f'{prefix}-'):
                try:
                    last_num = int(last_entity['customer_code'].split('-')[1])
                    customer_code = f"{prefix}-{last_num + 1:04d}"
                except:
                    customer_code = f"{prefix}-0001"
            else:
                customer_code = f"{prefix}-0001"
            
            entity_doc = {
                'id': entity_id,
                'name': order_data.customer_name,
                'customer_code': customer_code,
                'phone': clean_phone,
                'email': order_data.customer_email or '',
                'address': order_data.customer_address or '',
                'lead_status': 'Customer',
                'priority': 'moderate',
                'created_at': now.isoformat()
            }
            
            # Add reg_no for doctors
            if entity_type == 'doctor':
                entity_doc['reg_no'] = ''
            
            await collection.insert_one(entity_doc)
    
    # Create order
    order_doc = {
        'id': order_id,
        'order_number': order_number,
        'doctor_id': entity_id,  # Keep as doctor_id for backward compatibility
        'customer_type': entity_type,
        'doctor_name': order_data.customer_name,
        'doctor_phone': clean_phone,
        'doctor_email': order_data.customer_email,
        'doctor_address': order_data.customer_address,
        'items': [item.dict() for item in order_data.items],
        'status': 'pending',
        'created_by': current_user['name'],
        'created_by_id': current_user['id'],
        'source': 'admin_panel',
        'created_at': now.isoformat()
    }
    
    await db.orders.insert_one(order_doc)
    
    # Handle pending items (out of stock items)
    if order_data.pending_items and len(order_data.pending_items) > 0:
        for pending_item in order_data.pending_items:
            pending_doc = {
                'id': str(uuid.uuid4()),
                'order_id': order_id,
                'original_order_id': order_id,  # Add original_order_id for compatibility
                'original_order_number': order_number,
                'original_order_date': now.isoformat(),  # Add order date
                'doctor_phone': clean_phone,
                'doctor_name': order_data.customer_name,
                'item_id': pending_item.get('item_id'),
                'item_code': pending_item.get('item_code'),
                'item_name': pending_item.get('item_name'),
                'quantity': pending_item.get('quantity', '1'),
                'status': 'pending',
                'created_at': now.isoformat()
            }
            await db.pending_items.insert_one(pending_doc)
        
        # Send WhatsApp notification for out of stock items
        background_tasks.add_task(send_whatsapp_out_of_stock, order_doc, order_data.pending_items)
    
    # Send WhatsApp confirmation
    background_tasks.add_task(
        send_whatsapp_order,
        clean_phone,
        order_data.items,
        order_number,
        order_data.customer_name
    )
    
    # Send Email confirmation to customer
    background_tasks.add_task(
        send_order_confirmation_email,
        order_doc,
        [item.dict() for item in order_data.items]
    )
    
    return {
        "message": "Order created successfully",
        "order_number": order_number,
        "order_id": order_id,
        "customer_type": entity_type,
        "customer_id": entity_id,
        "pending_items_count": len(order_data.pending_items) if order_data.pending_items else 0
    }

@api_router.get("/orders", response_model=List[OrderResponse])
async def get_orders(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query['status'] = status
    
    orders = await db.orders.find(query, {'_id': 0}).sort('created_at', -1).to_list(1000)
    
    result = []
    for order in orders:
        created_at = order['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        items = [OrderItem(**item) for item in order.get('items', [])]
        
        result.append(OrderResponse(
            id=order['id'],
            order_number=order['order_number'],
            doctor_id=order.get('doctor_id'),
            doctor_name=order.get('doctor_name'),
            doctor_phone=order['doctor_phone'],
            doctor_email=order.get('doctor_email'),
            doctor_address=order.get('doctor_address'),
            items=items,
            status=order['status'],
            transport_id=order.get('transport_id'),
            transport_name=order.get('transport_name'),
            tracking_number=order.get('tracking_number'),
            tracking_url=order.get('tracking_url'),
            delivery_station=order.get('delivery_station'),
            payment_mode=order.get('payment_mode'),
            payment_amount=order.get('payment_amount'),
            expense_paid_by=order.get('expense_paid_by'),
            expense_account=order.get('expense_account'),
            boxes_count=order.get('boxes_count'),
            cans_count=order.get('cans_count'),
            bags_count=order.get('bags_count'),
            invoice_number=order.get('invoice_number'),
            invoice_date=order.get('invoice_date'),
            invoice_value=order.get('invoice_value'),
            cancellation_reason=order.get('cancellation_reason'),
            ip_address=order.get('ip_address'),
            location=order.get('location'),
            device_info=order.get('device_info'),
            created_at=created_at
        ))
    
    return result

async def auto_create_transport_expense(order: dict, update_data: dict, current_user: dict):
    """Auto-create an expense entry when order is shipped with 'paid' payment mode"""
    try:
        # Ensure default categories exist
        await ensure_default_categories()
        
        # Find Transport/Shipping category
        transport_category = await db.expense_categories.find_one({'name': 'Transport/Shipping'}, {'_id': 0})
        if not transport_category:
            return  # Skip if category not found
        
        now = datetime.now(timezone.utc)
        transport_name = order.get('transport_name') or update_data.get('transport_name', 'Unknown Transport')
        delivery_station = order.get('delivery_station') or update_data.get('delivery_station', '')
        order_number = order.get('order_number', 'N/A')
        
        # Use payment_amount for expense, fallback to invoice_value for backward compatibility
        payment_amount = update_data.get('payment_amount') or order.get('payment_amount') or order.get('invoice_value', 0)
        expense_paid_by = update_data.get('expense_paid_by') or order.get('expense_paid_by')
        expense_account = update_data.get('expense_account') or order.get('expense_account', 'company_account')
        
        if not payment_amount or payment_amount <= 0:
            logger.info(f"Skipping expense creation for order {order_number} - no payment amount")
            return
        
        # Check if expense already exists for this order
        existing = await db.expenses.find_one({'order_id': order['id'], 'is_auto_generated': True}, {'_id': 0})
        if existing:
            return  # Don't create duplicate
        
        expense_doc = {
            'id': str(uuid.uuid4()),
            'category_id': transport_category['id'],
            'category_name': transport_category['name'],
            'date': now.strftime('%Y-%m-%d'),
            'amount': payment_amount,
            'payment_type': 'cash' if expense_account == 'cash' else 'net_banking',
            'payment_account': expense_account or 'company_account',
            'paid_by': expense_paid_by,
            'reason': f'Transport for Order #{order_number} via {transport_name}',
            'transport_id': order.get('transport_id'),
            'transport_name': transport_name,
            'transport_location': delivery_station,
            'order_id': order['id'],
            'order_number': order_number,
            'is_auto_generated': True,
            'created_at': now.isoformat(),
            'updated_at': now.isoformat(),
            'created_by': current_user['id']
        }
        
        await db.expenses.insert_one(expense_doc)
        logger.info(f"Auto-created transport expense ₹{payment_amount} for order {order_number}")
    except Exception as e:
        logger.error(f"Failed to auto-create transport expense: {str(e)}")

@api_router.put("/orders/{order_id}/status")
async def update_order_status(order_id: str, status_data: OrderStatusUpdate, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Update order status with optional transport details (for ready_to_despatch/shipped) or cancellation reason"""
    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    new_status = status_data.status
    update_data = {
        'status': new_status,
        'updated_at': datetime.now(timezone.utc).isoformat()
    }
    
    # Add transport/shipping details if status is 'ready_to_despatch'
    if new_status == 'ready_to_despatch':
        if status_data.transport_id:
            update_data['transport_id'] = status_data.transport_id
        if status_data.transport_name:
            update_data['transport_name'] = status_data.transport_name
        if status_data.delivery_station:
            update_data['delivery_station'] = status_data.delivery_station
        if status_data.payment_mode:
            update_data['payment_mode'] = status_data.payment_mode
        # Payment amount (for both to_pay and paid)
        if status_data.payment_amount is not None:
            update_data['payment_amount'] = status_data.payment_amount
        # Expense details (only relevant for 'paid' mode)
        if status_data.expense_paid_by:
            update_data['expense_paid_by'] = status_data.expense_paid_by
        if status_data.expense_account:
            update_data['expense_account'] = status_data.expense_account
        # Package counts
        if status_data.boxes_count is not None:
            update_data['boxes_count'] = status_data.boxes_count
        if status_data.cans_count is not None:
            update_data['cans_count'] = status_data.cans_count
        if status_data.bags_count is not None:
            update_data['bags_count'] = status_data.bags_count
        # Invoice details
        if status_data.invoice_number:
            update_data['invoice_number'] = status_data.invoice_number
        if status_data.invoice_date:
            update_data['invoice_date'] = status_data.invoice_date
        if status_data.invoice_value is not None:
            update_data['invoice_value'] = status_data.invoice_value
    
    # Only add tracking number if status is 'shipped'
    if new_status == 'shipped':
        if status_data.tracking_number:
            update_data['tracking_number'] = status_data.tracking_number
        if status_data.tracking_url:
            update_data['tracking_url'] = status_data.tracking_url
        # Also update transport if provided (for cases where ready_to_despatch was skipped)
        if status_data.transport_id:
            update_data['transport_id'] = status_data.transport_id
        if status_data.transport_name:
            update_data['transport_name'] = status_data.transport_name
        
        # Auto-create expense for shipped orders with 'paid' payment mode
        payment_mode = order.get('payment_mode') or update_data.get('payment_mode', '')
        payment_amount = order.get('payment_amount') or update_data.get('payment_amount', 0)
        if payment_mode == 'paid' and payment_amount > 0:
            await auto_create_transport_expense(order, update_data, current_user)
    
    # Only add cancellation reason if status is 'cancelled'
    if new_status == 'cancelled' and status_data.cancellation_reason:
        update_data['cancellation_reason'] = status_data.cancellation_reason
    
    await db.orders.update_one({'id': order_id}, {'$set': update_data})
    
    # Send WhatsApp notification for status change
    if new_status in ['confirmed', 'ready_to_despatch', 'shipped', 'delivered', 'cancelled']:
        # For ready_to_despatch, we need to send to both transporter and customer
        if new_status == 'ready_to_despatch':
            background_tasks.add_task(send_whatsapp_ready_to_despatch, order, update_data)
        else:
            background_tasks.add_task(send_whatsapp_status_update, order, new_status, update_data)
    
    return {"message": f"Order status updated to {new_status}"}

# Keep legacy endpoint for backward compatibility
@api_router.put("/orders/{order_id}/transport")
async def update_order_transport(order_id: str, transport_data: OrderStatusUpdate, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Legacy endpoint - redirects to status update"""
    return await update_order_status(order_id, transport_data, background_tasks, current_user)

@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an order (admin only or with delete_orders permission)"""
    if current_user.get('role') != 'admin':
        perms = current_user.get('permissions', {})
        if not perms.get('delete_orders', False):
            raise HTTPException(status_code=403, detail="You don't have permission to delete orders")
    
    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    await db.orders.delete_one({'id': order_id})
    return {"message": "Order deleted successfully"}

@api_router.put("/orders/{order_id}/items")
async def update_order_items(order_id: str, update_data: OrderItemsUpdate, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):

    """Update order items and optionally create pending items for removed items"""
    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Update order items
    items_data = [item.dict() for item in update_data.items]
    await db.orders.update_one(
        {'id': order_id},
        {'$set': {
            'items': items_data,
            'updated_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Create pending items if any and send WhatsApp notification
    if update_data.pending_items and len(update_data.pending_items) > 0:
        now = datetime.now(timezone.utc)
        order_date = order.get('created_at')
        if isinstance(order_date, str):
            order_date = datetime.fromisoformat(order_date.replace('Z', '+00:00'))
        
        for pending_item in update_data.pending_items:
            pending_doc = {
                'id': str(uuid.uuid4()),
                'doctor_phone': order.get('doctor_phone'),
                'doctor_name': order.get('doctor_name'),
                'item_id': pending_item.get('item_id'),
                'item_code': pending_item.get('item_code'),
                'item_name': pending_item.get('item_name'),
                'quantity': pending_item.get('quantity'),
                'original_order_id': order_id,
                'original_order_number': order.get('order_number'),
                'original_order_date': order_date.isoformat() if isinstance(order_date, datetime) else order_date,
                'created_at': now.isoformat()
            }
            await db.pending_items.insert_one(pending_doc)
        
        # Send WhatsApp notification for out of stock items
        background_tasks.add_task(send_whatsapp_out_of_stock, order, update_data.pending_items)
    
    return {"message": "Order items updated successfully"}

@api_router.put("/orders/{order_id}/customer")
async def update_order_customer(order_id: str, customer_data: OrderCustomerUpdate, current_user: dict = Depends(get_current_user)):
    """Update customer/doctor information for an order"""
    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    update_data = {
        'updated_at': datetime.now(timezone.utc).isoformat()
    }
    
    if customer_data.doctor_name:
        update_data['doctor_name'] = customer_data.doctor_name
    if customer_data.doctor_email:
        update_data['doctor_email'] = customer_data.doctor_email
    if customer_data.doctor_address:
        update_data['doctor_address'] = customer_data.doctor_address
    if customer_data.doctor_phone:
        update_data['doctor_phone'] = customer_data.doctor_phone
    
    # Update the order
    await db.orders.update_one({'id': order_id}, {'$set': update_data})
    
    # Also update pending items for this customer if phone changed
    old_phone = order.get('doctor_phone')
    new_phone = customer_data.doctor_phone or old_phone
    new_name = customer_data.doctor_name or order.get('doctor_name')
    
    if old_phone:
        await db.pending_items.update_many(
            {'doctor_phone': old_phone},
            {'$set': {'doctor_phone': new_phone, 'doctor_name': new_name}}
        )
    
    # If link_to_doctor is true, find or create doctor record
    if customer_data.link_to_doctor:
        phone = new_phone or old_phone
        existing_doctor = await db.doctors.find_one({'phone': phone}, {'_id': 0})
        
        if existing_doctor:
            # Update existing doctor
            doctor_update = {}
            if new_name:
                doctor_update['name'] = new_name
            if customer_data.doctor_email:
                doctor_update['email'] = customer_data.doctor_email
            if customer_data.doctor_address:
                doctor_update['address'] = customer_data.doctor_address
            
            if doctor_update:
                await db.doctors.update_one({'phone': phone}, {'$set': doctor_update})
            
            # Link order to doctor
            await db.orders.update_one({'id': order_id}, {'$set': {'doctor_id': existing_doctor['id']}})
            
            return {"message": "Customer info updated and linked to existing doctor", "doctor_id": existing_doctor['id']}
        else:
            # Create new doctor
            count = await db.doctors.count_documents({})
            customer_code = f"VMP-{str(count + 1).zfill(4)}"
            
            new_doctor = {
                'id': str(uuid.uuid4()),
                'name': new_name or 'Unknown',
                'reg_no': '',
                'address': customer_data.doctor_address or '',
                'email': customer_data.doctor_email or '',
                'phone': phone,
                'dob': None,
                'lead_status': 'Customer',
                'customer_code': customer_code,
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
            await db.doctors.insert_one(new_doctor)
            await db.orders.update_one({'id': order_id}, {'$set': {'doctor_id': new_doctor['id']}})
            
            return {"message": "Customer info updated and new doctor created", "doctor_id": new_doctor['id'], "customer_code": customer_code}
    
    return {"message": "Customer info updated successfully"}

@api_router.get("/orders/{order_id}/lookup-doctor")
async def lookup_doctor_for_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """Look up existing doctor by order's phone number"""
    order = await db.orders.find_one({'id': order_id}, {'_id': 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    phone = order.get('doctor_phone')
    if not phone:
        return {"found": False, "doctor": None}
    
    # Clean phone number - get last 10 digits
    clean_phone = ''.join(filter(str.isdigit, phone))[-10:]
    
    # Search for doctor with matching phone
    doctor = await db.doctors.find_one(
        {'phone': {'$regex': clean_phone}},
        {'_id': 0}
    )
    
    if doctor:
        return {"found": True, "doctor": doctor}
    
    return {"found": False, "doctor": None}

# ============== PENDING ITEMS ROUTES ==============

@api_router.get("/pending-items")
async def get_all_pending_items(current_user: dict = Depends(get_current_user)):
    """Get all pending items"""
    pending_items = await db.pending_items.find({}, {'_id': 0}).sort('created_at', -1).to_list(1000)
    
    result = []
    for item in pending_items:
        created_at = item.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        # Use original_order_date if available, otherwise use created_at as fallback
        order_date = item.get('original_order_date') or item.get('created_at')
        if isinstance(order_date, str):
            order_date = datetime.fromisoformat(order_date.replace('Z', '+00:00'))
        
        result.append(PendingItemResponse(
            id=item['id'],
            doctor_phone=item['doctor_phone'],
            doctor_name=item.get('doctor_name'),
            item_id=item['item_id'],
            item_code=item['item_code'],
            item_name=item['item_name'],
            quantity=item['quantity'],
            original_order_id=item.get('original_order_id') or item.get('order_id'),
            original_order_number=item['original_order_number'],
            original_order_date=order_date or created_at,
            created_at=created_at
        ))
    
    return result

@api_router.get("/pending-items/stats")
async def get_pending_items_stats(current_user: dict = Depends(get_current_user)):
    """Get pending items statistics"""
    total_count = await db.pending_items.count_documents({})
    
    # Get unique doctors with pending items
    pipeline = [
        {"$group": {"_id": "$doctor_phone", "count": {"$sum": 1}}},
        {"$count": "unique_doctors"}
    ]
    doctors_result = await db.pending_items.aggregate(pipeline).to_list(1)
    unique_doctors = doctors_result[0]['unique_doctors'] if doctors_result else 0
    
    return {
        "total_pending_items": total_count,
        "doctors_with_pending": unique_doctors
    }

@api_router.get("/pending-items/doctor/{phone}")
async def get_pending_items_by_doctor(phone: str, current_user: dict = Depends(get_current_user)):
    """Get pending items for a specific doctor by phone"""
    pending_items = await db.pending_items.find({'doctor_phone': phone}, {'_id': 0}).sort('created_at', -1).to_list(100)
    
    result = []
    for item in pending_items:
        created_at = item.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        order_date = item.get('original_order_date') or item.get('created_at')
        if isinstance(order_date, str):
            order_date = datetime.fromisoformat(order_date.replace('Z', '+00:00'))
        
        result.append(PendingItemResponse(
            id=item['id'],
            doctor_phone=item['doctor_phone'],
            doctor_name=item.get('doctor_name'),
            item_id=item['item_id'],
            item_code=item['item_code'],
            item_name=item['item_name'],
            quantity=item['quantity'],
            original_order_id=item.get('original_order_id') or item.get('order_id'),
            original_order_number=item['original_order_number'],
            original_order_date=order_date,
            created_at=created_at
        ))
    
    return result

@api_router.delete("/pending-items/{item_id}")
async def delete_pending_item(item_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a pending item after customer contact"""
    result = await db.pending_items.delete_one({'id': item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pending item not found")
    return {"message": "Pending item deleted successfully"}

@api_router.get("/pending-items/by-item")
async def get_pending_items_grouped_by_item(current_user: dict = Depends(get_current_user)):
    """Get pending items grouped by item (for stock arrival notifications)"""
    pipeline = [
        {
            "$group": {
                "_id": {
                    "item_id": "$item_id",
                    "item_code": "$item_code",
                    "item_name": "$item_name"
                },
                "doctors": {
                    "$push": {
                        "pending_id": "$id",
                        "doctor_phone": "$doctor_phone",
                        "doctor_name": "$doctor_name",
                        "quantity": "$quantity",
                        "original_order_number": "$original_order_number",
                        "created_at": "$created_at"
                    }
                },
                "total_quantity": {"$sum": 1},  # Count of pending requests (quantity is text)
                "doctor_count": {"$sum": 1}
            }
        },
        {
            "$project": {
                "_id": 0,
                "item_id": "$_id.item_id",
                "item_code": "$_id.item_code",
                "item_name": "$_id.item_name",
                "doctors": 1,
                "total_quantity": 1,
                "doctor_count": 1
            }
        },
        {"$sort": {"doctor_count": -1}}
    ]
    
    result = await db.pending_items.aggregate(pipeline).to_list(100)
    return result

@api_router.post("/pending-items/notify-stock-arrived/{item_code}")
async def notify_stock_arrived_by_item(item_code: str, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Send stock arrived notification to all doctors waiting for a specific item"""
    # Find all pending items with this item code
    pending_items = await db.pending_items.find({'item_code': item_code}, {'_id': 0}).to_list(100)
    
    if not pending_items:
        raise HTTPException(status_code=404, detail="No pending items found with this item code")
    
    # Get item details from first pending item
    item_name = pending_items[0].get('item_name', item_code)
    
    # Send notifications to all doctors
    notifications_sent = 0
    notifications_failed = 0
    doctors_notified = []
    
    for pending in pending_items:
        doctor_phone = pending.get('doctor_phone')
        doctor_name = pending.get('doctor_name')
        quantity = pending.get('quantity')
        
        # Avoid sending duplicate notifications to same phone
        if doctor_phone in doctors_notified:
            continue
        
        success = await send_whatsapp_stock_arrived(
            doctor_phone=doctor_phone,
            doctor_name=doctor_name,
            item_name=item_name,
            item_code=item_code,
            quantity=quantity
        )
        
        if success:
            notifications_sent += 1
            doctors_notified.append(doctor_phone)
        else:
            notifications_failed += 1
    
    return {
        "message": f"Stock arrived notifications sent for {item_name}",
        "item_code": item_code,
        "item_name": item_name,
        "notifications_sent": notifications_sent,
        "notifications_failed": notifications_failed,
        "doctors_notified": len(doctors_notified)
    }

@api_router.post("/pending-items/{pending_id}/notify-stock-arrived")
async def notify_stock_arrived_single(pending_id: str, current_user: dict = Depends(get_current_user)):
    """Send stock arrived notification for a single pending item"""
    pending_item = await db.pending_items.find_one({'id': pending_id}, {'_id': 0})
    
    if not pending_item:
        raise HTTPException(status_code=404, detail="Pending item not found")
    
    success = await send_whatsapp_stock_arrived(
        doctor_phone=pending_item.get('doctor_phone'),
        doctor_name=pending_item.get('doctor_name'),
        item_name=pending_item.get('item_name'),
        item_code=pending_item.get('item_code'),
        quantity=pending_item.get('quantity')
    )
    
    if success:
        return {"message": "Stock arrived notification sent successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to send notification")

# ============== TRANSPORT ROUTES ==============

@api_router.post("/transports", response_model=TransportResponse)
async def create_transport(transport: TransportCreate, current_user: dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can add transports")
    
    transport_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    transport_doc = {
        'id': transport_id,
        'name': transport.name,
        'tracking_url_template': transport.tracking_url_template if not transport.is_local else None,
        'is_local': transport.is_local,
        'contact_number': transport.contact_number,
        'alternate_number': transport.alternate_number,
        'created_at': now.isoformat()
    }
    
    await db.transports.insert_one(transport_doc)
    
    return TransportResponse(
        id=transport_id,
        name=transport.name,
        tracking_url_template=transport.tracking_url_template if not transport.is_local else None,
        is_local=transport.is_local,
        contact_number=transport.contact_number,
        alternate_number=transport.alternate_number,
        created_at=now
    )

@api_router.get("/transports", response_model=List[TransportResponse])
async def get_transports(current_user: dict = Depends(get_current_user)):
    transports = await db.transports.find({}, {'_id': 0}).sort('name', 1).to_list(100)
    
    result = []
    for t in transports:
        created_at = t['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append(TransportResponse(
            id=t['id'],
            name=t['name'],
            tracking_url_template=t.get('tracking_url_template'),
            is_local=t.get('is_local', False),
            contact_number=t.get('contact_number'),
            alternate_number=t.get('alternate_number'),
            created_at=created_at
        ))
    
    return result

@api_router.get("/public/transports", response_model=List[TransportResponse])
async def get_public_transports():
    """Public transports list for customer portal selection"""
    transports = await db.transports.find({}, {'_id': 0}).sort('name', 1).to_list(100)
    
    result = []
    for t in transports:
        created_at = t['created_at']
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append(TransportResponse(
            id=t['id'],
            name=t['name'],
            tracking_url_template=t.get('tracking_url_template'),
            is_local=t.get('is_local', False),
            contact_number=t.get('contact_number'),
            alternate_number=t.get('alternate_number'),
            created_at=created_at
        ))
    
    return result

@api_router.delete("/transports/{transport_id}")
async def delete_transport(transport_id: str, current_user: dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can delete transports")
    
    result = await db.transports.delete_one({'id': transport_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transport not found")
    return {"message": "Transport deleted successfully"}

# ============== EXPENSE ROUTES ==============

# Default expense categories
DEFAULT_EXPENSE_CATEGORIES = [
    {"name": "Transport/Shipping", "description": "Shipping and transport costs"},
    {"name": "Office Supplies", "description": "Office materials and supplies"},
    {"name": "Salaries", "description": "Employee salaries and wages"},
    {"name": "Utilities", "description": "Electricity, water, internet bills"},
    {"name": "Marketing", "description": "Advertising and marketing expenses"},
    {"name": "Miscellaneous", "description": "Other miscellaneous expenses"},
]

async def ensure_default_categories():
    """Create default expense categories if they don't exist"""
    existing = await db.expense_categories.count_documents({})
    if existing == 0:
        now = datetime.now(timezone.utc)
        for cat in DEFAULT_EXPENSE_CATEGORIES:
            await db.expense_categories.insert_one({
                'id': str(uuid.uuid4()),
                'name': cat['name'],
                'description': cat['description'],
                'is_default': True,
                'created_at': now.isoformat()
            })

@api_router.get("/expense-categories", response_model=List[ExpenseCategoryResponse])
async def get_expense_categories(current_user: dict = Depends(get_current_user)):
    """Get all expense categories"""
    await ensure_default_categories()
    categories = await db.expense_categories.find({}, {'_id': 0}).sort('name', 1).to_list(100)
    
    result = []
    for cat in categories:
        created_at = cat.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append(ExpenseCategoryResponse(
            id=cat['id'],
            name=cat['name'],
            description=cat.get('description'),
            is_default=cat.get('is_default', False),
            created_at=created_at
        ))
    
    return result

@api_router.post("/expense-categories", response_model=ExpenseCategoryResponse)
async def create_expense_category(category: ExpenseCategoryCreate, current_user: dict = Depends(get_current_user)):
    """Create a new expense category"""
    cat_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    cat_doc = {
        'id': cat_id,
        'name': category.name,
        'description': category.description,
        'is_default': False,
        'created_at': now.isoformat()
    }
    
    await db.expense_categories.insert_one(cat_doc)
    
    return ExpenseCategoryResponse(
        id=cat_id,
        name=category.name,
        description=category.description,
        is_default=False,
        created_at=now
    )

@api_router.delete("/expense-categories/{category_id}")
async def delete_expense_category(category_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an expense category (only non-default)"""
    category = await db.expense_categories.find_one({'id': category_id}, {'_id': 0})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    if category.get('is_default', False):
        raise HTTPException(status_code=400, detail="Cannot delete default categories")
    
    expense_count = await db.expenses.count_documents({'category_id': category_id})
    if expense_count > 0:
        raise HTTPException(status_code=400, detail=f"Category is used by {expense_count} expenses")
    
    await db.expense_categories.delete_one({'id': category_id})
    return {"message": "Category deleted successfully"}

@api_router.post("/expenses", response_model=ExpenseResponse)
async def create_expense(expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    """Create a new expense"""
    category = await db.expense_categories.find_one({'id': expense.category_id}, {'_id': 0})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    expense_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    expense_doc = {
        'id': expense_id,
        'category_id': expense.category_id,
        'category_name': category['name'],
        'date': expense.date,
        'amount': expense.amount,
        'payment_type': expense.payment_type,
        'payment_account': expense.payment_account,
        'paid_by': expense.paid_by,
        'reason': expense.reason,
        'transport_id': expense.transport_id,
        'transport_name': expense.transport_name,
        'transport_location': expense.transport_location,
        'order_id': expense.order_id,
        'order_number': expense.order_number,
        'is_auto_generated': False,
        'created_at': now.isoformat(),
        'updated_at': now.isoformat(),
        'created_by': current_user['id']
    }
    
    await db.expenses.insert_one(expense_doc)
    
    return ExpenseResponse(
        id=expense_id,
        category_id=expense.category_id,
        category_name=category['name'],
        date=expense.date,
        amount=expense.amount,
        payment_type=expense.payment_type,
        payment_account=expense.payment_account,
        paid_by=expense.paid_by,
        reason=expense.reason,
        transport_id=expense.transport_id,
        transport_name=expense.transport_name,
        transport_location=expense.transport_location,
        order_id=expense.order_id,
        order_number=expense.order_number,
        is_auto_generated=False,
        created_at=now,
        updated_at=now
    )

@api_router.get("/expenses", response_model=List[ExpenseResponse])
async def get_expenses(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category_id: Optional[str] = None,
    payment_type: Optional[str] = None,
    payment_account: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all expenses with optional filters"""
    query = {}
    
    if start_date and end_date:
        query['date'] = {'$gte': start_date, '$lte': end_date}
    elif start_date:
        query['date'] = {'$gte': start_date}
    elif end_date:
        query['date'] = {'$lte': end_date}
    
    if category_id:
        query['category_id'] = category_id
    if payment_type:
        query['payment_type'] = payment_type
    if payment_account:
        query['payment_account'] = payment_account
    
    expenses = await db.expenses.find(query, {'_id': 0}).sort('date', -1).to_list(1000)
    
    result = []
    for exp in expenses:
        created_at = exp.get('created_at')
        updated_at = exp.get('updated_at') or exp.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        if isinstance(updated_at, str):
            updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
        
        result.append(ExpenseResponse(
            id=exp['id'],
            category_id=exp['category_id'],
            category_name=exp.get('category_name'),
            date=exp['date'],
            amount=exp['amount'],
            payment_type=exp['payment_type'],
            payment_account=exp['payment_account'],
            paid_by=exp.get('paid_by'),
            reason=exp['reason'],
            transport_id=exp.get('transport_id'),
            transport_name=exp.get('transport_name'),
            transport_location=exp.get('transport_location'),
            order_id=exp.get('order_id'),
            order_number=exp.get('order_number'),
            is_auto_generated=exp.get('is_auto_generated', False),
            created_at=created_at,
            updated_at=updated_at
        ))
    
    return result

@api_router.put("/expenses/{expense_id}", response_model=ExpenseResponse)
async def update_expense(expense_id: str, expense_data: ExpenseUpdate, current_user: dict = Depends(get_current_user)):
    """Update an expense"""
    expense = await db.expenses.find_one({'id': expense_id}, {'_id': 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    update_data = {k: v for k, v in expense_data.model_dump().items() if v is not None}
    
    if 'category_id' in update_data:
        category = await db.expense_categories.find_one({'id': update_data['category_id']}, {'_id': 0})
        if category:
            update_data['category_name'] = category['name']
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.expenses.update_one({'id': expense_id}, {'$set': update_data})
    
    updated = await db.expenses.find_one({'id': expense_id}, {'_id': 0})
    created_at = updated.get('created_at')
    updated_at = updated.get('updated_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    return ExpenseResponse(
        id=updated['id'],
        category_id=updated['category_id'],
        category_name=updated.get('category_name'),
        date=updated['date'],
        amount=updated['amount'],
        payment_type=updated['payment_type'],
        payment_account=updated['payment_account'],
        paid_by=updated.get('paid_by'),
        reason=updated['reason'],
        transport_id=updated.get('transport_id'),
        transport_name=updated.get('transport_name'),
        transport_location=updated.get('transport_location'),
        order_id=updated.get('order_id'),
        order_number=updated.get('order_number'),
        is_auto_generated=updated.get('is_auto_generated', False),
        created_at=created_at,
        updated_at=updated_at
    )

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an expense"""
    result = await db.expenses.delete_one({'id': expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}

@api_router.get("/expenses/stats/monthly")
async def get_monthly_expense_stats(current_user: dict = Depends(get_current_user)):
    """Get monthly expense statistics"""
    import calendar
    now = datetime.now(timezone.utc)
    current_month_start = now.replace(day=1).strftime('%Y-%m-%d')
    current_month_end = now.strftime('%Y-%m-%d')
    
    if now.month == 1:
        prev_month_start = now.replace(year=now.year-1, month=12, day=1).strftime('%Y-%m-%d')
        prev_month_end = now.replace(year=now.year-1, month=12, day=31).strftime('%Y-%m-%d')
    else:
        prev_month_start = now.replace(month=now.month-1, day=1).strftime('%Y-%m-%d')
        last_day = calendar.monthrange(now.year, now.month-1)[1]
        prev_month_end = now.replace(month=now.month-1, day=last_day).strftime('%Y-%m-%d')
    
    current_expenses = await db.expenses.find({
        'date': {'$gte': current_month_start, '$lte': current_month_end}
    }, {'_id': 0}).to_list(1000)
    current_total = sum(exp['amount'] for exp in current_expenses)
    
    prev_expenses = await db.expenses.find({
        'date': {'$gte': prev_month_start, '$lte': prev_month_end}
    }, {'_id': 0, 'amount': 1}).to_list(1000)
    prev_total = sum(exp['amount'] for exp in prev_expenses)
    
    by_category = {}
    by_payment_type = {}
    for exp in current_expenses:
        cat_name = exp.get('category_name', 'Uncategorized')
        by_category[cat_name] = by_category.get(cat_name, 0) + exp['amount']
        pt = exp.get('payment_type', 'other')
        by_payment_type[pt] = by_payment_type.get(pt, 0) + exp['amount']
    
    return {
        'current_month_total': current_total,
        'previous_month_total': prev_total,
        'change_percent': round(((current_total - prev_total) / prev_total * 100) if prev_total > 0 else 0, 1),
        'by_category': by_category,
        'by_payment_type': by_payment_type,
        'expense_count': len(current_expenses)
    }

# ============== FOLLOWUP ROUTES ==============

@api_router.post("/followups")
async def create_followup(followup: FollowUpCreate, current_user: dict = Depends(get_current_user)):
    """Add a follow-up entry for a lead. Auto-closes previous open follow-up and updates entity."""
    entity_type = followup.entity_type
    entity_id = followup.entity_id

    # Validate entity type
    collection_map = {'doctor': 'doctors', 'medical': 'medicals', 'agency': 'agencies'}
    if entity_type not in collection_map:
        raise HTTPException(status_code=400, detail="Invalid entity type")

    collection = db[collection_map[entity_type]]
    entity = await collection.find_one({'id': entity_id}, {'_id': 0})
    if not entity:
        raise HTTPException(status_code=404, detail=f"{entity_type.title()} not found")

    now = datetime.now(timezone.utc)
    followup_id = str(uuid.uuid4())

    # Close all previous open follow-ups for this entity
    await db.followups.update_many(
        {'entity_type': entity_type, 'entity_id': entity_id, 'status': 'open'},
        {'$set': {'status': 'closed', 'closed_at': now.isoformat()}}
    )

    # Create new follow-up entry
    followup_doc = {
        'id': followup_id,
        'entity_type': entity_type,
        'entity_id': entity_id,
        'entity_name': entity.get('name', ''),
        'notes': followup.notes,
        'new_status': followup.new_status,
        'next_follow_up_date': followup.next_follow_up_date,
        'next_follow_up_time': followup.next_follow_up_time,
        'status': 'open' if followup.next_follow_up_date else 'closed',
        'created_by': current_user.get('name', current_user.get('email', '')),
        'created_at': now.isoformat()
    }
    await db.followups.insert_one(followup_doc)

    # Update the entity: last_contact_date, follow_up_date, lead_status
    update_data = {
        'last_contact_date': now.strftime('%Y-%m-%d'),
        'updated_at': now.isoformat()
    }
    if followup.next_follow_up_date:
        update_data['follow_up_date'] = followup.next_follow_up_date
    else:
        update_data['follow_up_date'] = None
    if followup.new_status:
        update_data['lead_status'] = followup.new_status

    await collection.update_one({'id': entity_id}, {'$set': update_data})

    followup_doc.pop('_id', None)
    return followup_doc


@api_router.get("/followups/{entity_type}/{entity_id}")
async def get_followup_history(entity_type: str, entity_id: str, current_user: dict = Depends(get_current_user)):
    """Get all follow-up history for a specific entity, newest first."""
    followups = await db.followups.find(
        {'entity_type': entity_type, 'entity_id': entity_id},
        {'_id': 0}
    ).sort('created_at', -1).to_list(100)
    return followups


# ============== REMINDER ROUTES ==============

@api_router.post("/reminders", response_model=ReminderResponse)
async def create_reminder(reminder: ReminderCreate, current_user: dict = Depends(get_current_user)):
    """Create a new reminder"""
    reminder_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    reminder_doc = {
        'id': reminder_id,
        'title': reminder.title,
        'description': reminder.description,
        'reminder_type': reminder.reminder_type,
        'reminder_date': reminder.reminder_date,
        'reminder_time': reminder.reminder_time,
        'entity_type': reminder.entity_type,
        'entity_id': reminder.entity_id,
        'entity_name': reminder.entity_name,
        'priority': reminder.priority,
        'is_completed': False,
        'is_auto_generated': False,
        'created_at': now.isoformat(),
        'created_by': current_user['id']
    }
    
    await db.reminders.insert_one(reminder_doc)
    
    return ReminderResponse(
        id=reminder_id,
        title=reminder.title,
        description=reminder.description,
        reminder_type=reminder.reminder_type,
        reminder_date=reminder.reminder_date,
        reminder_time=reminder.reminder_time,
        entity_type=reminder.entity_type,
        entity_id=reminder.entity_id,
        entity_name=reminder.entity_name,
        priority=reminder.priority,
        is_completed=False,
        is_auto_generated=False,
        created_at=now
    )

@api_router.get("/reminders", response_model=List[ReminderResponse])
async def get_reminders(
    date: Optional[str] = None,
    reminder_type: Optional[str] = None,
    is_completed: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all reminders with optional filters"""
    query = {}
    
    if date:
        query['reminder_date'] = date
    if reminder_type:
        query['reminder_type'] = reminder_type
    if is_completed is not None:
        query['is_completed'] = is_completed
    
    reminders = await db.reminders.find(query, {'_id': 0}).sort('reminder_date', 1).to_list(1000)
    
    result = []
    for rem in reminders:
        created_at = rem.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        
        result.append(ReminderResponse(
            id=rem['id'],
            title=rem['title'],
            description=rem.get('description'),
            reminder_type=rem['reminder_type'],
            reminder_date=rem['reminder_date'],
            reminder_time=rem.get('reminder_time'),
            entity_type=rem.get('entity_type'),
            entity_id=rem.get('entity_id'),
            entity_name=rem.get('entity_name'),
            priority=rem.get('priority', 'moderate'),
            is_completed=rem.get('is_completed', False),
            is_auto_generated=rem.get('is_auto_generated', False),
            created_at=created_at
        ))
    
    return result

@api_router.get("/reminders/today")
async def get_today_reminders(current_user: dict = Depends(get_current_user)):
    """Get all reminders for today including auto-generated ones"""
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    today_md = datetime.now(timezone.utc).strftime('%m-%d')  # For birthday/anniversary matching
    
    reminders = []
    
    # 1. Get manual reminders for today
    manual_reminders = await db.reminders.find({
        'reminder_date': today,
        'is_completed': False
    }, {'_id': 0}).to_list(100)
    
    for rem in manual_reminders:
        created_at = rem.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        reminders.append({
            'id': rem['id'],
            'title': rem['title'],
            'description': rem.get('description'),
            'reminder_type': rem['reminder_type'],
            'reminder_date': rem['reminder_date'],
            'reminder_time': rem.get('reminder_time'),
            'entity_type': rem.get('entity_type'),
            'entity_id': rem.get('entity_id'),
            'entity_name': rem.get('entity_name'),
            'priority': rem.get('priority', 'moderate'),
            'is_completed': rem.get('is_completed', False),
            'is_auto_generated': rem.get('is_auto_generated', False),
            'phone': None
        })
    
    # 2. Get doctors with follow-up due today or overdue
    doctors_followup = await db.doctors.find({
        'follow_up_date': {'$lte': today, '$ne': None},
        'lead_status': {'$nin': ['Not Interested', 'Closed', 'Converted', 'Lost']}
    }, {'_id': 0}).to_list(200)
    
    for doc in doctors_followup:
        is_overdue = doc.get('follow_up_date', '') < today
        # Get last follow-up notes
        last_fu = await db.followups.find_one(
            {'entity_type': 'doctor', 'entity_id': doc['id']},
            {'_id': 0, 'notes': 1, 'created_at': 1}
        )
        reminders.append({
            'id': f"auto_followup_doctor_{doc['id']}",
            'title': f"{'OVERDUE: ' if is_overdue else ''}Follow-up: {doc['name']}",
            'description': last_fu.get('notes', f"Follow-up due for {doc['name']}") if last_fu else f"Follow-up due for {doc['name']} ({doc.get('customer_code', '')})",
            'reminder_type': 'follow_up',
            'reminder_date': doc.get('follow_up_date', today),
            'reminder_time': None,
            'entity_type': 'doctor',
            'entity_id': doc['id'],
            'entity_name': doc['name'],
            'priority': 'high' if is_overdue else doc.get('priority', 'moderate'),
            'is_completed': False,
            'is_auto_generated': True,
            'is_overdue': is_overdue,
            'lead_status': doc.get('lead_status'),
            'phone': doc.get('phone')
        })
    
    # 3. Get medicals with follow-up due today or overdue
    medicals_followup = await db.medicals.find({
        'follow_up_date': {'$lte': today, '$ne': None},
        'lead_status': {'$nin': ['Not Interested', 'Closed', 'Converted', 'Lost']}
    }, {'_id': 0}).to_list(200)
    
    for med in medicals_followup:
        is_overdue = med.get('follow_up_date', '') < today
        last_fu = await db.followups.find_one(
            {'entity_type': 'medical', 'entity_id': med['id']},
            {'_id': 0, 'notes': 1}
        )
        reminders.append({
            'id': f"auto_followup_medical_{med['id']}",
            'title': f"{'OVERDUE: ' if is_overdue else ''}Follow-up: {med['name']}",
            'description': last_fu.get('notes', f"Follow-up due for {med['name']}") if last_fu else f"Follow-up due for {med['name']} ({med.get('customer_code', '')})",
            'reminder_type': 'follow_up',
            'reminder_date': med.get('follow_up_date', today),
            'reminder_time': None,
            'entity_type': 'medical',
            'entity_id': med['id'],
            'entity_name': med['name'],
            'priority': 'high' if is_overdue else med.get('priority', 'moderate'),
            'is_completed': False,
            'is_auto_generated': True,
            'is_overdue': is_overdue,
            'lead_status': med.get('lead_status'),
            'phone': med.get('phone')
        })
    
    # 4. Get agencies with follow-up due today or overdue
    agencies_followup = await db.agencies.find({
        'follow_up_date': {'$lte': today, '$ne': None},
        'lead_status': {'$nin': ['Not Interested', 'Closed', 'Converted', 'Lost']}
    }, {'_id': 0}).to_list(200)
    
    for agy in agencies_followup:
        is_overdue = agy.get('follow_up_date', '') < today
        last_fu = await db.followups.find_one(
            {'entity_type': 'agency', 'entity_id': agy['id']},
            {'_id': 0, 'notes': 1}
        )
        reminders.append({
            'id': f"auto_followup_agency_{agy['id']}",
            'title': f"{'OVERDUE: ' if is_overdue else ''}Follow-up: {agy['name']}",
            'description': last_fu.get('notes', f"Follow-up due for {agy['name']}") if last_fu else f"Follow-up due for {agy['name']} ({agy.get('customer_code', '')})",
            'reminder_type': 'follow_up',
            'reminder_date': agy.get('follow_up_date', today),
            'reminder_time': None,
            'entity_type': 'agency',
            'entity_id': agy['id'],
            'entity_name': agy['name'],
            'priority': 'high' if is_overdue else agy.get('priority', 'moderate'),
            'is_completed': False,
            'is_auto_generated': True,
            'is_overdue': is_overdue,
            'lead_status': agy.get('lead_status'),
            'phone': agy.get('phone')
        })
    
    # 5. Get birthdays today (doctors)
    doctors_all = await db.doctors.find({'dob': {'$exists': True, '$ne': None}}, {'_id': 0}).to_list(1000)
    for doc in doctors_all:
        if doc.get('dob') and doc['dob'][5:] == today_md:  # Match MM-DD
            reminders.append({
                'id': f"auto_birthday_doctor_{doc['id']}",
                'title': f"Birthday: {doc['name']}",
                'description': f"Today is {doc['name']}'s birthday!",
                'reminder_type': 'birthday',
                'reminder_date': today,
                'reminder_time': None,
                'entity_type': 'doctor',
                'entity_id': doc['id'],
                'entity_name': doc['name'],
                'priority': 'high',
                'is_completed': False,
                'is_auto_generated': True,
                'phone': doc.get('phone')
            })
    
    # 6. Get birthdays today (medicals)
    medicals_all = await db.medicals.find({'birthday': {'$exists': True, '$ne': None}}, {'_id': 0}).to_list(1000)
    for med in medicals_all:
        if med.get('birthday') and med['birthday'][5:] == today_md:
            reminders.append({
                'id': f"auto_birthday_medical_{med['id']}",
                'title': f"Birthday: {med['name']} (Prop: {med.get('proprietor_name', '')})",
                'description': f"Today is {med.get('proprietor_name', med['name'])}'s birthday!",
                'reminder_type': 'birthday',
                'reminder_date': today,
                'reminder_time': None,
                'entity_type': 'medical',
                'entity_id': med['id'],
                'entity_name': med['name'],
                'priority': 'high',
                'is_completed': False,
                'is_auto_generated': True,
                'phone': med.get('phone')
            })
    
    # 7. Get birthdays today (agencies)
    agencies_all = await db.agencies.find({'birthday': {'$exists': True, '$ne': None}}, {'_id': 0}).to_list(1000)
    for agy in agencies_all:
        if agy.get('birthday') and agy['birthday'][5:] == today_md:
            reminders.append({
                'id': f"auto_birthday_agency_{agy['id']}",
                'title': f"Birthday: {agy['name']} (Prop: {agy.get('proprietor_name', '')})",
                'description': f"Today is {agy.get('proprietor_name', agy['name'])}'s birthday!",
                'reminder_type': 'birthday',
                'reminder_date': today,
                'reminder_time': None,
                'entity_type': 'agency',
                'entity_id': agy['id'],
                'entity_name': agy['name'],
                'priority': 'high',
                'is_completed': False,
                'is_auto_generated': True,
                'phone': agy.get('phone')
            })
    
    # 8. Get anniversaries today (medicals)
    for med in medicals_all:
        if med.get('anniversary') and med['anniversary'][5:] == today_md:
            reminders.append({
                'id': f"auto_anniversary_medical_{med['id']}",
                'title': f"Anniversary: {med['name']}",
                'description': f"Today is {med['name']}'s business anniversary!",
                'reminder_type': 'anniversary',
                'reminder_date': today,
                'reminder_time': None,
                'entity_type': 'medical',
                'entity_id': med['id'],
                'entity_name': med['name'],
                'priority': 'moderate',
                'is_completed': False,
                'is_auto_generated': True,
                'phone': med.get('phone')
            })
    
    # 9. Get anniversaries today (agencies)
    for agy in agencies_all:
        if agy.get('anniversary') and agy['anniversary'][5:] == today_md:
            reminders.append({
                'id': f"auto_anniversary_agency_{agy['id']}",
                'title': f"Anniversary: {agy['name']}",
                'description': f"Today is {agy['name']}'s business anniversary!",
                'reminder_type': 'anniversary',
                'reminder_date': today,
                'reminder_time': None,
                'entity_type': 'agency',
                'entity_id': agy['id'],
                'entity_name': agy['name'],
                'priority': 'moderate',
                'is_completed': False,
                'is_auto_generated': True,
                'phone': agy.get('phone')
            })
    
    # Sort by priority (high first) then by type
    priority_order = {'high': 0, 'moderate': 1, 'low': 2}
    reminders.sort(key=lambda x: (priority_order.get(x['priority'], 1), x['reminder_type']))
    
    return {
        'date': today,
        'total_count': len(reminders),
        'reminders': reminders
    }

@api_router.put("/reminders/{reminder_id}", response_model=ReminderResponse)
async def update_reminder(reminder_id: str, reminder_data: ReminderUpdate, current_user: dict = Depends(get_current_user)):
    """Update a reminder"""
    reminder = await db.reminders.find_one({'id': reminder_id}, {'_id': 0})
    if not reminder:
        raise HTTPException(status_code=404, detail="Reminder not found")
    
    update_data = {k: v for k, v in reminder_data.model_dump().items() if v is not None}
    
    await db.reminders.update_one({'id': reminder_id}, {'$set': update_data})
    
    updated = await db.reminders.find_one({'id': reminder_id}, {'_id': 0})
    created_at = updated.get('created_at')
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
    
    return ReminderResponse(
        id=updated['id'],
        title=updated['title'],
        description=updated.get('description'),
        reminder_type=updated['reminder_type'],
        reminder_date=updated['reminder_date'],
        reminder_time=updated.get('reminder_time'),
        entity_type=updated.get('entity_type'),
        entity_id=updated.get('entity_id'),
        entity_name=updated.get('entity_name'),
        priority=updated.get('priority', 'moderate'),
        is_completed=updated.get('is_completed', False),
        is_auto_generated=updated.get('is_auto_generated', False),
        created_at=created_at
    )

@api_router.delete("/reminders/{reminder_id}")
async def delete_reminder(reminder_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a reminder"""
    result = await db.reminders.delete_one({'id': reminder_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reminder not found")
    return {"message": "Reminder deleted successfully"}

@api_router.post("/reminders/{reminder_id}/complete")
async def mark_reminder_complete(reminder_id: str, current_user: dict = Depends(get_current_user)):
    """Mark a reminder as completed"""
    # Check if it's an auto-generated reminder (starts with 'auto_')
    if reminder_id.startswith('auto_'):
        # For auto-generated, we need to create a completed record
        await db.completed_reminders.insert_one({
            'id': str(uuid.uuid4()),
            'original_reminder_id': reminder_id,
            'completed_at': datetime.now(timezone.utc).isoformat(),
            'completed_by': current_user['id']
        })
        return {"message": "Reminder marked as completed"}
    
    # For manual reminders
    result = await db.reminders.update_one(
        {'id': reminder_id},
        {'$set': {'is_completed': True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reminder not found")
    
    return {"message": "Reminder marked as completed"}

@api_router.post("/reminders/send-whatsapp-summary")
async def send_whatsapp_reminder_summary(current_user: dict = Depends(get_current_user)):
    """Send today's reminders summary to admin via WhatsApp"""
    # Get WhatsApp config
    wa_config = await db.whatsapp_config.find_one({}, {'_id': 0})
    if not wa_config:
        raise HTTPException(status_code=400, detail="WhatsApp not configured")
    
    # Get company settings for admin number
    company = await db.company_settings.find_one({}, {'_id': 0})
    admin_phone = company.get('phone') if company else None
    
    if not admin_phone:
        raise HTTPException(status_code=400, detail="Admin phone number not configured in company settings")
    
    # Get today's reminders
    today_data = await get_today_reminders(current_user)
    reminders = today_data['reminders']
    
    if not reminders:
        return {"message": "No reminders for today", "sent": False}
    
    # Build detailed message
    message_lines = [
        f"🌅 *Good Morning!*",
        f"📅 *Today's Reminders ({today_data['date']})*",
        f"Total: {today_data['total_count']} reminder(s)",
        ""
    ]
    
    # Group by type
    followups = [r for r in reminders if r['reminder_type'] == 'follow_up']
    birthdays = [r for r in reminders if r['reminder_type'] == 'birthday']
    anniversaries = [r for r in reminders if r['reminder_type'] == 'anniversary']
    custom = [r for r in reminders if r['reminder_type'] == 'custom']
    
    if followups:
        overdue = [r for r in followups if r.get('is_overdue')]
        today_fu = [r for r in followups if not r.get('is_overdue')]
        if overdue:
            message_lines.append(f"🚨 *Overdue Follow-ups ({len(overdue)}):*")
            for r in overdue[:10]:
                status_tag = f" [{r.get('lead_status', '')}]" if r.get('lead_status') else ''
                message_lines.append(f"  • {r['entity_name']} - {r.get('phone', 'N/A')}{status_tag}")
            if len(overdue) > 10:
                message_lines.append(f"  +{len(overdue) - 10} more...")
            message_lines.append("")
        if today_fu:
            message_lines.append(f"📞 *Follow-ups Today ({len(today_fu)}):*")
            for r in today_fu[:10]:
                status_tag = f" [{r.get('lead_status', '')}]" if r.get('lead_status') else ''
                message_lines.append(f"  • {r['entity_name']} - {r.get('phone', 'N/A')}{status_tag}")
            if len(today_fu) > 10:
                message_lines.append(f"  +{len(today_fu) - 10} more...")
            message_lines.append("")
    
    if birthdays:
        message_lines.append(f"🎂 *Birthdays ({len(birthdays)}):*")
        for r in birthdays:
            message_lines.append(f"  • {r['entity_name']} - {r.get('phone', 'N/A')} ({r.get('entity_type', '').title()})")
        message_lines.append("")
    
    if anniversaries:
        message_lines.append(f"🎉 *Anniversaries ({len(anniversaries)}):*")
        for r in anniversaries:
            message_lines.append(f"  • {r['entity_name']} - {r.get('phone', 'N/A')} ({r.get('entity_type', '').title()})")
        message_lines.append("")
    
    if custom:
        message_lines.append(f"📝 *Custom ({len(custom)}):*")
        for r in custom[:5]:
            name_part = f" - {r['entity_name']}" if r.get('entity_name') else ''
            message_lines.append(f"  • {r['title']}{name_part}")
        if len(custom) > 5:
            message_lines.append(f"  +{len(custom) - 5} more...")
    
    message_lines.append("")
    message_lines.append("Login to CRM to view details.")
    
    message = "\n".join(message_lines)
    
    # Send WhatsApp
    try:
        import httpx
        async with httpx.AsyncClient() as client:
            response = await client.post(
                wa_config['api_url'],
                json={
                    "token": wa_config['api_token'],
                    "phone": admin_phone,
                    "message": message
                },
                timeout=30.0
            )
            
            if response.status_code == 200:
                logger.info(f"Reminder summary sent to admin: {admin_phone}")
                return {"message": "Reminder summary sent successfully", "sent": True, "phone": admin_phone}
            else:
                logger.error(f"Failed to send reminder summary: {response.text}")
                raise HTTPException(status_code=500, detail="Failed to send WhatsApp message")
    except Exception as e:
        logger.error(f"Error sending reminder summary: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============== WHATSAPP CONFIG ROUTES ==============

@api_router.post("/whatsapp-config", response_model=WhatsAppConfigResponse)
async def save_whatsapp_config(config: WhatsAppConfigCreate, current_user: dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can configure WhatsApp settings")
    
    config_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Delete existing config (only one allowed)
    await db.whatsapp_config.delete_many({})
    
    config_doc = {
        'id': config_id,
        'api_url': config.api_url,
        'auth_token': config.auth_token,
        'sender_id': config.sender_id,
        'updated_at': now.isoformat()
    }
    
    await db.whatsapp_config.insert_one(config_doc)
    
    return WhatsAppConfigResponse(
        id=config_id,
        api_url=config.api_url,
        sender_id=config.sender_id,
        updated_at=now
    )

@api_router.get("/whatsapp-config", response_model=Optional[WhatsAppConfigResponse])
async def get_whatsapp_config_route(current_user: dict = Depends(get_current_user)):
    config = await db.whatsapp_config.find_one({}, {'_id': 0, 'auth_token': 0})
    if not config:
        # Return default config info (without token)
        return WhatsAppConfigResponse(
            id='default',
            api_url='https://api.botmastersender.com/api/v1/',
            sender_id='919944472488',
            updated_at=datetime.now(timezone.utc)
        )
    
    updated_at = config['updated_at']
    if isinstance(updated_at, str):
        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
    
    return WhatsAppConfigResponse(
        id=config['id'],
        api_url=config['api_url'],
        sender_id=config['sender_id'],
        updated_at=updated_at
    )

@api_router.post("/whatsapp-config/test")
async def test_whatsapp_config(mobile: str, current_user: dict = Depends(get_current_user)):
    """Send a test message to verify WhatsApp configuration"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can test WhatsApp")
    
    config = await get_whatsapp_config()
    
    clean_mobile = ''.join(filter(str.isdigit, mobile))
    if not clean_mobile.startswith('91'):
        clean_mobile = f"91{clean_mobile[-10:]}"
    
    message = "Test message from VMP CRM. WhatsApp integration is working!"
    
    params = {
        'action': 'send',
        'senderId': config['sender_id'],
        'authToken': config['auth_token'],
        'messageText': message,
        'receiverId': clean_mobile
    }
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(config['api_url'], params=params, timeout=30)
            if response.status_code == 200:
                return {"message": "Test message sent successfully", "status": "success"}
            else:
                return {"message": f"API returned status {response.status_code}", "status": "failed", "response": response.text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send test message: {str(e)}")

# ============== WHATSAPP LOGS ROUTES ==============

@api_router.get("/whatsapp-logs")
async def get_whatsapp_logs(
    skip: int = 0,
    limit: int = 50,
    message_type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get WhatsApp logs with filtering and pagination"""
    query = {}
    
    if message_type:
        query['message_type'] = message_type
    
    if status:
        query['status'] = status
    
    if search:
        query['$or'] = [
            {'recipient_phone': {'$regex': search, '$options': 'i'}},
            {'recipient_name': {'$regex': search, '$options': 'i'}},
            {'message_preview': {'$regex': search, '$options': 'i'}}
        ]
    
    total = await db.whatsapp_logs.count_documents(query)
    
    logs = await db.whatsapp_logs.find(query, {'_id': 0}).sort('created_at', -1).skip(skip).limit(limit).to_list(limit)
    
    # Parse dates
    for log in logs:
        if isinstance(log.get('created_at'), str):
            log['created_at'] = datetime.fromisoformat(log['created_at'].replace('Z', '+00:00'))
    
    return {
        'logs': logs,
        'total': total,
        'skip': skip,
        'limit': limit
    }

@api_router.get("/whatsapp-logs/stats")
async def get_whatsapp_logs_stats(current_user: dict = Depends(get_current_user)):
    """Get WhatsApp logs statistics"""
    total = await db.whatsapp_logs.count_documents({})
    success = await db.whatsapp_logs.count_documents({'status': 'success'})
    failed = await db.whatsapp_logs.count_documents({'status': 'failed'})
    
    # Get counts by message type
    pipeline = [
        {'$group': {'_id': '$message_type', 'count': {'$sum': 1}}},
        {'$sort': {'count': -1}}
    ]
    type_counts = await db.whatsapp_logs.aggregate(pipeline).to_list(100)
    
    return {
        'total': total,
        'success': success,
        'failed': failed,
        'by_type': {item['_id']: item['count'] for item in type_counts}
    }

@api_router.delete("/whatsapp-logs/{log_id}")
async def delete_whatsapp_log(log_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a WhatsApp log entry"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can delete logs")
    
    result = await db.whatsapp_logs.delete_one({'id': log_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Log not found")
    
    return {"message": "Log deleted successfully"}

@api_router.delete("/whatsapp-logs")
async def clear_whatsapp_logs(current_user: dict = Depends(get_current_user)):
    """Clear all WhatsApp logs (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can clear logs")
    
    result = await db.whatsapp_logs.delete_many({})
    return {"message": f"Deleted {result.deleted_count} logs"}


# ============== ADMIN PROFILE ROUTES ==============

class AdminProfileUpdate(BaseModel):
    name: str
    email: str

class AdminPasswordChange(BaseModel):
    current_password: str
    new_password: str

@api_router.put("/admin/profile")
async def update_admin_profile(profile_data: AdminProfileUpdate, current_user: dict = Depends(get_current_user)):
    """Update admin profile (name and email)"""
    # Check if email is already taken by another user
    existing = await db.users.find_one({
        'email': profile_data.email.lower(),
        'id': {'$ne': current_user['id']}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Email already in use by another user")
    
    await db.users.update_one(
        {'id': current_user['id']},
        {'$set': {
            'name': profile_data.name.strip(),
            'email': profile_data.email.lower().strip(),
            'updated_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Profile updated successfully"}

@api_router.put("/admin/change-password")
async def change_admin_password(password_data: AdminPasswordChange, current_user: dict = Depends(get_current_user)):
    """Change admin password"""
    # Get current user with password
    user = await db.users.find_one({'id': current_user['id']})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify current password
    if not bcrypt.checkpw(password_data.current_password.encode(), user['password'].encode()):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Hash new password
    new_password_hash = bcrypt.hashpw(password_data.new_password.encode(), bcrypt.gensalt()).decode()
    
    await db.users.update_one(
        {'id': current_user['id']},
        {'$set': {
            'password': new_password_hash,
            'password_changed_at': datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Password changed successfully"}


# ============== DATABASE BACKUP ROUTES ==============

# Global variable for backup scheduler
backup_scheduler_task = None

@api_router.get("/database/backup-settings")
async def get_backup_settings(current_user: dict = Depends(get_current_user)):
    """Get database backup settings"""
    settings = await db.system_settings.find_one({'type': 'backup_settings'}, {'_id': 0})
    if not settings:
        # Default settings
        settings = {
            'auto_backup_enabled': True,
            'backup_times': ['09:00', '17:00'],
            'whatsapp_number': '9486544884',
            'email_address': 'vetmech2server@gmail.com'
        }
    return settings

@api_router.put("/database/backup-settings")
async def update_backup_settings(settings: dict, current_user: dict = Depends(get_current_user)):
    """Update database backup settings"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can modify backup settings")
    
    settings['type'] = 'backup_settings'
    settings['updated_at'] = datetime.now(timezone.utc).isoformat()
    settings['updated_by'] = current_user['name']
    
    await db.system_settings.update_one(
        {'type': 'backup_settings'},
        {'$set': settings},
        upsert=True
    )
    
    return {"message": "Backup settings updated successfully"}

@api_router.get("/database/backup-history")
async def get_backup_history(current_user: dict = Depends(get_current_user)):
    """Get database backup history"""
    backups = await db.backup_history.find({}, {'_id': 0}).sort('created_at', -1).limit(20).to_list(20)
    return {"backups": backups}

@api_router.get("/database/export")
async def export_database(current_user: dict = Depends(get_current_user)):
    """Export entire database as JSON"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can export database")
    
    # Get all collections
    collections_to_export = [
        'doctors', 'medicals', 'agencies', 'items', 'orders', 'expenses',
        'reminders', 'pending_items', 'portal_customers', 'support_tickets',
        'users', 'company_settings', 'item_categories', 'transports',
        'email_logs', 'whatsapp_logs', 'marketing_campaigns'
    ]
    
    export_data = {
        'export_date': datetime.now(timezone.utc).isoformat(),
        'exported_by': current_user['name'],
        'collections': {}
    }
    
    for collection_name in collections_to_export:
        collection = db[collection_name]
        documents = await collection.find({}, {'_id': 0}).to_list(100000)
        export_data['collections'][collection_name] = documents
    
    # Create JSON response
    json_str = json.dumps(export_data, default=str, indent=2)
    
    # Log the backup
    await db.backup_history.insert_one({
        'id': str(uuid.uuid4()),
        'filename': f"vmp_crm_backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json",
        'status': 'success',
        'type': 'manual_download',
        'created_at': datetime.now(timezone.utc).isoformat(),
        'created_by': current_user['name'],
        'size_bytes': len(json_str)
    })
    
    return Response(
        content=json_str,
        media_type="application/json",
        headers={
            "Content-Disposition": f"attachment; filename=vmp_crm_backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
        }
    )

@api_router.post("/database/trigger-backup")
async def trigger_backup(background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """Manually trigger a backup and send via WhatsApp and Email"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can trigger backups")
    
    background_tasks.add_task(perform_scheduled_backup, current_user['name'])
    return {"message": "Backup triggered. You will receive it shortly via WhatsApp and Email."}

async def perform_scheduled_backup(triggered_by: str = "System"):
    """Perform database backup and send via WhatsApp and Email"""
    try:
        # Get backup settings
        settings = await db.system_settings.find_one({'type': 'backup_settings'}, {'_id': 0})
        if not settings:
            settings = {
                'whatsapp_number': '9486544884',
                'email_address': 'vetmech2server@gmail.com'
            }
        
        # Export database
        collections_to_export = [
            'doctors', 'medicals', 'agencies', 'items', 'orders', 'expenses',
            'reminders', 'pending_items', 'portal_customers', 'support_tickets',
            'users', 'company_settings', 'item_categories', 'transports'
        ]
        
        export_data = {
            'export_date': datetime.now(timezone.utc).isoformat(),
            'exported_by': triggered_by,
            'collections': {}
        }
        
        total_records = 0
        for collection_name in collections_to_export:
            collection = db[collection_name]
            documents = await collection.find({}, {'_id': 0}).to_list(100000)
            export_data['collections'][collection_name] = documents
            total_records += len(documents)
        
        json_str = json.dumps(export_data, default=str)
        file_size = len(json_str)
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        filename = f"vmp_crm_backup_{timestamp}.json"
        
        sent_whatsapp = False
        sent_email = False
        
        # Send WhatsApp notification
        wa_number = settings.get('whatsapp_number', '9486544884')
        if wa_number:
            config = await get_whatsapp_config()
            if config.get('api_url') and config.get('auth_token'):
                try:
                    message = f"*VMP CRM Database Backup*\n\n" \
                              f"Backup completed successfully!\n\n" \
                              f"*Timestamp:* {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC\n" \
                              f"*Triggered by:* {triggered_by}\n" \
                              f"*Total Records:* {total_records}\n" \
                              f"*File Size:* {file_size / 1024:.1f} KB\n\n" \
                              f"The backup has been sent to your email.\n\n" \
                              f"Filename: {filename}"
                    
                    wa_mobile = wa_number if wa_number.startswith('91') else f"91{wa_number[-10:]}"
                    params = {
                        'action': 'send',
                        'senderId': config['sender_id'],
                        'authToken': config['auth_token'],
                        'messageText': message,
                        'receiverId': wa_mobile
                    }
                    
                    async with httpx.AsyncClient(timeout=30.0) as http_client:
                        response = await http_client.get(config['api_url'], params=params)
                        if response.status_code == 200:
                            sent_whatsapp = True
                            logger.info(f"Backup notification sent via WhatsApp to {wa_mobile}")
                except Exception as e:
                    logger.error(f"Failed to send WhatsApp backup notification: {str(e)}")
        
        # Send Email with backup attachment
        email_address = settings.get('email_address', 'vetmech2server@gmail.com')
        if email_address:
            smtp_settings = await db.smtp_settings.find_one({}, {'_id': 0})
            if smtp_settings and smtp_settings.get('host'):
                try:
                    import smtplib
                    from email.mime.multipart import MIMEMultipart
                    from email.mime.text import MIMEText
                    from email.mime.application import MIMEApplication
                    
                    msg = MIMEMultipart()
                    msg['From'] = smtp_settings.get('from_email', smtp_settings['username'])
                    msg['To'] = email_address
                    msg['Subject'] = f"VMP CRM Database Backup - {timestamp}"
                    
                    body = f"""
VMP CRM Database Backup

Backup Details:
- Timestamp: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC
- Triggered by: {triggered_by}
- Total Records: {total_records}
- File Size: {file_size / 1024:.1f} KB

The backup file is attached to this email.

This is an automated backup from VMP CRM system.
                    """
                    
                    msg.attach(MIMEText(body, 'plain'))
                    
                    # Attach backup file
                    attachment = MIMEApplication(json_str.encode(), _subtype='json')
                    attachment.add_header('Content-Disposition', 'attachment', filename=filename)
                    msg.attach(attachment)
                    
                    # Send email
                    server = smtplib.SMTP(smtp_settings['host'], smtp_settings.get('port', 587))
                    server.starttls()
                    server.login(smtp_settings['username'], smtp_settings['password'])
                    server.send_message(msg)
                    server.quit()
                    
                    sent_email = True
                    logger.info(f"Backup email sent to {email_address}")
                except Exception as e:
                    logger.error(f"Failed to send backup email: {str(e)}")
        
        # Log the backup
        await db.backup_history.insert_one({
            'id': str(uuid.uuid4()),
            'filename': filename,
            'status': 'success',
            'type': 'scheduled' if triggered_by == 'System' else 'manual',
            'created_at': datetime.now(timezone.utc).isoformat(),
            'created_by': triggered_by,
            'size_bytes': file_size,
            'total_records': total_records,
            'sent_whatsapp': sent_whatsapp,
            'sent_email': sent_email
        })
        
        logger.info(f"Database backup completed: {filename}")
        
    except Exception as e:
        logger.error(f"Database backup failed: {str(e)}")
        # Log failure
        await db.backup_history.insert_one({
            'id': str(uuid.uuid4()),
            'filename': f"backup_failed_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}",
            'status': 'failed',
            'error': str(e),
            'created_at': datetime.now(timezone.utc).isoformat(),
            'created_by': triggered_by
        })

async def run_scheduled_backups():
    """Background task to run scheduled backups at 9 AM and 5 PM"""
    while True:
        try:
            # Get settings
            settings = await db.system_settings.find_one({'type': 'backup_settings'}, {'_id': 0})
            if not settings or not settings.get('auto_backup_enabled', True):
                await asyncio.sleep(300)  # Check again in 5 minutes
                continue
            
            backup_times = settings.get('backup_times', ['09:00', '17:00'])
            
            # Get current time in IST (UTC+5:30)
            now_utc = datetime.now(timezone.utc)
            ist_offset = timedelta(hours=5, minutes=30)
            now_ist = now_utc + ist_offset
            current_time = now_ist.strftime('%H:%M')
            
            # Check if current time matches any backup time (within 1 minute window)
            for backup_time in backup_times:
                if current_time == backup_time:
                    logger.info(f"Running scheduled backup at {current_time} IST")
                    await perform_scheduled_backup("System (Scheduled)")
                    await asyncio.sleep(120)  # Wait 2 minutes to avoid duplicate runs
                    break
            
            await asyncio.sleep(60)  # Check every minute
            
        except asyncio.CancelledError:
            logger.info("Backup scheduler task cancelled")
            break
        except Exception as e:
            logger.error(f"Error in backup scheduler: {str(e)}")
            await asyncio.sleep(300)


# ============== USER MANAGEMENT ROUTES ==============

@api_router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    """Get all users (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can view users")
    
    users = await db.users.find({}, {'_id': 0, 'password': 0}).to_list(100)
    
    # Parse dates
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'].replace('Z', '+00:00'))
    
    return users

@api_router.post("/users")
async def create_user(user_data: UserCreateByAdmin, current_user: dict = Depends(get_current_user)):
    """Create a new user (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can create users")
    
    # Check if email already exists
    existing = await db.users.find_one({'email': user_data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Default permissions for staff
    default_permissions = UserPermissions()
    permissions = user_data.permissions or default_permissions
    
    # Admin gets all permissions by default
    if user_data.role == 'admin':
        permissions = UserPermissions(
            doctors=True, medicals=True, agencies=True, items=True, orders=True,
            expenses=True, reminders=True, pending_items=True, email_logs=True,
            whatsapp_logs=True, users=True, smtp_settings=True, company_settings=True,
            whatsapp_settings=True
        )
    
    user_doc = {
        'id': user_id,
        'email': user_data.email.lower(),
        'password': hash_password(user_data.password),
        'name': user_data.name,
        'role': user_data.role,
        'permissions': permissions.model_dump(),
        'created_at': now.isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    return {
        'id': user_id,
        'email': user_data.email.lower(),
        'name': user_data.name,
        'role': user_data.role,
        'permissions': permissions.model_dump(),
        'created_at': now
    }

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, user_data: UserUpdateByAdmin, current_user: dict = Depends(get_current_user)):
    """Update a user (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can update users")
    
    # Get existing user
    existing = await db.users.find_one({'id': user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_doc = {}
    
    if user_data.email:
        # Check if new email already exists for another user
        email_exists = await db.users.find_one({'email': user_data.email.lower(), 'id': {'$ne': user_id}})
        if email_exists:
            raise HTTPException(status_code=400, detail="Email already in use")
        update_doc['email'] = user_data.email.lower()
    
    if user_data.name:
        update_doc['name'] = user_data.name
    
    if user_data.role:
        update_doc['role'] = user_data.role
    
    if user_data.password:
        update_doc['password'] = hash_password(user_data.password)
    
    if user_data.permissions:
        update_doc['permissions'] = user_data.permissions.model_dump()
    
    if update_doc:
        await db.users.update_one({'id': user_id}, {'$set': update_doc})
    
    # Return updated user
    updated = await db.users.find_one({'id': user_id}, {'_id': 0, 'password': 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'].replace('Z', '+00:00'))
    
    return updated

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a user (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    # Prevent deleting yourself
    if user_id == current_user['id']:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({'id': user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

@api_router.get("/users/{user_id}")
async def get_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """Get a specific user (admin only)"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can view user details")
    
    user = await db.users.find_one({'id': user_id}, {'_id': 0, 'password': 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if isinstance(user.get('created_at'), str):
        user['created_at'] = datetime.fromisoformat(user['created_at'].replace('Z', '+00:00'))
    
    return user

# ============== HEALTH CHECK ==============

@api_router.get("/")
async def root():
    return {"message": "VMP CRM API is running"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    """Start background tasks on app startup"""
    global daily_reminder_task, backup_scheduler_task, greeting_task
    daily_reminder_task = asyncio.create_task(send_daily_reminder_summary())
    logger.info("Daily reminder background task started")
    backup_scheduler_task = asyncio.create_task(run_scheduled_backups())
    logger.info("Backup scheduler task started")
    greeting_task = asyncio.create_task(send_birthday_anniversary_greetings())
    logger.info("Birthday/Anniversary greeting task started")
    # Seed default templates
    await seed_default_greeting_templates()

@app.on_event("shutdown")
async def shutdown_db_client():
    global daily_reminder_task, backup_scheduler_task, greeting_task
    if daily_reminder_task:
        daily_reminder_task.cancel()
        try:
            await daily_reminder_task
        except asyncio.CancelledError:
            pass
        logger.info("Daily reminder background task stopped")
    if backup_scheduler_task:
        backup_scheduler_task.cancel()
        try:
            await backup_scheduler_task
        except asyncio.CancelledError:
            pass
        logger.info("Backup scheduler task stopped")
    if greeting_task:
        greeting_task.cancel()
        try:
            await greeting_task
        except asyncio.CancelledError:
            pass
        logger.info("Birthday/Anniversary greeting task stopped")
    client.close()
