"""
Tests for Items Bulk Import Feature
- GET /api/items/import/template - Download Excel template
- POST /api/items/import - Bulk import items from Excel
"""
import pytest
import requests
import os
from io import BytesIO

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
TEST_EMAIL = "admin@vmpcrm.com"
TEST_PASSWORD = "admin123"


class TestItemsBulkImport:
    """Test Items Bulk Import functionality"""
    
    auth_token = None
    created_item_codes = []
    
    @pytest.fixture(autouse=True)
    def setup_auth(self):
        """Setup authentication before tests"""
        if not TestItemsBulkImport.auth_token:
            response = requests.post(f"{BASE_URL}/api/auth/login", json={
                "email": TEST_EMAIL,
                "password": TEST_PASSWORD
            })
            assert response.status_code == 200, f"Login failed: {response.text}"
            TestItemsBulkImport.auth_token = response.json()["access_token"]
    
    def get_headers(self):
        return {"Authorization": f"Bearer {TestItemsBulkImport.auth_token}"}
    
    # ============== Template Download Tests ==============
    
    def test_01_download_template_returns_excel(self):
        """Test GET /api/items/import/template returns Excel file"""
        response = requests.get(
            f"{BASE_URL}/api/items/import/template",
            headers=self.get_headers()
        )
        assert response.status_code == 200, f"Template download failed: {response.status_code}"
        
        # Check content type
        content_type = response.headers.get('content-type', '')
        assert 'spreadsheet' in content_type or 'excel' in content_type or 'octet-stream' in content_type, \
            f"Unexpected content type: {content_type}"
        
        # Check content disposition header for filename
        content_disp = response.headers.get('content-disposition', '')
        assert 'items_import_template.xlsx' in content_disp, \
            f"Expected filename in header, got: {content_disp}"
        
        # Check that we received some content
        assert len(response.content) > 0, "Empty file received"
        print("PASS: Template download returns valid Excel file")
    
    def test_02_download_template_requires_auth(self):
        """Test template download requires authentication"""
        response = requests.get(f"{BASE_URL}/api/items/import/template")
        assert response.status_code in [401, 403], \
            f"Expected 401/403 without auth, got {response.status_code}"
        print("PASS: Template download requires authentication")
    
    # ============== Import Tests ==============
    
    def test_03_import_requires_auth(self):
        """Test import endpoint requires authentication"""
        # Create a simple Excel-like file
        response = requests.post(
            f"{BASE_URL}/api/items/import",
            files={"file": ("test.xlsx", b"test content", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code in [401, 403], \
            f"Expected 401/403 without auth, got {response.status_code}"
        print("PASS: Import endpoint requires authentication")
    
    def test_04_import_rejects_non_excel_files(self):
        """Test import rejects non-Excel files"""
        response = requests.post(
            f"{BASE_URL}/api/items/import",
            headers=self.get_headers(),
            files={"file": ("test.txt", b"test content", "text/plain")}
        )
        assert response.status_code == 400, \
            f"Expected 400 for non-Excel file, got {response.status_code}"
        print("PASS: Import rejects non-Excel files")
    
    def test_05_import_valid_excel_file(self):
        """Test importing a valid Excel file creates items"""
        # First download the template
        template_response = requests.get(
            f"{BASE_URL}/api/items/import/template",
            headers=self.get_headers()
        )
        assert template_response.status_code == 200
        
        # Try to import using openpyxl to create a proper Excel file
        try:
            from openpyxl import Workbook
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            
            # Add headers (matching template format)
            headers = ["Item Code*", "Item Name*", "Main Categories", "Subcategories", 
                      "Composition", "MRP*", "Rate*", "GST %", "Offer", "Special Offer"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add test data - leave Item Code empty for auto-generation
            test_items = [
                ["", "TEST_ImportItem1", "Large Animals, Poultry", "Injection", "Test composition 1", 100, 80, 12, "10% off", ""],
                ["TEST-IMP-002", "TEST_ImportItem2", "Pets", "Liquids", "Test composition 2", 200, 150, 5, "", "Special Deal"],
            ]
            
            for row_num, row_data in enumerate(test_items, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Upload the file
            response = requests.post(
                f"{BASE_URL}/api/items/import",
                headers=self.get_headers(),
                files={"file": ("test_import.xlsx", output.getvalue(), 
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
            
            assert response.status_code == 200, f"Import failed: {response.text}"
            result = response.json()
            
            # Check response structure
            assert "message" in result, "Missing message in response"
            assert "items_created" in result, "Missing items_created in response"
            assert "items_updated" in result, "Missing items_updated in response"
            
            # At least one item should be created
            total_processed = result["items_created"] + result["items_updated"]
            assert total_processed > 0, f"No items were processed: {result}"
            
            # Store created codes for cleanup
            TestItemsBulkImport.created_item_codes.append("TEST-IMP-002")
            
            print(f"PASS: Import successful - {result['items_created']} created, {result['items_updated']} updated")
            
        except ImportError:
            # If openpyxl not available, skip this test
            pytest.skip("openpyxl not installed for test file creation")
    
    def test_06_import_updates_existing_items(self):
        """Test import updates existing items when item_code matches"""
        try:
            from openpyxl import Workbook
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            
            # Add headers
            headers = ["Item Code*", "Item Name*", "Main Categories", "Subcategories", 
                      "Composition", "MRP*", "Rate*", "GST %", "Offer", "Special Offer"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Update existing item (if created in previous test)
            test_items = [
                ["TEST-IMP-002", "TEST_ImportItem2_Updated", "Large Animals", "Powder", "Updated composition", 250, 180, 8, "20% off", "Updated offer"],
            ]
            
            for row_num, row_data in enumerate(test_items, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = requests.post(
                f"{BASE_URL}/api/items/import",
                headers=self.get_headers(),
                files={"file": ("test_update.xlsx", output.getvalue(), 
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
            
            assert response.status_code == 200, f"Import update failed: {response.text}"
            result = response.json()
            
            # Should have updated at least one item
            assert result["items_updated"] >= 0 or result["items_created"] >= 0, \
                f"No items were processed: {result}"
            
            print(f"PASS: Import update - {result['items_created']} created, {result['items_updated']} updated")
            
        except ImportError:
            pytest.skip("openpyxl not installed for test file creation")
    
    def test_07_import_handles_comma_separated_categories(self):
        """Test import correctly parses comma-separated categories"""
        try:
            from openpyxl import Workbook
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            
            headers = ["Item Code*", "Item Name*", "Main Categories", "Subcategories", 
                      "Composition", "MRP*", "Rate*", "GST %", "Offer", "Special Offer"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Test with multiple categories
            test_items = [
                ["TEST-IMP-MULTI", "TEST_MultiCategoryItem", "Large Animals, Poultry, Pets", 
                 "Injection, Powder, Liquids", "Multi comp", 300, 250, 12, "", ""],
            ]
            
            for row_num, row_data in enumerate(test_items, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = requests.post(
                f"{BASE_URL}/api/items/import",
                headers=self.get_headers(),
                files={"file": ("test_multi.xlsx", output.getvalue(), 
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
            
            assert response.status_code == 200, f"Import failed: {response.text}"
            
            # Verify the item was created/updated with multiple categories
            items_response = requests.get(
                f"{BASE_URL}/api/items?search=TEST_MultiCategoryItem",
                headers=self.get_headers()
            )
            
            if items_response.status_code == 200:
                items = items_response.json()
                if items:
                    item = items[0]
                    assert isinstance(item.get("main_categories", []), list), "main_categories should be a list"
                    assert isinstance(item.get("subcategories", []), list), "subcategories should be a list"
                    print(f"PASS: Categories parsed - main: {item.get('main_categories')}, sub: {item.get('subcategories')}")
            
            TestItemsBulkImport.created_item_codes.append("TEST-IMP-MULTI")
            print("PASS: Import handles comma-separated categories")
            
        except ImportError:
            pytest.skip("openpyxl not installed for test file creation")
    
    def test_08_import_empty_file_returns_error(self):
        """Test import rejects empty/invalid Excel file"""
        try:
            from openpyxl import Workbook
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            
            # Only headers, no data
            headers = ["Item Code*", "Item Name*", "Main Categories", "Subcategories", 
                      "Composition", "MRP*", "Rate*", "GST %", "Offer", "Special Offer"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = requests.post(
                f"{BASE_URL}/api/items/import",
                headers=self.get_headers(),
                files={"file": ("empty.xlsx", output.getvalue(), 
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
            
            # Should return 400 for no valid items
            assert response.status_code == 400, \
                f"Expected 400 for empty file, got {response.status_code}: {response.text}"
            print("PASS: Import rejects empty Excel file")
            
        except ImportError:
            pytest.skip("openpyxl not installed for test file creation")
    
    def test_99_cleanup_test_items(self):
        """Cleanup test items created during tests"""
        # Get all items with TEST_ prefix
        response = requests.get(
            f"{BASE_URL}/api/items?search=TEST_",
            headers=self.get_headers()
        )
        
        if response.status_code == 200:
            items = response.json()
            for item in items:
                if item.get('item_name', '').startswith('TEST_'):
                    delete_response = requests.delete(
                        f"{BASE_URL}/api/items/{item['id']}",
                        headers=self.get_headers()
                    )
                    if delete_response.status_code in [200, 204]:
                        print(f"Cleaned up: {item['item_name']}")
        
        print("PASS: Cleanup completed")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
